# -*- coding: utf-8 -*-
"""
Enhanced Stock Options Analysis v2
Fixes all bugs in v1 and adds Greeks, full chain, vol skew, parity check.
"""

import os
import warnings
import numpy as np
from scipy.stats import norm
from scipy.optimize import brentq
from datetime import datetime
import yfinance as yf
import pandas as pd
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

warnings.filterwarnings('ignore')


# ─── Black-Scholes Core ────────────────────────────────────────────────────────

def bs_price(S, K, T, r, sigma, option_type='call'):
    """Black-Scholes price. Returns np.nan for degenerate inputs."""
    if T <= 0 or sigma <= 0 or S <= 0 or K <= 0:
        return np.nan
    d1 = (np.log(S / K) + (r + 0.5 * sigma**2) * T) / (sigma * np.sqrt(T))
    d2 = d1 - sigma * np.sqrt(T)
    if option_type == 'call':
        return S * norm.cdf(d1) - K * np.exp(-r * T) * norm.cdf(d2)
    return K * np.exp(-r * T) * norm.cdf(-d2) - S * norm.cdf(-d1)


def bs_greeks(S, K, T, r, sigma, option_type='call'):
    """
    All Black-Scholes Greeks.
    Theta is per calendar day. Vega is per 1% change in vol. Rho per 1%.
    """
    if T <= 0 or sigma <= 0 or S <= 0 or K <= 0:
        return {g: np.nan for g in ['delta', 'gamma', 'theta', 'vega', 'rho']}

    d1 = (np.log(S / K) + (r + 0.5 * sigma**2) * T) / (sigma * np.sqrt(T))
    d2 = d1 - sigma * np.sqrt(T)
    nd1 = norm.pdf(d1)

    gamma = nd1 / (S * sigma * np.sqrt(T))
    vega  = S * np.sqrt(T) * nd1 / 100       # per 1% vol move

    if option_type == 'call':
        delta = norm.cdf(d1)
        theta = (-(S * nd1 * sigma) / (2 * np.sqrt(T))
                 - r * K * np.exp(-r * T) * norm.cdf(d2)) / 365
        rho   = K * T * np.exp(-r * T) * norm.cdf(d2) / 100
    else:
        delta = norm.cdf(d1) - 1
        theta = (-(S * nd1 * sigma) / (2 * np.sqrt(T))
                 + r * K * np.exp(-r * T) * norm.cdf(-d2)) / 365
        rho   = -K * T * np.exp(-r * T) * norm.cdf(-d2) / 100

    return {'delta': delta, 'gamma': gamma, 'theta': theta, 'vega': vega, 'rho': rho}


# ─── Implied Volatility ────────────────────────────────────────────────────────

def implied_volatility(S, K, T, r, market_price, option_type='call'):
    """
    Robust IV using Brent's method (scipy.optimize.brentq).

    Fix over v1: Newton-Raphson crashed when vega≈0 (deep OTM, short T) and
    had no sigma bounds, producing NaN/inf silently. Brent's method is
    bracket-safe, guaranteed to converge, and needs no vega computation.
    """
    if T <= 0 or market_price <= 0 or S <= 0 or K <= 0:
        return np.nan

    intrinsic = max(0, S - K) if option_type == 'call' else max(0, K - S)
    if market_price < intrinsic - 0.01:   # below intrinsic → bad data
        return np.nan

    def objective(sigma):
        return bs_price(S, K, T, r, sigma, option_type) - market_price

    try:
        return brentq(objective, 1e-6, 10.0, xtol=1e-6, maxiter=500)
    except (ValueError, RuntimeError):
        return np.nan


# ─── Historical Volatility ─────────────────────────────────────────────────────

def historical_volatility(symbol, period='1y'):
    """Annualised close-to-close HV over the last trading year."""
    hist = yf.Ticker(symbol).history(period=period)
    if len(hist) < 10:
        return np.nan
    log_ret = np.log(hist['Close'] / hist['Close'].shift(1)).dropna()
    return log_ret.std() * np.sqrt(252)


# ─── Risk-Free Rate ────────────────────────────────────────────────────────────

def get_risk_free_rate():
    """
    13-week T-bill yield from ^IRX via yfinance.

    Fix over v1: pandas_datareader FRED requires a separate API key and
    frequently breaks. ^IRX is available directly through yfinance.
    """
    try:
        irx = yf.Ticker('^IRX').history(period='5d')['Close']
        return irx.iloc[-1] / 100
    except Exception:
        print("  [Warning] Could not fetch ^IRX; defaulting to 5.00%")
        return 0.05


# ─── Option Chain Helpers ──────────────────────────────────────────────────────

def mid_price(row):
    """
    Bid/ask midpoint when available, else lastPrice.

    Fix over v1: lastPrice on illiquid options can be hours/days stale.
    """
    bid = row.get('bid', 0) or 0
    ask = row.get('ask', 0) or 0
    if bid > 0 and ask > 0:
        return (bid + ask) / 2.0
    return row['lastPrice']


def classify_moneyness(S, K, option_type):
    ratio = S / K
    if option_type == 'call':
        if ratio > 1.02: return 'ITM'
        if ratio < 0.98: return 'OTM'
    else:
        if ratio < 0.98: return 'ITM'
        if ratio > 1.02: return 'OTM'
    return 'ATM'


def prob_expire_itm(S, K, T, r, sigma, option_type='call'):
    """Risk-neutral P(expiry ITM): N(d2) for calls, N(-d2) for puts."""
    if T <= 0 or sigma <= 0:
        return np.nan
    d2 = (np.log(S / K) + (r - 0.5 * sigma**2) * T) / (sigma * np.sqrt(T))
    return norm.cdf(d2) if option_type == 'call' else norm.cdf(-d2)


def expected_move(S, iv, T):
    """1-sigma expected dollar move at expiry: S × IV × √T."""
    if np.isnan(iv) or iv <= 0 or T <= 0:
        return np.nan
    return S * iv * np.sqrt(T)


# ─── Full Chain Analysis ───────────────────────────────────────────────────────

def analyze_chain(symbol, expiry_str, S, r, hv):
    """
    Analyse every strike (calls and puts) for one expiry date.

    Fix over v1: v1 only read iloc[0] — the first call at the lowest strike.
    This function iterates the entire option chain.
    """
    chain = yf.Ticker(symbol).option_chain(expiry_str)
    trade_dt = datetime.now().replace(tzinfo=None)
    expiry_dt = datetime.strptime(expiry_str, '%Y-%m-%d')
    T = max((expiry_dt - trade_dt).days / 365.0, 1 / 365.0)

    rows = []
    for opt_type, df in [('call', chain.calls), ('put', chain.puts)]:
        for _, row in df.iterrows():
            K     = row['strike']
            price = mid_price(row)
            if price <= 0:
                continue

            iv = implied_volatility(S, K, T, r, price, opt_type)
            sigma_for_greeks = iv if not np.isnan(iv) else hv
            g = bs_greeks(S, K, T, r, sigma_for_greeks, opt_type)

            intrinsic = max(0, S - K) if opt_type == 'call' else max(0, K - S)
            time_val  = max(0, price - intrinsic)
            p_itm     = prob_expire_itm(S, K, T, r, sigma_for_greeks, opt_type)
            exp_mv    = expected_move(S, sigma_for_greeks, T)
            iv_hv     = (iv - hv) * 100 if not np.isnan(iv) else np.nan

            rows.append({
                'Expiry':            expiry_str,
                'Type':              opt_type.upper(),
                'Strike':            K,
                'Moneyness':         classify_moneyness(S, K, opt_type),
                'Mid Price':         round(price, 4),
                'Bid':               row.get('bid', np.nan),
                'Ask':               row.get('ask', np.nan),
                'Last Price':        row['lastPrice'],
                'Volume':            row.get('volume', 0) or 0,
                'Open Interest':     row.get('openInterest', 0) or 0,
                'IV (%)':            round(iv * 100, 2) if not np.isnan(iv) else np.nan,
                'HV (%)':            round(hv * 100, 2),
                'IV-HV Spread (%)':  round(iv_hv, 2) if not np.isnan(iv_hv) else np.nan,
                'Intrinsic Value':   round(intrinsic, 4),
                'Time Value':        round(time_val, 4),
                'Delta':             round(g['delta'], 4),
                'Gamma':             round(g['gamma'], 6),
                'Theta ($/day)':     round(g['theta'], 4),
                'Vega ($/1%vol)':    round(g['vega'], 4),
                'Rho ($/1%)':        round(g['rho'], 4),
                'Prob ITM (%)':      round(p_itm * 100, 2) if not np.isnan(p_itm) else np.nan,
                'Expected Move ($)': round(exp_mv, 2) if not np.isnan(exp_mv) else np.nan,
                'T (years)':         round(T, 4),
            })

    return pd.DataFrame(rows)


# ─── Volatility Skew ───────────────────────────────────────────────────────────

def vol_skew(df):
    """IV by strike for calls and puts — visualises the vol smile/skew."""
    calls = (df[df['Type'] == 'CALL'][['Strike', 'IV (%)']]
             .rename(columns={'IV (%)': 'Call IV (%)'})
             .set_index('Strike'))
    puts  = (df[df['Type'] == 'PUT'][['Strike', 'IV (%)']]
             .rename(columns={'IV (%)': 'Put IV (%)'})
             .set_index('Strike'))
    return calls.join(puts, how='outer').reset_index().sort_values('Strike')


# ─── Put-Call Parity Check ─────────────────────────────────────────────────────

def parity_check(df, S, r):
    """
    C - P = S - K·e^(−rT)
    Large errors indicate stale prices or data anomalies worth filtering.
    """
    calls = (df[df['Type'] == 'CALL'][['Strike', 'Mid Price', 'T (years)']]
             .rename(columns={'Mid Price': 'C'}))
    puts  = (df[df['Type'] == 'PUT'][['Strike', 'Mid Price', 'T (years)']]
             .rename(columns={'Mid Price': 'P'}))
    m = pd.merge(calls, puts, on=['Strike', 'T (years)'])
    m['C-P']           = m['C'] - m['P']
    m['S-PV(K)']       = S - m['Strike'] * np.exp(-r * m['T (years)'])
    m['Parity Err ($)'] = (m['C-P'] - m['S-PV(K)']).abs().round(4)
    return m[['Strike', 'C', 'P', 'C-P', 'S-PV(K)', 'Parity Err ($)']].sort_values('Strike')


# ─── Excel Export ──────────────────────────────────────────────────────────────

def _autofit(ws):
    """Auto-size columns and bold headers. Fixes v1 len(numeric) crash."""
    for col in ws.columns:
        max_len = max(
            (len(str(c.value)) for c in col if c.value is not None),
            default=8
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 3, 28)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')


def export_excel(symbol, S, r, hv, all_data, skew_df, parity_df, filename):
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:

        # Summary
        summary = pd.DataFrame([{
            'Symbol':                 symbol,
            'Spot Price ($)':         round(S, 2),
            'Risk-Free Rate (%)':     round(r * 100, 3),
            'Historical Vol (%)':     round(hv * 100, 2),
            'Analysis Date':          datetime.now().strftime('%Y-%m-%d %H:%M'),
            'Total Options Analyzed': len(all_data),
            'Expiries Analyzed':      all_data['Expiry'].nunique(),
        }])
        summary.T.to_excel(writer, sheet_name='Summary', header=False)

        all_data.to_excel(writer, index=False, sheet_name='Full Chain')
        all_data[all_data['Type'] == 'CALL'].to_excel(writer, index=False, sheet_name='Calls')
        all_data[all_data['Type'] == 'PUT'].to_excel(writer, index=False, sheet_name='Puts')
        skew_df.to_excel(writer, index=False, sheet_name='Vol Skew')
        parity_df.to_excel(writer, index=False, sheet_name='Parity Check')

        for ws in writer.sheets.values():
            _autofit(ws)


# ─── Main ──────────────────────────────────────────────────────────────────────

def main():
    symbol = input("Enter stock symbol: ").strip().upper()

    print(f"\n[1/5] Fetching spot price for {symbol}...")
    hist = yf.Ticker(symbol).history(period='2d')
    if hist.empty:
        print(f"Error: no data for {symbol}")
        return
    S = hist['Close'].iloc[-1]
    print(f"  Spot: ${S:.2f}")

    print("[2/5] Fetching risk-free rate (^IRX)...")
    r = get_risk_free_rate()
    print(f"  Rate: {r * 100:.3f}%")

    print("[3/5] Computing 1-year historical volatility...")
    hv = historical_volatility(symbol)
    print(f"  HV:   {hv * 100:.2f}%")

    print("[4/5] Fetching option expiry dates...")
    expirations = yf.Ticker(symbol).options
    if not expirations:
        print("No options available.")
        return
    expirations = expirations[:8]   # first 8 expiries is plenty for one run
    print(f"  Analysing {len(expirations)} expiry dates")

    print("[5/5] Analysing full option chain...")
    all_dfs = []
    for i, exp in enumerate(expirations, 1):
        try:
            print(f"  [{i}/{len(expirations)}] {exp} ...", end=' ', flush=True)
            df = analyze_chain(symbol, exp, S, r, hv)
            all_dfs.append(df)
            print(f"{len(df)} contracts")
        except Exception as e:
            print(f"skipped ({e})")

    if not all_dfs:
        print("Could not retrieve any option data.")
        return

    all_data  = pd.concat(all_dfs, ignore_index=True)
    skew_df   = vol_skew(all_dfs[0])        # nearest expiry shows the clearest skew
    parity_df = parity_check(all_dfs[0], S, r)

    num = sum(1 for f in os.listdir('.') if f.startswith(f"{symbol}_options_v2"))
    filename = f"{symbol}_options_v2_{num + 1}.xlsx"
    export_excel(symbol, S, r, hv, all_data, skew_df, parity_df, filename)
    print(f"\nSaved: {filename}")

    # Quick console summary
    atm = all_data[all_data['Moneyness'] == 'ATM']
    if not atm.empty:
        atm_iv = atm['IV (%)'].mean()
        print(f"\nATM IV:  {atm_iv:.1f}%")
        print(f"HV:      {hv * 100:.1f}%")
        print(f"IV Prem: {atm_iv - hv * 100:+.1f}%  "
              f"({'options expensive — vol sellers edge' if atm_iv > hv*100 else 'options cheap — vol buyers edge'})")

    # Optional Colab download (graceful — no hard crash if not in Colab)
    try:
        from google.colab import files
        files.download(filename)
    except ImportError:
        pass


if __name__ == '__main__':
    main()
