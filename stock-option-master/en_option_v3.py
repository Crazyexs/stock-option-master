# -*- coding: utf-8 -*-
"""
Enhanced Stock Options Analysis v3

Integrations over v2:
  1. Term-matched risk-free rate  — wallstreet idea: interpolate the Treasury
     yield curve so LEAPS use the 5-yr rate, weeklies use the 3-month rate.
  2. Dividend yield (q) in BS     — correct pricing for dividend-paying stocks.
  3. Income screener (CSP + CC)   — options_lab ARR screener, compounded annual.
  4. Strategy P&L + PoP           — optionlab (pip install optionlab) for any
     single- or multi-leg strategy evaluation; graceful fallback if not installed.
  5. Buying power + commissions   — OptionSuite (sirnfs): CBOE margin formula
     (20% rule vs 10% rule, take max) and TastyWorks-style commission model.
     Also adds delta-targeted entry and risk management exit rules.

Skipped repos:
  mirajgodha/options  → India/NSE only, wrong market.
  hfwebbed/...        → WIP prototype, no production-ready logic.
  symfony/options-resolver → PHP config library, not financial options.
  QuantConnect/Lean        → C# runtime required; Python examples are
                             not standalone. Used as reference for multi-
                             leg strategy construction patterns only.

New in this version:
  6. American option pricing  — optlib (dbrojas): Bjerksund-Stensland 2002.
     All US equity options are American-style. European BS underprices deep
     ITM puts and high-dividend calls. B-S 2002 is the industry standard
     closed-form approximation for American early exercise premium.
  7. CBOE direct data source  — OpenBB CBOE provider reveals a free public
     endpoint (cdn.cboe.com) with 15-min delayed exchange data. No API key.
     Added as a second data source alongside yfinance.
  8. Multi-leg strategy scanner — Lean reference: iron condor and calendar
     spread construction logic ported as native Python from Lean's
     IronCondorStrategyAlgorithm.py and CallCalendarSpread patterns.
"""

import os
import math
import warnings
import numpy as np
from scipy.stats import norm, mvn
from scipy.optimize import brentq
from scipy.interpolate import interp1d
from datetime import datetime, date
from collections import namedtuple
from datetime import timezone as _tz
import yfinance as yf
import pandas as pd
import requests
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

warnings.filterwarnings('ignore')


# ─── Yahoo Finance Direct Client  (fc.yahoo.com bypass) ───────────────────────
# yfinance ≥ 1.0 uses curl_cffi to hit fc.yahoo.com for cookie auth.
# That domain is unreachable in many environments (corporate VPN, ISP blocks,
# certain macOS network configs). This class performs the same auth flow —
# cookies from finance.yahoo.com + crumb from getcrumb endpoint — but using
# curl_cffi directly against query1/query2, bypassing fc.yahoo.com entirely.

_OptionChain = namedtuple('OptionChain', ['calls', 'puts'])

# ── Module-level data cache (TTL = 5 min) — prevents duplicate API calls ──────
import time as _time
_DATA_CACHE: dict = {}
_CACHE_TTL = 300   # seconds


def _cache_get(key):
    entry = _DATA_CACHE.get(key)
    if entry and (_time.time() - entry[1]) < _CACHE_TTL:
        return entry[0]
    return None


def _cache_set(key, value):
    _DATA_CACHE[key] = (value, _time.time())


class _AuthError(Exception):
    """Raised when Yahoo Finance returns HTTP 401 and re-auth also fails."""


class _YFDirect:
    """
    Drop-in replacement for yf.Ticker() that bypasses fc.yahoo.com entirely,
    adds per-call data caching (5-min TTL) and retry with exponential back-off
    so a single 429 does not crash the whole analysis run.

    Usage is identical to yf.Ticker:
        t = _YFDirect('AAPL')
        t.history(period='1y')              → DataFrame with 'Close' column
        t.options                           → tuple of expiry date strings
        t.option_chain('2025-06-20')        → namedtuple(calls=df, puts=df)
        t.info                              → dict with 'dividendYield' etc.
        t.calendar                          → dict with 'Earnings Date' list
    """

    _session  = None
    _crumb    = None

    _Q1 = 'https://query1.finance.yahoo.com'
    _HEADERS = {
        'User-Agent': ('Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) '
                       'AppleWebKit/537.36 (KHTML, like Gecko) '
                       'Chrome/124.0.0.0 Safari/537.36'),
        'Accept': 'application/json',
    }

    def __init__(self, symbol: str):
        self.symbol           = symbol.upper().replace('^', '%5E')
        self._raw_symbol      = symbol.upper()
        self._use_yf_fallback = False
        self._yftk            = None
        _YFDirect._ensure_session()

    def _yf_fallback(self):
        """Lazily created yf.Ticker — used when direct API gets persistent 401."""
        if self._yftk is None:
            self._yftk = yf.Ticker(self._raw_symbol)
        return self._yftk

    @classmethod
    def _ensure_session(cls):
        if cls._session is not None and cls._crumb is not None:
            return
        try:
            from curl_cffi import requests as _cr
            cls._session = _cr.Session(impersonate='chrome')
        except ImportError:
            cls._session = requests.Session()
            cls._session.headers.update(cls._HEADERS)
        try:
            cls._session.get('https://finance.yahoo.com', timeout=12)
            r = cls._session.get(f'{cls._Q1}/v1/test/getcrumb', timeout=10)
            cls._crumb = r.text.strip() if r.status_code == 200 else ''
        except Exception:
            cls._crumb = ''

    @classmethod
    def _reset_session(cls):
        """Force re-auth on next call (used after persistent 429)."""
        cls._session = None
        cls._crumb   = None

    def _get(self, path, params=None, _retries=4):
        """
        GET with exponential back-off on 429 / 5xx.
        On 401: tries one re-auth (reset + new session), then raises _AuthError
        so callers can fall back to standard yfinance.
        """
        p = dict(params or {})
        if self._crumb:
            p['crumb'] = self._crumb

        last_err      = None
        _auth_retried = False
        for attempt in range(_retries):
            try:
                r = self._session.get(
                    f'{self._Q1}{path}', params=p, timeout=15
                )
                if r.status_code == 429:
                    wait = 2 ** (attempt + 1)
                    _time.sleep(wait)
                    if attempt == _retries - 1:
                        raise RuntimeError(
                            "Too Many Requests. Rate limited. "
                            f"Try after a while. (waited {wait}s)"
                        )
                    continue
                if r.status_code == 401:
                    if not _auth_retried:
                        _auth_retried = True
                        _YFDirect._reset_session()
                        _YFDirect._ensure_session()
                        p = dict(params or {})
                        if self._crumb:
                            p['crumb'] = self._crumb
                        continue
                    raise _AuthError(
                        "HTTP 401 Unauthorized – Yahoo Finance blocked this server IP. "
                        "Switching to standard yfinance fallback."
                    )
                r.raise_for_status()
                return r.json()
            except (_AuthError, RuntimeError):
                raise
            except Exception as e:
                last_err = e
                if attempt < _retries - 1:
                    _time.sleep(2 ** attempt)
        raise last_err or RuntimeError("Request failed after retries")

    def history(self, period='1y', interval='1d'):
        """Returns DataFrame with DatetimeIndex and 'Close' column (cached)."""
        cache_key = (self._raw_symbol, 'history', period, interval)
        cached = _cache_get(cache_key)
        if cached is not None:
            return cached

        try:
            if self._use_yf_fallback:
                raise _AuthError("fallback mode active")
            yf_period = '5d' if period == '2d' else period
            data   = self._get('/v8/finance/chart/' + self.symbol,
                               {'interval': interval, 'range': yf_period, 'events': 'div'})
            result = data['chart']['result'][0]
            ts     = result.get('timestamp', [])
            closes = result['indicators']['quote'][0].get('close', [])
            vols   = result['indicators']['quote'][0].get('volume', [])
            df = pd.DataFrame(
                {'Close': closes, 'Volume': vols},
                index=pd.to_datetime(ts, unit='s', utc=True).tz_convert('America/New_York')
            ).dropna(subset=['Close'])
            if period == '2d':
                df = df.tail(2)
            self._meta   = result['meta']
            self._events = result.get('events', {})
        except _AuthError:
            self._use_yf_fallback = True
            yf_period = '5d' if period == '2d' else period
            df = self._yf_fallback().history(period=yf_period, interval=interval)
            if period == '2d':
                df = df.tail(2)

        _cache_set(cache_key, df)
        return df

    @property
    def info(self):
        """Dict with dividendYield (cached)."""
        cache_key = (self._raw_symbol, 'info')
        cached = _cache_get(cache_key)
        if cached is not None:
            return cached

        if not hasattr(self, '_meta') or not getattr(self, '_events', {}).get('dividends'):
            self.history(period='1y')

        if self._use_yf_fallback or not hasattr(self, '_meta'):
            try:
                raw = self._yf_fallback().info or {}
            except Exception:
                raw = {}
            result = {
                'dividendYield':      (raw.get('dividendYield')
                                       or raw.get('trailingAnnualDividendYield') or 0.0),
                'regularMarketPrice': raw.get('regularMarketPrice') or raw.get('currentPrice'),
                'symbol':             raw.get('symbol', self._raw_symbol),
            }
            _cache_set(cache_key, result)
            return result

        meta      = self._meta
        div_yield = meta.get('dividendYield') or meta.get('trailingAnnualDividendYield')
        if not div_yield:
            divs = getattr(self, '_events', {}).get('dividends', {})
            if divs:
                amounts = sorted(divs.values(), key=lambda x: x['date'])[-4:]
                annual  = sum(d['amount'] for d in amounts)
                price   = meta.get('regularMarketPrice', 1)
                div_yield = annual / price if price else 0.0
        result = {
            'dividendYield':      div_yield or 0.0,
            'regularMarketPrice': meta.get('regularMarketPrice'),
            'symbol':             meta.get('symbol'),
        }
        _cache_set(cache_key, result)
        return result

    @property
    def options(self):
        """Tuple of expiry date strings — cached."""
        cache_key = (self._raw_symbol, 'options')
        cached = _cache_get(cache_key)
        if cached is not None:
            return cached

        try:
            if self._use_yf_fallback:
                raise _AuthError("fallback mode active")
            data   = self._get('/v7/finance/options/' + self.symbol)
            result = data['optionChain']['result'][0]
            out    = tuple(
                datetime.utcfromtimestamp(ts).strftime('%Y-%m-%d')
                for ts in result['expirationDates']
            )
        except _AuthError:
            self._use_yf_fallback = True
            out = tuple(self._yf_fallback().options)

        _cache_set(cache_key, out)
        return out

    def option_chain(self, date_str: str) -> _OptionChain:
        """Returns namedtuple(calls=DataFrame, puts=DataFrame) — cached."""
        cache_key = (self._raw_symbol, 'chain', date_str)
        cached = _cache_get(cache_key)
        if cached is not None:
            return cached

        def _to_df(contracts):
            if not contracts:
                return pd.DataFrame()
            df = pd.DataFrame(contracts)
            for col in ['bid', 'ask', 'volume', 'openInterest',
                        'impliedVolatility', 'lastPrice']:
                if col not in df.columns:
                    df[col] = 0
            df['volume']       = pd.to_numeric(df['volume'],       errors='coerce').fillna(0)
            df['openInterest'] = pd.to_numeric(df['openInterest'], errors='coerce').fillna(0)
            return df

        try:
            if self._use_yf_fallback:
                raise _AuthError("fallback mode active")
            unix_ts = int(datetime.strptime(date_str, '%Y-%m-%d')
                          .replace(tzinfo=_tz.utc).timestamp())
            data   = self._get('/v7/finance/options/' + self.symbol, {'date': unix_ts})
            result = data['optionChain']['result'][0]
            opts   = result['options'][0] if result['options'] else {}
            out = _OptionChain(
                calls=_to_df(opts.get('calls', [])),
                puts =_to_df(opts.get('puts',  [])),
            )
        except _AuthError:
            self._use_yf_fallback = True
            raw = self._yf_fallback().option_chain(date_str)
            out = _OptionChain(calls=raw.calls, puts=raw.puts)

        _cache_set(cache_key, out)
        return out

    @property
    def calendar(self):
        """
        Dict with 'Earnings Date' list — sourced from quoteSummary calendarEvents.
        Returns None if unavailable (no crash — earnings check is advisory only).
        """
        cache_key = (self._raw_symbol, 'calendar')
        cached = _cache_get(cache_key)
        if cached is not None:
            return cached

        try:
            data   = self._get('/v10/finance/quoteSummary/' + self.symbol,
                               {'modules': 'calendarEvents'})
            events = (data.get('quoteSummary', {})
                         .get('result', [{}])[0]
                         .get('calendarEvents', {})
                         .get('earnings', {}))
            raw_dates = events.get('earningsDate', [])
            dates     = [datetime.utcfromtimestamp(d['raw'])
                         for d in raw_dates if isinstance(d, dict) and 'raw' in d]
            result = {'Earnings Date': dates} if dates else None
        except Exception:
            result = None

        _cache_set(cache_key, result)
        return result


# ── Ticker factory: always returns _YFDirect (with caching + retry) ───────────
_TICKER_INSTANCES: dict = {}

def _ticker(symbol: str) -> _YFDirect:
    """
    Returns a cached _YFDirect instance for the symbol.
    All data fetched through this instance is also cached for 5 minutes,
    so repeated calls to .history() or .option_chain() within one analysis
    run hit the in-process cache instead of Yahoo Finance's API.
    """
    sym = symbol.upper()
    if sym not in _TICKER_INSTANCES:
        _TICKER_INSTANCES[sym] = _YFDirect(sym)
    return _TICKER_INSTANCES[sym]


# ─── Treasury Yield Curve (wallstreet-inspired) ───────────────────────────────
# wallstreet scraped the Treasury XML feed and interpolated across maturities.
# We replicate that concept using yfinance treasury tickers instead — no extra
# HTTP dependency, same result: a continuous r(T) function.

_TREASURY_TICKERS = {
    0.25:  '^IRX',   # 13-week T-bill
    5.0:   '^FVX',   # 5-year note
    10.0:  '^TNX',   # 10-year note
    30.0:  '^TYX',   # 30-year bond
}
_FALLBACK_RATE = 0.05   # used only if all fetches fail

_yield_curve_fn = None   # cached after first fetch


def _build_yield_curve() -> interp1d:
    """Fetch 4 treasury points and return a linear interpolation function."""
    import time
    maturities, rates = [], []
    for T, ticker in _TREASURY_TICKERS.items():
        try:
            hist = _ticker(ticker).history(period='5d')['Close']
            if not hist.empty:
                maturities.append(T)
                rates.append(hist.iloc[-1] / 100)
            time.sleep(0.2)
        except Exception:
            pass
    if len(maturities) < 2:
        return lambda _: _FALLBACK_RATE
    # fill_value="extrapolate" handles T outside [0.25, 30]
    return interp1d(maturities, rates, kind='linear', fill_value='extrapolate')


def get_risk_free_rate(T: float = 0.25) -> float:
    """
    Term-matched risk-free rate (continuously compounded) for maturity T years.

    Improvement over v2: v2 used ^IRX (13-week) for all maturities — wrong for
    LEAPS where the 5- or 10-year rate is more appropriate.
    Improvement over optionlab default: their docs default to a flat 0.0 rate.
    """
    global _yield_curve_fn
    if _yield_curve_fn is None:
        _yield_curve_fn = _build_yield_curve()
    return float(_yield_curve_fn(T))


# ─── Dividend Yield ───────────────────────────────────────────────────────────

def get_dividend_yield(symbol: str) -> float:
    """
    Continuous dividend yield q from yfinance info.
    Improvement over v1/v2: both versions used q=0 always, mispricing ITM
    calls on high-dividend stocks (energy, utilities, REITs).
    """
    try:
        info = _ticker(symbol).info
        return info.get('dividendYield') or 0.0
    except Exception:
        return 0.0


# ─── Black-Scholes with Dividends (Merton 1973) ───────────────────────────────

def bs_price(S, K, T, r, sigma, option_type='call', q=0.0):
    """Black-Scholes-Merton price. q = continuous dividend yield."""
    if T <= 0 or sigma <= 0 or S <= 0 or K <= 0:
        return np.nan
    d1 = (np.log(S / K) + (r - q + 0.5 * sigma**2) * T) / (sigma * np.sqrt(T))
    d2 = d1 - sigma * np.sqrt(T)
    if option_type == 'call':
        return S * np.exp(-q * T) * norm.cdf(d1) - K * np.exp(-r * T) * norm.cdf(d2)
    return K * np.exp(-r * T) * norm.cdf(-d2) - S * np.exp(-q * T) * norm.cdf(-d1)


def bs_greeks(S, K, T, r, sigma, option_type='call', q=0.0):
    """All Greeks with dividend adjustment. Theta per calendar day, Vega per 1% vol."""
    if T <= 0 or sigma <= 0 or S <= 0 or K <= 0:
        return {g: np.nan for g in ['delta', 'gamma', 'theta', 'vega', 'rho']}
    d1 = (np.log(S / K) + (r - q + 0.5 * sigma**2) * T) / (sigma * np.sqrt(T))
    d2 = d1 - sigma * np.sqrt(T)
    nd1 = norm.pdf(d1)
    Sq = S * np.exp(-q * T)

    gamma = nd1 / (Sq * sigma * np.sqrt(T))
    vega  = Sq * np.sqrt(T) * nd1 / 100

    if option_type == 'call':
        delta = np.exp(-q * T) * norm.cdf(d1)
        theta = (-(Sq * nd1 * sigma) / (2 * np.sqrt(T))
                 + q * Sq * norm.cdf(d1)
                 - r * K * np.exp(-r * T) * norm.cdf(d2)) / 365
        rho   = K * T * np.exp(-r * T) * norm.cdf(d2) / 100
    else:
        delta = np.exp(-q * T) * (norm.cdf(d1) - 1)
        theta = (-(Sq * nd1 * sigma) / (2 * np.sqrt(T))
                 - q * Sq * norm.cdf(-d1)
                 + r * K * np.exp(-r * T) * norm.cdf(-d2)) / 365
        rho   = -K * T * np.exp(-r * T) * norm.cdf(-d2) / 100

    return {'delta': delta, 'gamma': gamma, 'theta': theta, 'vega': vega, 'rho': rho}


def bs_second_order(S, K, T, r, sigma, option_type='call', q=0.0):
    """
    Second-order Greeks: Vanna, Volga, Charm, Speed, Color.

    Vanna  (∂²V/∂S∂σ): sensitivity of delta to vol change. High vanna → skew risk.
    Volga  (∂²V/∂σ²):  convexity of option value w.r.t. vol (vol of vol exposure).
    Charm  (∂Δ/∂t):    rate of delta decay per calendar day (pin-risk indicator).
    Speed  (∂Γ/∂S):    rate of gamma change per $1 move.
    Color  (∂Γ/∂t):    rate of gamma decay per calendar day.

    Sources: Haug (2007) "The Complete Guide to Option Pricing Formulas";
             Vanna-Volga method: Castagna & Mercurio (2007).
    """
    if T <= 0 or sigma <= 0 or S <= 0 or K <= 0:
        return {g: np.nan for g in ['vanna', 'volga', 'charm', 'speed', 'color']}
    d1 = (np.log(S / K) + (r - q + 0.5 * sigma**2) * T) / (sigma * np.sqrt(T))
    d2 = d1 - sigma * np.sqrt(T)
    nd1 = norm.pdf(d1)
    Sq  = S * np.exp(-q * T)
    vega_raw = Sq * np.sqrt(T) * nd1   # raw vega (not /100)

    vanna = -np.exp(-q * T) * nd1 * d2 / sigma          # ∂Δ/∂σ = ∂Vega/∂S
    volga = vega_raw * d1 * d2 / sigma                   # ∂²V/∂σ² (raw)
    # Speed = ∂Γ/∂S = -Γ × (1 + d1/(σ√T)) / S  (Haug 2007, corrected)
    gamma_raw = nd1 * np.exp(-q * T) / (S * sigma * np.sqrt(T))
    speed = -gamma_raw * (1 + d1 / (sigma * np.sqrt(T))) / S

    if option_type == 'call':
        charm = -np.exp(-q * T) * nd1 * (2*(r-q)*T - d2*sigma*np.sqrt(T)) / (2*T*sigma*np.sqrt(T))
        color = -np.exp(-q * T) * nd1 / (2*S*T*sigma*np.sqrt(T)) * (2*q*T + 1 + d1*(2*(r-q)*T - d2*sigma*np.sqrt(T)) / (sigma*np.sqrt(T)))
    else:
        charm = np.exp(-q * T) * nd1 * (2*(r-q)*T - d2*sigma*np.sqrt(T)) / (2*T*sigma*np.sqrt(T))
        color = np.exp(-q * T) * nd1 / (2*S*T*sigma*np.sqrt(T)) * (2*q*T + 1 + d1*(2*(r-q)*T - d2*sigma*np.sqrt(T)) / (sigma*np.sqrt(T)))

    return {
        'vanna': round(vanna, 6),
        'volga': round(volga / 100, 6),   # per 1% vol, consistent with vega
        'charm': round(charm / 365, 6),   # per calendar day
        'speed': round(speed, 8),
        'color': round(color / 365, 8),   # per calendar day
    }


def gamma_breakeven_move(gamma, theta, S):
    """
    Gamma scalping breakeven — minimum daily move for long gamma to cover theta cost.

    From P&L identity (Taleb 1997, Hull 2018):
        P&L_gamma = ½ × Γ × (ΔS)² − |Θ| × Δt
    Setting P&L = 0:
        |ΔS_BE| = √(2|Θ|/Γ)           [in dollars]
        σ_BE    = |ΔS_BE| / (S × √(1/252))  [annualised vol breakeven]

    If realised vol > σ_BE → long gamma generates positive expected P&L after theta.
    Source: Taleb (1997) Dynamic Hedging, Ch.7; confirmed Derman & Miller (2016).
    """
    if gamma <= 0 or theta == 0:
        return {'delta_s_be': np.nan, 'sigma_be_pct': np.nan}
    delta_s_be = float(np.sqrt(2 * abs(theta) / gamma))
    sigma_be   = delta_s_be / (S * np.sqrt(1.0 / 252.0))
    return {'delta_s_be': round(delta_s_be, 4), 'sigma_be_pct': round(sigma_be * 100, 2)}


def delta_gamma_var(delta, gamma, S, sigma, confidence=0.99, horizon_days=1):
    """
    Delta-Gamma VaR using Cornish-Fisher expansion (Britten-Jones & Neuberger 1999).

    Standard delta-only VaR is linear; adding gamma captures the convexity of options.
    VaR_α ≈ −(Δ·ΔS·z + ½Γ·ΔS²·(z²−1))
    where ΔS = σ_daily × S, z = Φ⁻¹(1−α) (negative tail quantile).

    Returns 1-day 99% loss estimate per contract (100 shares).
    Source: Britten-Jones & Neuberger (1999); JP Morgan RiskMetrics framework.
    """
    z    = norm.ppf(1 - confidence)            # e.g. −2.326 at 99%
    dS   = sigma * S * np.sqrt(horizon_days / 252.0)
    var_ = -(delta * dS * z + 0.5 * gamma * dS**2 * (z**2 - 1))
    return round(float(var_) * 100, 2)         # per contract


def dp_exit_threshold(K, mu_annual, sigma, r):
    """
    Dixit-Pindyck (1994) real-options optimal exit threshold for a long call.

    Solves the continuation-value ODE: ½σ²S²V_SS + μSV_S − rV = 0
    Solution exponent: β₁ = ½ − μ/σ² + √((μ/σ² − ½)² + 2r/σ²)
    Optimal exit (exercise) threshold: S* = β₁/(β₁−1) × K

    For a long call, S* > K is the price at which holding the option is suboptimal
    relative to taking the intrinsic value. Practical interpretation: if spot crosses
    S* from above (gap-fill scenario), the call has peaked in expected value.
    Source: Dixit & Pindyck (1994) "Investment under Uncertainty", Ch.5.
    """
    if sigma <= 0 or r <= 0:
        return np.nan
    mu = mu_annual
    b  = mu / sigma**2
    beta1 = 0.5 - b + float(np.sqrt((b - 0.5)**2 + 2 * r / sigma**2))
    if beta1 <= 1:
        return np.nan
    return round(beta1 / (beta1 - 1) * K, 2)


def garch_vol_forecast(log_returns, horizon=21):
    """
    GARCH(1,1) volatility forecast (Bollerslev 1986).

    σ²_t = ω + α·ε²_(t-1) + β·σ²_(t-1)
    Multi-step forecast: E[σ²_(t+h)] = σ²_LR + (α+β)^h × (σ²_current − σ²_LR)
    where σ²_LR = ω/(1−α−β)  [long-run variance]

    Falls back to HAR-RV if arch library is not installed.
    Horizon-averaged annualised vol is the main output (matches IV convention).
    Source: Bollerslev (1986) J. Econometrics; review in Andersen et al. (2006).
    """
    try:
        from arch import arch_model
        ret_pct = np.array(log_returns) * 100
        if len(ret_pct) < 50:
            raise ValueError("insufficient data")
        am  = arch_model(ret_pct, vol='Garch', p=1, q=1, dist='normal', rescale=False)
        res = am.fit(disp='off', show_warning=False)
        fc    = res.forecast(horizon=horizon)
        var_h = fc.variance.values[-1]         # horizon daily variances (pct²)
        # Mean integrated variance over [0, h] — matches IV convention.
        # IV is the BS-equivalent of √(E[∫₀ᵀσ²(t)dt]/T), so the *average*
        # of the per-step variance forecasts is the right object to
        # compare against an option's implied vol. Using only the terminal
        # point E[σ²_{t+h}] diverges from IV in trending vol regimes.
        # Reference: Andersen, Bollerslev, Christoffersen, Diebold (2006),
        # "Volatility and Correlation Forecasting", §3.
        ann_vol = float(np.sqrt(np.mean(var_h) * 252)) / 100
        params  = res.params
        persistence = float(params.get('alpha[1]', 0) + params.get('beta[1]', 0))
        return {'vol': round(ann_vol, 4), 'persistence': round(persistence, 4),
                'source': 'GARCH(1,1)'}
    except Exception:
        return None


def pin_risk_score(S, chain_df, bandwidth_pct=0.03,
                   T: float = None, r: float = 0.05, sigma: float = None, q: float = 0.0):
    """
    Pin-risk gravity score: identifies the strike where price is most likely to
    be 'pinned' near expiry due to dealer delta-hedging flows.

    Avellaneda & Lipkin (2003) derive the pinning drift as proportional to
    Σ Γ_i × OI_i × (K_i − S), i.e. **gamma-weighted** open interest — not
    raw OI. Raw OI over-weights deep ITM/OTM strikes which contribute almost
    nothing to the pinning force; the near-ATM strikes where Γ is largest
    do all the actual pinning work.

    If T and sigma are supplied, computes the true BS gamma at every strike.
    Otherwise falls back to a Gaussian kernel centred on S as a Γ-proxy
    (peaks ATM where real Γ peaks, with width bw·S).

    Source: Avellaneda & Lipkin (2003) "A Market-Induced Mechanism for Stock Pinning";
            Ni, Pearson & Poteshman (2005) "Stock Price Clustering on Option Expiration Dates".
    """
    strikes = chain_df['Strike'].values
    oi_vals = chain_df['Open Interest'].fillna(0).values
    bw = S * bandwidth_pct

    if T is not None and sigma is not None and T > 0 and sigma > 0:
        # True BS gamma per strike (Avellaneda-Lipkin canonical form).
        gammas = np.array([
            bs_greeks(S, float(K), float(T), float(r), float(sigma), 'call', float(q))['gamma']
            for K in strikes
        ])
        gammas = np.nan_to_num(gammas, nan=0.0, posinf=0.0, neginf=0.0)
        weights = oi_vals * gammas
    else:
        # Gaussian proxy for gamma (ATM-peaked) — used when σ/T not supplied.
        weights = oi_vals * np.exp(-0.5 * ((S - strikes) / bw) ** 2)

    if weights.sum() == 0:
        return None
    pin_K = float(strikes[np.argmax(weights)])
    pin_strength = float(weights.max() / weights.sum())   # 0–1, higher = stronger pull
    return {'pin_strike': round(pin_K, 2), 'pin_strength': round(pin_strength, 4),
            'distance_pct': round(abs(S - pin_K) / S * 100, 2)}


def sabr_iv_smile(F, K_array, T, alpha, beta, rho, nu):
    """
    SABR model implied volatility smile (Hagan et al 2002, "Managing Smile Risk").

    For each strike K, approximates the BS-equivalent IV accounting for skew and
    vol-of-vol, calibrated from 5 market quotes (ATM, 25D RR, 25D BF, 10D RR, 10D BF).

    Parameters: α (ATM vol level), β (backbone, fixed at 0.5 for equities),
                ρ (vol-spot correlation, drives skew), ν (vol of vol, drives smile).
    Source: Hagan, Kumar, Lesniewski, Woodward (2002) Wilmott Magazine.
    """
    ivs = []
    for K in K_array:
        if K <= 0 or F <= 0 or T <= 0:
            ivs.append(np.nan); continue
        if abs(F - K) < 1e-8:   # ATM formula
            atm_vol = (alpha / (F ** (1 - beta))
                       * (1 + ((1-beta)**2 * alpha**2 / (24 * F**(2-2*beta))
                               + rho*beta*nu*alpha / (4 * F**(1-beta))
                               + (2 - 3*rho**2) * nu**2 / 24) * T))
            ivs.append(atm_vol); continue
        logFK = np.log(F / K)
        FK_mid = (F * K) ** ((1 - beta) / 2)
        z  = (nu / alpha) * FK_mid * logFK
        xi = np.log((np.sqrt(1 - 2*rho*z + z**2) + z - rho) / (1 - rho))
        zxi = z / xi if abs(xi) > 1e-10 else 1.0
        numer = alpha
        denom = (FK_mid * (1 + (1-beta)**2/24 * logFK**2 + (1-beta)**4/1920 * logFK**4))
        body  = (1 + ((1-beta)**2 * alpha**2 / (24 * FK_mid**2)
                      + rho*beta*nu*alpha / (4 * FK_mid)
                      + (2-3*rho**2)*nu**2/24) * T)
        ivs.append(numer / denom * zxi * body)
    return np.array(ivs)


def fit_sabr_from_chain(chain_df, F, T, beta=0.5):
    """
    Fit SABR parameters (α, ρ, ν) from the available IV surface in chain_df.
    Returns None if fewer than 5 usable strikes exist.
    """
    from scipy.optimize import minimize
    sub = chain_df[chain_df['IV (%)'].notna()].copy()
    if len(sub) < 5:
        return None
    Ks   = sub['Strike'].values
    ivs  = sub['IV (%)'].values / 100.0

    def loss(params):
        a, rho, nu = params
        pred = sabr_iv_smile(F, Ks, T, a, beta, rho, nu)
        mask = ~np.isnan(pred)
        if mask.sum() < 3:
            return 1e9
        return float(np.mean((pred[mask] - ivs[mask])**2))

    try:
        from scipy.optimize import minimize
        res = minimize(loss, [ivs.mean(), -0.3, 0.4],
                       bounds=[(0.001, 5), (-0.999, 0.999), (0.001, 5)],
                       method='L-BFGS-B')
        alpha, rho, nu = res.x
        return {'alpha': round(alpha, 4), 'rho': round(rho, 4),
                'nu': round(nu, 4), 'beta': beta,
                'fit_rmse': round(np.sqrt(res.fun) * 100, 3)}
    except Exception:
        return None


# ─── Implied Volatility (Brent) ───────────────────────────────────────────────

def implied_volatility(S, K, T, r, market_price, option_type='call', q=0.0):
    if T <= 0 or market_price <= 0 or S <= 0 or K <= 0:
        return np.nan
    intrinsic = max(0, S - K) if option_type == 'call' else max(0, K - S)
    if market_price < intrinsic - 0.01:
        return np.nan
    try:
        return brentq(
            lambda sigma: bs_price(S, K, T, r, sigma, option_type, q) - market_price,
            1e-6, 10.0, xtol=1e-6, maxiter=500
        )
    except (ValueError, RuntimeError):
        return np.nan


# ─── Historical Volatility ────────────────────────────────────────────────────

def historical_volatility(symbol: str, period: str = '1y') -> float:
    hist = _ticker(symbol).history(period=period)
    if len(hist) < 10:
        return np.nan
    log_ret = np.log(hist['Close'] / hist['Close'].shift(1)).dropna()
    return log_ret.std() * np.sqrt(252)


# ─── HAR-RV Volatility Forecast  (Corsi 2009) ────────────────────────────────
# Heterogeneous Autoregressive model of Realized Volatility.
# Regresses next-period RV on daily, weekly (5d), and monthly (22d) RV components.
# Outperforms GARCH and simple rolling HV at 5-22 day horizons out-of-sample.
# Source: Corsi (2009) Journal of Financial Econometrics; confirmed by ML studies.

def har_rv_forecast(prices: list) -> float:
    """
    Fit HAR-RV on historical daily squared log-returns and return the
    one-step-ahead annualised volatility forecast (as a decimal, e.g. 0.35).

    Falls back to 30-day close-to-close HV when data is insufficient.
    """
    if len(prices) < 50:
        rets = [np.log(prices[k] / prices[k-1]) for k in range(1, len(prices))]
        return float(np.std(rets) * np.sqrt(252)) if rets else 0.30

    log_rets = [np.log(prices[k] / prices[k-1]) for k in range(1, len(prices))]
    # Daily realised variance (annualised)
    rv_d = [r**2 * 252 for r in log_rets]

    X, y = [], []
    for t in range(22, len(rv_d) - 1):
        rv1  = rv_d[t]
        rv5  = float(np.mean(rv_d[t-4 : t+1]))
        rv22 = float(np.mean(rv_d[t-21: t+1]))
        X.append([1.0, rv1, rv5, rv22])
        y.append(rv_d[t + 1])

    if len(X) < 10:
        return float(np.std(log_rets[-30:]) * np.sqrt(252))

    beta, *_ = np.linalg.lstsq(np.array(X), np.array(y), rcond=None)
    rv1  = rv_d[-1]
    rv5  = float(np.mean(rv_d[-5:]))
    rv22 = float(np.mean(rv_d[-22:]))
    forecast_var = beta[0] + beta[1]*rv1 + beta[2]*rv5 + beta[3]*rv22
    return float(np.sqrt(max(forecast_var, 1e-6)))


# ─── VIX Term Structure  (regime filter) ─────────────────────────────────────
# VIX (1-month) vs VIX3M (3-month) slope signals the vol supply/demand regime.
# Steep contango (>10%) → dealers selling vol → short-vol edge.
# Backwardation (<0%)   → vol spike risk → avoid selling, reduce size.
# Source: Quantpedia / AQR; empirically validated 2004-present.

def vix_term_structure() -> dict | None:
    try:
        vix   = float(_ticker('^VIX').history(period='5d')['Close'].iloc[-1])
        vix3m = float(_ticker('^VIX3M').history(period='5d')['Close'].iloc[-1])
        slope = (vix3m - vix) / vix
        if slope > 0.10:
            regime = 'CONTANGO >10% — short-vol edge (IV sellers favoured)'
        elif slope > 0:
            regime = 'MILD CONTANGO — neutral environment'
        elif slope > -0.05:
            regime = 'FLAT/SLIGHT BACKWARDATION — caution, reduce size'
        else:
            regime = 'BACKWARDATION — vol spike risk, avoid selling options'
        return {'VIX': round(vix, 2), 'VIX3M': round(vix3m, 2),
                'slope': round(slope, 3), 'regime': regime}
    except Exception:
        return None


# ─── 25-Delta Risk Reversal  (Bali & Murray 2013) ────────────────────────────
# RR = IV(25Δ put) − IV(25Δ call).
# Negative → put skew → market fears downside → bearish lean.
# Positive → call skew → unusual bullish demand.
# Source: Bali & Murray (2013) JFQA — predicts cross-section of option returns.

def risk_reversal_25d(chain_df: pd.DataFrame) -> dict | None:
    try:
        calls = chain_df[(chain_df['Type'] == 'CALL') &
                         (chain_df['Delta'].between(0.20, 0.30))]
        puts  = chain_df[(chain_df['Type'] == 'PUT')  &
                         (chain_df['Delta'].between(-0.30, -0.20))]
        if calls.empty or puts.empty:
            return None
        iv_c = calls.loc[calls['Delta'].sub(0.25).abs().idxmin(), 'IV (%)']
        iv_p = puts.loc[puts['Delta'].sub(-0.25).abs().idxmin(), 'IV (%)']
        rr   = float(iv_p - iv_c)
        if rr > 5:
            signal = 'PUT SKEW — fear of downside → bearish lean'
        elif rr < -5:
            signal = 'CALL SKEW — unusual upside demand → bullish lean'
        else:
            signal = 'BALANCED — no strong directional signal from skew'
        return {'iv_put_25': round(float(iv_p), 1),
                'iv_call_25': round(float(iv_c), 1),
                'RR': round(rr, 1), 'signal': signal}
    except Exception:
        return None


# ─── Fractional Kelly Position Sizing  (arXiv 2025) ──────────────────────────
# Kelly fraction f* = (p×b − q) / b  where b = avg_win / avg_loss ratio.
# Full Kelly is too aggressive for options (estimation error → ruin).
# 25% Kelly is the standard fractional application; scale down further in
# high-VIX regimes per the arXiv 2025 hybrid approach.

def kelly_contracts(win_rate: float, avg_win: float, avg_loss: float,
                    account: float, premium: float,
                    fraction: float = 0.25, vix: float = 20.0) -> dict:
    """
    Returns recommended number of contracts and the Kelly fraction used.
    Reduces fraction to 12.5% when VIX > 30 (high-risk regime).
    """
    if avg_loss <= 0 or premium <= 0 or account <= 0:
        return {'contracts': 1, 'kelly_f': 0.0, 'risk_pct': 0.0}
    b            = abs(avg_win / avg_loss)
    q            = 1.0 - win_rate
    kelly_full   = max(0.0, (win_rate * b - q) / b)
    # Halve fraction in high-VIX regime (arXiv 2025 hybrid approach)
    adj_fraction = fraction / 2 if vix > 30 else fraction
    kelly_frac   = kelly_full * adj_fraction
    max_risk     = account * kelly_frac
    n            = max(1, int(max_risk / (premium * 100)))
    return {
        'contracts':  n,
        'kelly_f':    round(kelly_full, 3),
        'kelly_frac': round(kelly_frac, 3),
        'risk_pct':   round(kelly_frac * 100, 1),
    }


# ─── American Option Pricing  (optlib / Bjerksund-Stensland 2002) ─────────────
# Ported from dbrojas/optlib gbs.py.
# Why: Every US equity option is American-style. For deep ITM puts and high-
# dividend calls, early exercise is rational — European BS ignores this and
# systematically underprices. B-S 2002 is the standard closed-form approximation.
#
# Cost-of-carry notation (GBS framework, same as optlib):
#   b = r       → Black-Scholes (no dividend)
#   b = r - q   → Merton / dividend-paying stock  ← what we use
#   b = 0       → Black-76 futures
#   b = r - rf  → Garman-Kohlhagen FX

def _phi(fs, t, gamma, h, i, r, b, v):
    """Helper for Bjerksund-Stensland (bivariate normal CDF term)."""
    d1 = -(math.log(fs / h) + (b + (gamma - 0.5) * v**2) * t) / (v * math.sqrt(t))
    d2 = -(math.log((i**2) / (fs * h)) + (b + (gamma - 0.5) * v**2) * t) / (v * math.sqrt(t))
    kappa = 2 * b / v**2 + 2 * gamma - 1
    return (math.exp((b - r) * t) * (fs**gamma) *
            (norm.cdf(d1) - ((i / fs)**kappa) * norm.cdf(d2)))


def _psi(fs, t2, gamma, h, i2, i1, t1, r, b, v):
    """Helper for Bjerksund-Stensland 2002 (bivariate normal CDF term)."""
    vsqrt_t1 = v * math.sqrt(t1)
    vsqrt_t2 = v * math.sqrt(t2)

    d1 = -(math.log(fs / i1) + (b + (gamma - 0.5) * v**2) * t1) / vsqrt_t1
    d3 = -(math.log(fs / i1) - (b + (gamma - 0.5) * v**2) * t1) / vsqrt_t1
    d2 = -(math.log((i2**2) / (fs * i1)) + (b + (gamma - 0.5) * v**2) * t1) / vsqrt_t1
    d4 = -(math.log((i2**2) / (fs * i1)) - (b + (gamma - 0.5) * v**2) * t1) / vsqrt_t1
    e1 = -(math.log(fs / h) + (b + (gamma - 0.5) * v**2) * t2) / vsqrt_t2
    e2 = -(math.log((i2**2) / (fs * h)) + (b + (gamma - 0.5) * v**2) * t2) / vsqrt_t2
    e3 = -(math.log((i1**2) / (fs * h)) + (b + (gamma - 0.5) * v**2) * t2) / vsqrt_t2
    e4 = -(math.log((fs * (i1**2)) / (h * (i2**2))) + (b + (gamma - 0.5) * v**2) * t2) / vsqrt_t2

    rho = math.sqrt(t1 / t2)
    kappa = 2 * b / v**2 + 2 * gamma - 1

    return (math.exp((b - r) * t2) * (fs**gamma) * (
        mvn.mvnun([d1, e1], [0, 0], [[1, -rho], [-rho, 1]])[0]
        - ((i2 / fs)**kappa) * mvn.mvnun([d2, e2], [0, 0], [[1, -rho], [-rho, 1]])[0]
        - ((i1 / fs)**kappa) * mvn.mvnun([d3, e3], [0, 0], [[1, rho], [rho, 1]])[0]
        + ((i1 / i2)**kappa) * mvn.mvnun([d4, e4], [0, 0], [[1, rho], [rho, 1]])[0]
    ))


def american_option_price(S, K, T, r, sigma, option_type='call', q=0.0):
    """
    American option price via Bjerksund-Stensland 2002 closed-form approximation.

    For puts, uses put-call transformation: price put as mirrored call.
    Returns np.nan for degenerate inputs. Falls back to European BS when
    early exercise is never optimal (b >= r, i.e. no dividend pressure).

    Args:
        S    : spot price
        K    : strike price
        T    : time to expiry in years
        r    : risk-free rate
        sigma: volatility
        option_type: 'call' or 'put'
        q    : continuous dividend yield (b = r - q internally)
    """
    if T <= 0 or sigma < 0.005 or S <= 0 or K <= 0:
        return np.nan

    b = r - q   # cost of carry

    def _bs2002_call(fs, x, t, _r, _b, v):
        """Bjerksund-Stensland 2002 for a call."""
        euro = bs_price(fs, x, t, _r, v, 'call', q)   # European lower bound

        if _b >= _r:   # early exercise never optimal
            return euro

        v2  = v ** 2
        t1  = 0.5 * (math.sqrt(5) - 1) * t

        beta_inner = abs((_b / v2 - 0.5)**2 + 2 * _r / v2)
        beta       = (0.5 - _b / v2) + math.sqrt(beta_inner)
        b_inf      = (beta / (beta - 1)) * x
        b_zero     = max(x, (_r / (_r - _b)) * x)

        h1 = -(_b * t1 + 2 * v * math.sqrt(t1)) * ((x**2) / ((b_inf - b_zero) * b_zero))
        h2 = -(_b * t  + 2 * v * math.sqrt(t))  * ((x**2) / ((b_inf - b_zero) * b_zero))

        i1 = b_zero + (b_inf - b_zero) * (1 - math.exp(h1))
        i2 = b_zero + (b_inf - b_zero) * (1 - math.exp(h2))

        alpha1 = (i1 - x) * (i1 ** (-beta))
        alpha2 = (i2 - x) * (i2 ** (-beta))

        if fs >= i2:
            return fs - x   # immediate exercise

        value = (alpha2 * (fs**beta)
                 - alpha2 * _phi(fs, t, beta, i2, i2, _r, _b, v)
                 + _phi(fs, t, 1, i2, i2, _r, _b, v)
                 - _phi(fs, t, 1, x, i2, _r, _b, v)
                 - x * _phi(fs, t, 0, i2, i2, _r, _b, v)
                 + x * _phi(fs, t, 0, x, i2, _r, _b, v)
                 - alpha1 * _phi(fs, t, beta, i1, i2, _r, _b, v)
                 + alpha1 * _psi(fs, t, beta, i1, i2, i1, t1, _r, _b, v)
                 - _psi(fs, t, 1, i1, i2, i1, t1, _r, _b, v)
                 + _psi(fs, t, 1, x, i2, i1, t1, _r, _b, v)
                 + x * _psi(fs, t, 0, i1, i2, i1, t1, _r, _b, v)
                 - x * _psi(fs, t, 0, x, i2, i1, t1, _r, _b, v))

        return max(value, euro)   # no less than European value

    try:
        if option_type == 'call':
            return _bs2002_call(S, K, T, r, b, sigma)
        else:
            # Put-call symmetry: American put on (S, K, r, b) = American call on (K, S, r-b, -b)
            return _bs2002_call(K, S, T, r - b, -b, sigma)
    except Exception:
        return np.nan


def american_early_exercise_premium(S, K, T, r, sigma, option_type='call', q=0.0):
    """Early exercise premium = American price − European price (≥ 0)."""
    amer = american_option_price(S, K, T, r, sigma, option_type, q)
    euro = bs_price(S, K, T, r, sigma, option_type, q)
    if np.isnan(amer) or np.isnan(euro):
        return np.nan
    return max(0.0, amer - euro)


# ─── CBOE Direct Data Source  (OpenBB CBOE provider endpoint) ─────────────────
# OpenBB's CBOE provider (openbb_platform/providers/cboe/) revealed this free
# public endpoint. No API key required. 15-minute delayed exchange data.
# More reliable than yfinance (Yahoo Finance) for liquid names.
#
# Endpoint:  https://cdn.cboe.com/api/global/delayed_quotes/options/{symbol}.json
# Index ETFs: https://cdn.cboe.com/api/global/delayed_quotes/options/_{symbol}.json

_CBOE_INDEX_SYMBOLS = {'SPX', 'VIX', 'RUT', 'NDX', 'DJX', 'XSP', 'MXEA', 'MXEF'}
_CBOE_HEADERS = {'User-Agent': 'Mozilla/5.0'}


def fetch_cboe_chain(symbol: str) -> pd.DataFrame:
    """
    Fetch the live option chain directly from CBOE's public CDN endpoint.

    Returns a DataFrame with columns matching our analyze_chain() output schema
    so it can be used as a drop-in replacement for the yfinance chain.
    Returns empty DataFrame if the symbol isn't found on CBOE.

    Source: OpenBB openbb_platform/providers/cboe/openbb_cboe/models/options_chains.py
    """
    sym = symbol.upper().replace('^', '')
    url = (f"https://cdn.cboe.com/api/global/delayed_quotes/options/_{sym}.json"
           if sym in _CBOE_INDEX_SYMBOLS
           else f"https://cdn.cboe.com/api/global/delayed_quotes/options/{sym}.json")

    try:
        resp = requests.get(url, headers=_CBOE_HEADERS, timeout=10)
        resp.raise_for_status()
        data = resp.json()
    except requests.exceptions.HTTPError as e:
        if e.response.status_code in (401, 403, 404):
            print(f"  [CBOE] Endpoint returned {e.response.status_code} for {symbol}. Falling back...")
        else:
            print(f"  [CBOE] HTTP Error for {symbol}: {e}")
        return pd.DataFrame()
    except Exception as e:
        print(f"  [CBOE] Could not fetch {symbol}: {e}")
        return pd.DataFrame()

    options = data.get('data', {}).get('options', [])
    if not options:
        return pd.DataFrame()

    underlying_price = data.get('data', {}).get('current_price',
                        data.get('data', {}).get('close', np.nan))

    rows = []
    for opt in options:
        try:
            name = opt.get('option', '')          # e.g. "AAPL240119C00150000"
            bid  = float(opt.get('bid',  0) or 0)
            ask  = float(opt.get('ask',  0) or 0)
            iv   = float(opt.get('iv',   0) or 0) / 100   # CBOE provides as percent
            vol  = int(opt.get('volume', 0) or 0)
            oi   = int(opt.get('open_interest', 0) or 0)
            last = float(opt.get('last_trade_price', 0) or 0)

            # parse expiry + strike + type from option name
            # Format: {SYM}{YY}{MM}{DD}{C/P}{8-digit-strike×1000}
            import re
            m = re.search(r'(\d{6})([CP])(\d{8})$', name)
            if not m:
                continue
            exp_str    = m.group(1)   # YYMMDD
            opt_type   = 'call' if m.group(2) == 'C' else 'put'
            strike     = int(m.group(3)) / 1000.0
            expiry     = datetime.strptime(exp_str, '%y%m%d').strftime('%Y-%m-%d')

            mid = (bid + ask) / 2.0 if bid > 0 and ask > 0 else last

            rows.append({
                'source':        'CBOE',
                'expiry':        expiry,
                'option_type':   opt_type,
                'strike':        strike,
                'mid_price':     round(mid, 4),
                'bid':           bid,
                'ask':           ask,
                'last_price':    last,
                'volume':        vol,
                'open_interest': oi,
                'iv_cboe':       round(iv * 100, 2),   # CBOE-computed IV %
                'underlying':    underlying_price,
            })
        except Exception:
            continue

    return pd.DataFrame(rows)


# ─── Multi-Leg Strategy Scanner  (Lean reference patterns) ────────────────────
# QuantConnect Lean has IronCondorStrategyAlgorithm.py, CallCalendarSpread, etc.
# Those only run inside the Lean/C# runtime. This is a standalone re-implementation
# of the same strategy construction logic using our live chain data.

def find_iron_condor(chain_df: pd.DataFrame, S: float,
                     wing_delta: float = 0.15, width: int = 5) -> dict:
    """
    Find the best iron condor setup from the nearest expiry chain.

    Logic mirrors Lean's IronCondorStrategyAlgorithm.py:
    - Short put at ~-wing_delta, long put `width` strikes below
    - Short call at ~+wing_delta, long call `width` strikes above

    Args:
        chain_df   : DataFrame from analyze_chain() for one expiry
        S          : current spot price
        wing_delta : target absolute delta for the short strikes
        width      : number of strikes between long and short (spread width)

    Returns dict with all four legs and strategy metrics, or empty dict if
    the chain doesn't have enough strikes.
    """
    # Short put: closest to -wing_delta
    puts = chain_df[chain_df['Type'] == 'PUT'][['Strike', 'Delta', 'Mid Price', 'IV (%)']].copy()
    puts = puts[puts['Delta'].notna()]
    if puts.empty:
        return {}
    short_put_row = puts.iloc[(puts['Delta'] - (-wing_delta)).abs().argsort().iloc[0]]
    short_put_K   = short_put_row['Strike']

    # Short call: closest to +wing_delta
    calls = chain_df[chain_df['Type'] == 'CALL'][['Strike', 'Delta', 'Mid Price', 'IV (%)']].copy()
    calls = calls[calls['Delta'].notna()]
    if calls.empty:
        return {}
    short_call_row = calls.iloc[(calls['Delta'] - wing_delta).abs().argsort().iloc[0]]
    short_call_K   = short_call_row['Strike']

    # Wing strikes from type-specific lists so the lookup is guaranteed to exist
    put_strikes  = sorted(puts['Strike'].unique())
    call_strikes = sorted(calls['Strike'].unique())

    if len(put_strikes) < 2 or len(call_strikes) < 2:
        return {}

    sp_idx = put_strikes.index(short_put_K)
    lp_idx = max(0, sp_idx - width)
    long_put_K = put_strikes[lp_idx]
    if long_put_K == short_put_K:          # not enough strikes below
        return {}
    long_put_row = puts[puts['Strike'] == long_put_K].iloc[0]

    sc_idx = call_strikes.index(short_call_K)
    lc_idx = min(len(call_strikes) - 1, sc_idx + width)
    long_call_K = call_strikes[lc_idx]
    if long_call_K == short_call_K:        # not enough strikes above
        return {}
    long_call_row = calls[calls['Strike'] == long_call_K].iloc[0]

    net_credit   = (short_put_row['Mid Price'] + short_call_row['Mid Price']
                    - long_put_row['Mid Price']  - long_call_row['Mid Price'])
    put_spread   = short_put_K  - long_put_K
    call_spread  = long_call_K  - short_call_K
    wing_width   = max(put_spread, call_spread)

    if net_credit > 0:
        # Standard iron condor: net credit received
        max_loss = wing_width - net_credit
        be_low   = short_put_K  - net_credit
        be_high  = short_call_K + net_credit
    else:
        # Net debit (unusual) — max loss = net debit paid + any spread reversal
        max_loss = wing_width + abs(net_credit)
        be_low   = short_put_K  + net_credit   # credit is negative, so shifts out
        be_high  = short_call_K - net_credit

    if max_loss <= 0:
        return {}   # zero/negative max_loss is nonsensical — data quality issue

    bp_required = max_loss * 100
    ror         = net_credit / max_loss * 100 if max_loss > 0 else np.nan

    return {
        'long_put_strike':   long_put_K,
        'short_put_strike':  short_put_K,
        'short_call_strike': short_call_K,
        'long_call_strike':  long_call_K,
        'net_credit ($)':    round(net_credit, 2),
        'max_loss ($)':      round(max_loss, 2),
        'bp_required ($)':   round(bp_required, 2),
        'return_on_risk (%)':round(ror, 2) if not np.isnan(ror) else np.nan,
        'breakeven_low':     round(be_low, 2),
        'breakeven_high':    round(be_high, 2),
        'expiry':            chain_df['Expiry'].iloc[0],
    }


# ─── Helpers ──────────────────────────────────────────────────────────────────

def mid_price(row):
    bid = row.get('bid', 0) or 0
    ask = row.get('ask', 0) or 0
    return (bid + ask) / 2.0 if (bid > 0 and ask > 0) else row['lastPrice']


def nearest_listed_strike(S: float, K_target: float) -> float:
    """
    Snap a target strike to the nearest listed-strike grid increment.

    OCC strike-listing standards (CBOE Rulebook §5.5):
      • S < $25    → $0.50 increments (was $1 pre-2010)
      • $25–200    → $1 or $2.50 (we use $1 as the conservative midpoint)
      • $200–500   → $5
      • S ≥ $500   → $10
    Mega-cap names (e.g. BRK.A) and weeklies for SPX/QQQ use $5 even at
    higher prices, but for backtest fidelity at single-stock level the
    above grid is a faithful approximation.
    """
    if S < 25:    grid = 0.5
    elif S < 200: grid = 1.0
    elif S < 500: grid = 5.0
    else:         grid = 10.0
    return round(round(K_target / grid) * grid, 2)


def classify_moneyness(S, K, option_type):
    ratio = S / K
    if option_type == 'call':
        if ratio > 1.02: return 'ITM'
        if ratio < 0.98: return 'OTM'
    else:
        if ratio < 0.98: return 'ITM'
        if ratio > 1.02: return 'OTM'
    return 'ATM'


def prob_expire_itm(S, K, T, r, sigma, option_type='call', q=0.0):
    if T <= 0 or sigma <= 0:
        return np.nan
    d2 = (np.log(S / K) + (r - q - 0.5 * sigma**2) * T) / (sigma * np.sqrt(T))
    return norm.cdf(d2) if option_type == 'call' else norm.cdf(-d2)


def expected_move(S, iv, T):
    if np.isnan(iv) or iv <= 0 or T <= 0:
        return np.nan
    return S * iv * np.sqrt(T)


# ─── Full Chain Analysis ──────────────────────────────────────────────────────

def analyze_chain(symbol, expiry_str, S, hv, q=0.0):
    chain = _ticker(symbol).option_chain(expiry_str)
    trade_dt  = datetime.now().replace(tzinfo=None)
    expiry_dt = datetime.strptime(expiry_str, '%Y-%m-%d')
    T = max((expiry_dt - trade_dt).days / 365.0, 1 / 365.0)
    r = get_risk_free_rate(T)   # ← term-matched rate

    rows = []
    for opt_type, df in [('call', chain.calls), ('put', chain.puts)]:
        for _, row in df.iterrows():
            K     = row['strike']
            price = mid_price(row)
            if price <= 0:
                continue

            iv = implied_volatility(S, K, T, r, price, opt_type, q)
            sig = iv if not np.isnan(iv) else hv
            g   = bs_greeks(S, K, T, r, sig, opt_type, q)

            intrinsic  = max(0, S - K) if opt_type == 'call' else max(0, K - S)
            time_val   = max(0, price - intrinsic)
            p_itm      = prob_expire_itm(S, K, T, r, sig, opt_type, q)
            amer_price = american_option_price(S, K, T, r, sig, opt_type, q)
            eep        = american_early_exercise_premium(S, K, T, r, sig, opt_type, q)

            rows.append({
                'Expiry':                  expiry_str,
                'Type':                    opt_type.upper(),
                'Strike':                  K,
                'Moneyness':               classify_moneyness(S, K, opt_type),
                'Mid Price':               round(price, 4),
                'Bid':                     row.get('bid', np.nan),
                'Ask':                     row.get('ask', np.nan),
                'Last Price':              row['lastPrice'],
                'Volume':                  int(pd.to_numeric(row.get('volume', 0), errors='coerce') or 0),
                'Open Interest':           int(pd.to_numeric(row.get('openInterest', 0), errors='coerce') or 0),
                'IV (%)':                  round(iv * 100, 2) if not np.isnan(iv) else np.nan,
                'HV (%)':                  round(hv * 100, 2),
                'IV-HV Spread (%)':        round((iv - hv) * 100, 2) if not np.isnan(iv) else np.nan,
                'Euro BS Price':           round(bs_price(S, K, T, r, sig, opt_type, q), 4),
                'American Price (B-S02)':  round(amer_price, 4) if not np.isnan(amer_price) else np.nan,
                'Early Exercise Prem ($)': round(eep, 4) if not np.isnan(eep) else np.nan,
                'Intrinsic Value':         round(intrinsic, 4),
                'Time Value':              round(time_val, 4),
                'Delta':                   round(g['delta'], 4),
                'Gamma':                   round(g['gamma'], 6),
                'Theta ($/day)':           round(g['theta'], 4),
                'Vega ($/1%vol)':          round(g['vega'], 4),
                'Rho ($/1%)':              round(g['rho'], 4),
                'Prob ITM (%)':            round(p_itm * 100, 2) if not np.isnan(p_itm) else np.nan,
                'Expected Move ($)':       round(expected_move(S, sig, T), 2),
                'Risk-Free Rate (%)':      round(r * 100, 3),
                'Div Yield (%)':           round(q * 100, 2),
                'T (years)':               round(T, 4),
            })

    return pd.DataFrame(rows)


# ─── Vol Skew + Put-Call Parity ───────────────────────────────────────────────

def vol_skew(df):
    calls = (df[df['Type'] == 'CALL'][['Strike', 'IV (%)']]
             .rename(columns={'IV (%)': 'Call IV (%)'}).set_index('Strike'))
    puts  = (df[df['Type'] == 'PUT'][['Strike', 'IV (%)']]
             .rename(columns={'IV (%)': 'Put IV (%)'}).set_index('Strike'))
    return calls.join(puts, how='outer').reset_index().sort_values('Strike')


def parity_check(df, S):
    calls = (df[df['Type'] == 'CALL'][['Strike', 'Mid Price', 'T (years)', 'Risk-Free Rate (%)']]
             .rename(columns={'Mid Price': 'C'}))
    puts  = (df[df['Type'] == 'PUT'][['Strike', 'Mid Price', 'T (years)', 'Risk-Free Rate (%)']]
             .rename(columns={'Mid Price': 'P'}))
    m = pd.merge(calls, puts, on=['Strike', 'T (years)', 'Risk-Free Rate (%)'])
    m['r']             = m['Risk-Free Rate (%)'] / 100
    m['C-P']           = m['C'] - m['P']
    m['S-PV(K)']       = S - m['Strike'] * np.exp(-m['r'] * m['T (years)'])
    m['Parity Err ($)'] = (m['C-P'] - m['S-PV(K)']).abs().round(4)
    return m[['Strike', 'C', 'P', 'C-P', 'S-PV(K)', 'Parity Err ($)']].sort_values('Strike')


# ─── Income Screener: CSP + CC  (options_lab-inspired) ───────────────────────
# options_lab used simple linear ARR: premium/collateral * (365/DTE).
# We use the compound formula (1 + r)^(1/years) - 1, which is more accurate
# when comparing strategies across very different DTE buckets.

def _compound_arr(premium: float, collateral: float, dte: int) -> float:
    """Compounded Annualized Rate of Return on collateral."""
    if collateral <= 0 or dte <= 0:
        return np.nan
    holding_return = premium / collateral
    years = dte / 365.0
    return ((1 + holding_return) ** (1.0 / years) - 1) * 100


def cash_secured_put_screener(symbol: str, max_breakeven: float,
                               min_volume: int = 10, top_k: int = 10) -> pd.DataFrame:
    """
    Ranks every put across all expiries by compounded ARR on cash collateral.
    Breakeven = strike - bid.  Collateral = strike (cash needed to secure the put).

    Based on options_lab find_best_put_to_sell, using compound ARR.
    """
    ticker = _ticker(symbol)
    rows = []
    for exp_date in ticker.options:
        dte = (datetime.strptime(exp_date, '%Y-%m-%d') - datetime.now()).days
        if dte <= 0:
            continue
        try:
            puts = ticker.option_chain(exp_date).puts
        except Exception:
            continue
        puts = puts[(puts['volume'] > min_volume) & (puts['bid'] > 0.01)].copy()
        puts['breakeven'] = puts['strike'] - puts['bid']
        puts = puts[puts['breakeven'] < max_breakeven]
        for _, r in puts.iterrows():
            arr = _compound_arr(r['bid'], r['strike'], dte)
            rows.append({
                'Expiry':      exp_date,
                'DTE':         dte,
                'Strike':      r['strike'],
                'Bid':         r['bid'],
                'Breakeven':   round(r['breakeven'], 2),
                'Volume':      r.get('volume', 0),
                'OI':          r.get('openInterest', 0),
                'IV (%)':      round(r.get('impliedVolatility', np.nan) * 100, 1),
                'ARR (%)':     round(arr, 1) if not np.isnan(arr) else np.nan,
            })

    df = pd.DataFrame(rows)
    if df.empty:
        return df
    return (df.sort_values('ARR (%)', ascending=False)
              .head(top_k)
              .reset_index(drop=True))


def covered_call_screener(symbol: str, S: float, min_breakeven: float,
                           min_volume: int = 10, top_k: int = 10) -> pd.DataFrame:
    """
    Ranks every call across all expiries by compounded ARR on equity collateral.
    Collateral = current stock price (cost of holding 100 shares).
    Breakeven = strike + bid.
    """
    ticker = _ticker(symbol)
    rows = []
    for exp_date in ticker.options:
        dte = (datetime.strptime(exp_date, '%Y-%m-%d') - datetime.now()).days
        if dte <= 0:
            continue
        try:
            calls = ticker.option_chain(exp_date).calls
        except Exception:
            continue
        calls = calls[(calls['volume'] > min_volume) & (calls['bid'] > 0.01)
                      & (calls['impliedVolatility'] > 0.01)
                      & (calls['impliedVolatility'] < 3.0)].copy()
        calls['breakeven'] = calls['strike'] + calls['bid']
        calls = calls[calls['breakeven'] > min_breakeven]
        for _, r in calls.iterrows():
            arr = _compound_arr(r['bid'], S, dte)
            rows.append({
                'Expiry':      exp_date,
                'DTE':         dte,
                'Strike':      r['strike'],
                'Bid':         r['bid'],
                'Breakeven':   round(r['breakeven'], 2),
                'Volume':      r.get('volume', 0),
                'OI':          r.get('openInterest', 0),
                'IV (%)':      round(r.get('impliedVolatility', np.nan) * 100, 1),
                'ARR (%)':     round(arr, 1) if not np.isnan(arr) else np.nan,
            })

    df = pd.DataFrame(rows)
    if df.empty:
        return df
    return (df.sort_values('ARR (%)', ascending=False)
              .head(top_k)
              .reset_index(drop=True))


# ─── Strategy P&L + PoP  (optionlab) ─────────────────────────────────────────
# optionlab provides a well-tested P&L engine: profit curves, probability of
# profit, expected profit/loss, max/min return in a price range.
# We wrap it with a graceful fallback so the script still runs without it.

def analyze_strategy(S: float, legs: list, target_date: str,
                     sigma: float = 0.30, r: float = 0.05) -> dict:
    """
    Evaluate a multi-leg option strategy using optionlab.

    `legs` is a list of dicts, each with keys:
        type       : 'call' or 'put'
        strike     : float
        premium    : float (positive = debit paid, negative = credit received)
        n          : int (number of contracts)
        action     : 'buy' or 'sell'
        expiration : 'YYYY-MM-DD'

    `sigma` : annualised volatility as a decimal (e.g. 0.30 for 30%)
    `r`     : annualised risk-free rate as a decimal (e.g. 0.045 for 4.5%)
    """
    try:
        from optionlab import run_strategy
    except ImportError:
        return {"error": "optionlab not installed — run: pip install optionlab"}

    inputs = {
        "stock_price":    float(S),
        "volatility":     float(sigma),
        "interest_rate":  float(r),
        "start_date":     datetime.now().strftime('%Y-%m-%d'),
        "target_date":    target_date,
        "min_stock":      S * 0.60,
        "max_stock":      S * 1.40,
        "strategy":       legs,
    }
    try:
        out = run_strategy(inputs)
        return {
            "probability_of_profit (%)":        round(out.probability_of_profit * 100, 2),
            "max_return_in_domain ($)":          round(out.maximum_return_in_the_domain, 2),
            "min_return_in_domain ($)":          round(out.minimum_return_in_the_domain, 2),
            "expected_profit_if_profitable ($)": round(out.expected_profit_if_profitable, 2),
            "expected_loss_if_loss ($)":         round(out.expected_loss_if_unprofitable, 2),
            "strategy_cost ($)":                 round(out.strategy_cost, 2),
        }
    except Exception as e:
        return {"error": str(e)}


# ─── Buying Power + Commissions  (OptionSuite-inspired) ───────────────────────
# OptionSuite's strangle.py implements the CBOE margin formula and TastyWorks
# commission structure. We port both as standalone functions so they work with
# our yfinance-based live data without needing the full event-driven framework.

# Commission schedule (TastyWorks structure, per OptionSuite pricingConfig.json)
_COMMISSIONS = {
    'open':  {'per_contract': 1.00, 'per_leg': 0.00},   # $1.00/contract open
    'close': {'per_contract': 0.00, 'per_leg': 0.00},   # free closing at TW
    'regulatory_fee': 0.03,                               # OCC + FINRA/leg
}


def buying_power_strangle(S: float, call_strike: float, put_strike: float,
                           call_price: float, put_price: float,
                           n_contracts: int = 1, multiplier: int = 100) -> dict:
    """
    CBOE margin requirement for a short strangle (cash-settled index style).

    OptionSuite source: optionPrimitives/strangle.py → getBuyingPower()

    Two methods defined by CBOE; the larger one is the actual requirement:
      Method 1 (20% rule): 20% × underlying − OTM_amount + premium
      Method 2 (10% rule): 10% × strike + premium

    Args:
        S            : current underlying price
        call_strike  : call strike price
        put_strike   : put strike price
        call_price   : call mid price
        put_price    : put mid price
        n_contracts  : number of strangles
        multiplier   : contract multiplier (100 for equity options)

    Returns dict with both methods and final required buying power.
    """
    S = float(S)

    # Method 1 — 20% rule
    call_bp1 = (0.20 * S - max(0, call_strike - S) + call_price) * n_contracts * multiplier
    put_bp1  = (0.20 * S - max(0, S - put_strike) + put_price)  * n_contracts * multiplier
    method1  = max(call_bp1, put_bp1)

    # Method 2 — 10% rule
    call_bp2 = (0.10 * call_strike + call_price) * n_contracts * multiplier
    put_bp2  = (0.10 * put_strike  + put_price)  * n_contracts * multiplier
    method2  = max(call_bp2, put_bp2)

    required = max(method1, method2)

    return {
        'buying_power_method1 ($)': round(method1, 2),
        'buying_power_method2 ($)': round(method2, 2),
        'buying_power_required ($)': round(required, 2),
        'credit_received ($)':       round((call_price + put_price) * n_contracts * multiplier, 2),
        'max_loss_estimate ($)':     round(required - (call_price + put_price) * n_contracts * multiplier, 2),
    }


def buying_power_put_vertical(S: float, long_strike: float, short_strike: float,
                               net_credit: float, n_contracts: int = 1,
                               multiplier: int = 100) -> dict:
    """
    Buying power for a short put vertical (bull put spread).

    OptionSuite source: optionPrimitives/putVertical.py → getBuyingPower()

    Max loss = spread width − net credit received.
    Buying power = (short_strike − long_strike − net_credit) × contracts × multiplier
    """
    spread_width = short_strike - long_strike
    max_loss     = (spread_width - net_credit) * n_contracts * multiplier
    return {
        'spread_width ($)':          round(spread_width * multiplier, 2),
        'credit_received ($)':       round(net_credit * n_contracts * multiplier, 2),
        'buying_power_required ($)': round(max_loss, 2),
        'max_loss ($)':              round(max_loss, 2),
        'max_profit ($)':            round(net_credit * n_contracts * multiplier, 2),
        'return_on_risk (%)':        round(net_credit / max(spread_width - net_credit, 0.01) * 100, 2),
    }


def commission_cost(n_legs: int, n_contracts: int, open_or_close: str = 'open') -> float:
    """
    Round-trip commission estimate (TastyWorks structure from OptionSuite).

    OptionSuite source: optionPrimitives/strangle.py → getCommissionsAndFees()

    Args:
        n_legs       : number of option legs (1 = naked, 2 = strangle/spread, 4 = condor)
        n_contracts  : contracts per leg
        open_or_close: 'open' or 'close'
    """
    cfg = _COMMISSIONS[open_or_close]
    base_comm = cfg['per_contract'] * n_contracts * n_legs
    reg_fee   = _COMMISSIONS['regulatory_fee'] * n_legs
    return round(base_comm + reg_fee, 2)


def find_by_delta(chain_df: pd.DataFrame, target_delta: float,
                  option_type: str = 'put', delta_tol: float = 0.10) -> pd.Series:
    """
    Select the option row closest to a target delta within a tolerance band.

    OptionSuite source: strategyManager/StrangleStrat.py → __updateWithOptimalOption()

    This is how short-premium traders enter: 'sell the -0.25 delta put', not
    'sell the $X strike'. Works on any DataFrame produced by analyze_chain().

    Args:
        chain_df     : DataFrame from analyze_chain()
        target_delta : e.g. -0.25 for puts, 0.25 for calls
        option_type  : 'call' or 'put'
        delta_tol    : ± window around target_delta to search within
    """
    col = option_type.upper()
    sub = chain_df[chain_df['Type'] == col].copy()
    sub = sub[sub['Delta'].notna()]
    lo, hi = target_delta - delta_tol, target_delta + delta_tol
    band = sub[(sub['Delta'] >= lo) & (sub['Delta'] <= hi)]
    if band.empty:
        band = sub   # fall back to full chain if band is empty
    best_idx = (band['Delta'] - target_delta).abs().idxmin()
    return band.loc[best_idx]


class RiskManager:
    """
    Mechanical exit rule checker for a short premium position.

    OptionSuite source: riskManager/strangleRiskManagement.py → managePosition()

    Supported rules (can combine):
        '50pct'       — close when P&L ≥ 50% of max credit received
        '21dte'       — close when DTE ≤ 21
        'halfloss'    — close when P&L ≤ −100% of max credit (i.e. lost 1× premium)

    Usage:
        rm = RiskManager(['50pct', '21dte'])
        should_close, reason = rm.check(current_pnl, max_credit, dte_remaining)
    """

    RULES = {'50pct', '21dte', 'halfloss', 'hold'}

    def __init__(self, rules: list):
        bad = set(rules) - self.RULES
        if bad:
            raise ValueError(f"Unknown rules: {bad}. Valid: {self.RULES}")
        self.rules = rules

    def check(self, current_pnl: float, max_credit: float,
              dte_remaining: int) -> tuple:
        """
        Returns (should_close: bool, reason: str).

        Args:
            current_pnl   : current P&L in dollars (positive = profit)
            max_credit     : total credit received when trade was opened
            dte_remaining  : calendar days left to expiration
        """
        if 'hold' in self.rules:
            if dte_remaining <= 1:
                return True, 'expiration'
            return False, ''

        pnl_pct = (current_pnl / max_credit * 100) if max_credit > 0 else 0

        if '50pct' in self.rules and pnl_pct >= 50:
            return True, '50% profit target hit'
        if '21dte' in self.rules and dte_remaining <= 21:
            return True, '21 DTE reached'
        if 'halfloss' in self.rules and pnl_pct <= -100:
            return True, 'full-premium loss stop'
        if dte_remaining <= 1:
            return True, 'expiration'

        return False, ''



# ─── Excel Export ─────────────────────────────────────────────────────────────

def _autofit(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 3, 30)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')


def export_excel(symbol, S, q, hv, all_data, skew_df, parity_df,
                 csp_df, cc_df, filename):
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        summary = pd.DataFrame([{
            'Symbol':                 symbol,
            'Spot Price ($)':         round(S, 2),
            'Dividend Yield (%)':     round(q * 100, 2),
            'Historical Vol (%)':     round(hv * 100, 2),
            'Risk-Free Rate Note':    'Term-matched from yield curve',
            'Analysis Date':          datetime.now().strftime('%Y-%m-%d %H:%M'),
            'Total Options Analyzed': len(all_data),
            'Expiries Analyzed':      all_data['Expiry'].nunique(),
        }])
        summary.T.to_excel(writer, sheet_name='Summary', header=False)
        all_data.to_excel(writer, index=False, sheet_name='Full Chain')
        all_data[all_data['Type'] == 'CALL'].to_excel(writer, index=False, sheet_name='Calls')
        all_data[all_data['Type'] == 'PUT'].to_excel(writer, index=False, sheet_name='Puts')
        skew_df.to_excel(writer, index=False, sheet_name='Vol Skew')
        parity_df.to_excel(writer, index=False, sheet_name='Parity Check')
        if not csp_df.empty:
            csp_df.to_excel(writer, index=False, sheet_name='CSP Screener')
        if not cc_df.empty:
            cc_df.to_excel(writer, index=False, sheet_name='CC Screener')
        for ws in writer.sheets.values():
            _autofit(ws)


# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    symbol = input("Enter stock symbol: ").strip().upper()

    print(f"\n[1/6] Spot price & dividend yield for {symbol}...")
    hist = _ticker(symbol).history(period='2d')
    if hist.empty:
        print(f"Error: no data for {symbol}")
        return
    S = hist['Close'].iloc[-1]
    q = get_dividend_yield(symbol)
    print(f"  Spot: ${S:.2f}  |  Div yield: {q*100:.2f}%")

    print("[2/6] Building Treasury yield curve...")
    # Force build once now so the per-expiry calls hit cache
    global _yield_curve_fn
    _yield_curve_fn = _build_yield_curve()
    r_short = get_risk_free_rate(0.25)
    r_long  = get_risk_free_rate(2.0)
    print(f"  3-month rate: {r_short*100:.3f}%  |  2-year rate: {r_long*100:.3f}%")

    print("[3/6] Computing 1-year historical volatility...")
    hv = historical_volatility(symbol)
    print(f"  HV: {hv*100:.2f}%")

    # Try CBOE direct feed first (OpenBB-revealed endpoint, no API key)
    print("[4/6] Fetching option chain  (CBOE → yfinance fallback)...")
    cboe_df = fetch_cboe_chain(symbol)
    if not cboe_df.empty:
        print(f"  CBOE direct: {len(cboe_df)} contracts (15-min delayed exchange data)")
    else:
        print("  CBOE unavailable — using yfinance")

    expirations = _ticker(symbol).options
    if not expirations:
        print("No options available.")
        return
    expirations = expirations[:8]
    print(f"  Analysing {len(expirations)} expiry dates (full BS chain)")

    all_dfs = []
    import time
    for i, exp in enumerate(expirations, 1):
        try:
            print(f"  [{i}/{len(expirations)}] {exp} ...", end=' ', flush=True)
            df = analyze_chain(symbol, exp, S, hv, q)
            all_dfs.append(df)
            print(f"{len(df)} contracts  "
                  f"(early exercise premium range: "
                  f"${df['Early Exercise Prem ($)'].dropna().min():.3f}–"
                  f"${df['Early Exercise Prem ($)'].dropna().max():.3f})")
            time.sleep(0.5) # Prevent rate limiting
        except Exception as e:
            print(f"skipped ({e})")

    if not all_dfs:
        print("No data retrieved.")
        return

    all_data  = pd.concat(all_dfs, ignore_index=True)
    skew_df   = vol_skew(all_dfs[0])
    parity_df = parity_check(all_dfs[0], S)

    print("[5/6] Running income + strategy screeners...")
    csp_df = cash_secured_put_screener(symbol, max_breakeven=S * 0.95)
    cc_df  = covered_call_screener(symbol, S, min_breakeven=S * 1.02)
    if not csp_df.empty:
        print(f"  Top CSP: Strike {csp_df.iloc[0]['Strike']}  ARR {csp_df.iloc[0]['ARR (%)']}%")
    if not cc_df.empty:
        print(f"  Top CC:  Strike {cc_df.iloc[0]['Strike']}  ARR {cc_df.iloc[0]['ARR (%)']}%")

    # Iron condor from nearest expiry (Lean-pattern construction)
    condor = find_iron_condor(all_dfs[0], S, wing_delta=0.15, width=5)
    if condor:
        print(f"  Iron Condor ({condor['expiry']}): "
              f"{condor['long_put_strike']}P/{condor['short_put_strike']}P / "
              f"{condor['short_call_strike']}C/{condor['long_call_strike']}C  "
              f"Credit: ${condor['net_credit ($)']:.2f}  "
              f"RoR: {condor['return_on_risk (%)']:.1f}%  "
              f"BE: [{condor['breakeven_low']} – {condor['breakeven_high']}]")

    # OptionSuite: delta-targeted entry + buying power for nearest expiry
    print("[6/6] OptionSuite margin + delta analysis (nearest expiry)...")
    nearest = all_dfs[0]
    try:
        call_entry = find_by_delta(nearest, target_delta=0.25,  option_type='call')
        put_entry  = find_by_delta(nearest, target_delta=-0.25, option_type='put')
        bp = buying_power_strangle(
            S,
            call_strike=call_entry['Strike'], put_strike=put_entry['Strike'],
            call_price=call_entry['Mid Price'], put_price=put_entry['Mid Price'],
        )
        open_comm  = commission_cost(n_legs=2, n_contracts=1, open_or_close='open')
        close_comm = commission_cost(n_legs=2, n_contracts=1, open_or_close='close')
        net_credit = bp['credit_received ($)'] - open_comm - close_comm

        print(f"  -0.25Δ strangle: {put_entry['Strike']}P / {call_entry['Strike']}C")
        print(f"  Gross credit:    ${bp['credit_received ($)']:.2f}")
        print(f"  Commissions:     ${open_comm + close_comm:.2f} (open+close)")
        print(f"  Net credit:      ${net_credit:.2f}")
        print(f"  Buying power:    ${bp['buying_power_required ($)']:.2f}  "
              f"(return on BP: {net_credit/bp['buying_power_required ($)']*100:.1f}%)")

        rm = RiskManager(['50pct', '21dte'])
        _, exit_rule = rm.check(0, net_credit, int(put_entry['T (years)'] * 365))
        print(f"  Exit rules:      50% profit OR 21 DTE  (current trigger: '{exit_rule or 'none'}')")
    except Exception as e:
        print(f"  Margin analysis skipped: {e}")

    print("\nSaving Excel report...")
    n = sum(1 for f in os.listdir('.') if f.startswith(f"{symbol}_options_v3"))
    filename = f"{symbol}_options_v3_{n + 1}.xlsx"
    export_excel(symbol, S, q, hv, all_data, skew_df, parity_df, csp_df, cc_df, filename)
    print(f"Saved: {filename}")

    # ATM vol summary + quant vol panel
    atm = all_data[all_data['Moneyness'] == 'ATM']
    if not atm.empty:
        atm_iv = atm['IV (%)'].mean()
        spread = atm_iv - hv * 100
        signal = 'vol sellers edge' if spread > 0 else 'vol buyers edge'
        print(f"\nATM IV: {atm_iv:.1f}%  |  HV: {hv*100:.1f}%  |  IV Premium: {spread:+.1f}%  ({signal})")

        # GARCH(1,1) vol forecast alongside HAR-RV (reuse cached history)
        try:
            _hist_px  = list(_ticker(symbol).history(period='1y')['Close'].dropna())
            _log_rets = [np.log(_hist_px[k]/_hist_px[k-1]) for k in range(1, len(_hist_px))]
            _garch    = garch_vol_forecast(_log_rets, horizon=21)
            if _garch:
                _g_vs_iv = _garch['vol'] * 100 - atm_iv
                _ge = ('vol rising → buyers edge' if _g_vs_iv > 2
                       else 'vol falling → sellers edge' if _g_vs_iv < -2
                       else 'vol fair')
                print(f"GARCH(1,1) 21d forecast: {_garch['vol']*100:.1f}%  "
                      f"(persistence={_garch['persistence']:.3f})  → {_ge}")
        except Exception:
            pass

        # SABR smile fit on nearest expiry
        try:
            _near  = all_dfs[0]
            _F     = S * np.exp((get_risk_free_rate(all_dfs[0]['T (years)'].iloc[0]) - q)
                                * all_dfs[0]['T (years)'].iloc[0])
            _T_s   = float(_near['T (years)'].iloc[0])
            _sabr  = fit_sabr_from_chain(_near, _F, _T_s)
            if _sabr:
                _skew_dir = 'put skew (bearish)' if _sabr['rho'] < -0.2 else ('call skew (bullish)' if _sabr['rho'] > 0.2 else 'balanced smile')
                print(f"SABR smile fit:  α={_sabr['alpha']:.3f}  ρ={_sabr['rho']:+.3f}  ν={_sabr['nu']:.3f}  "
                      f"RMSE={_sabr['fit_rmse']:.2f}%  → {_skew_dir}")
        except Exception:
            pass

        # Pin risk for nearest expiry (reuse all_dfs[0] — no extra API call)
        try:
            _combined  = all_dfs[0].copy()
            _combined.rename(columns={'Open Interest': 'Open Interest', 'Strike': 'Strike'}, inplace=True)
            _pin = pin_risk_score(S, _combined[['Strike','Open Interest']])
            if _pin:
                _pstr = 'STRONG' if _pin['pin_strength'] > 0.4 else 'MODERATE' if _pin['pin_strength'] > 0.2 else 'WEAK'
                print(f"Pin risk (nearest expiry):  ${_pin['pin_strike']:.2f}  "
                      f"[{_pstr}, {_pin['distance_pct']:.1f}% from spot]"
                      + ("   near pin" if _pin['distance_pct'] < 3 else ""))
        except Exception:
            pass

    # Optional: run a strategy example with optionlab
    print("\n[Optional] Enter strategy legs to evaluate P&L + PoP (or press Enter to skip):")
    ans = input("  Run strategy analysis? [y/N]: ").strip().lower()
    if ans == 'y':
        print("  Example: long ATM call for nearest expiry")
        atm_call = all_data[(all_data['Type'] == 'CALL') & (all_data['Moneyness'] == 'ATM')]
        if not atm_call.empty:
            row   = atm_call.iloc[0]
            sigma = (row['IV (%)'] / 100) if not np.isnan(row['IV (%)']) else hv
            r_leg = get_risk_free_rate(row['T (years)'])
            legs  = [{
                "type":       "call",
                "strike":     float(row['Strike']),
                "premium":    float(row['Mid Price']),
                "n":          1,
                "action":     "buy",
                "expiration": row['Expiry'],
            }]
            result = analyze_strategy(S, legs, row['Expiry'], sigma=sigma, r=r_leg)
            print("\n  Strategy Results:")
            for k, v in result.items():
                print(f"    {k}: {v}")

    # Optional Colab download
    try:
        from google.colab import files
        files.download(filename)
    except ImportError:
        pass

    # Auto-launch Trade Finder for the same stock
    print("\n" + "─"*60)
    run_tf = input(f"  Run Trade Finder for {symbol} now? [y/N]: ").strip().lower()
    if run_tf == 'y':
        find_trade(symbol=symbol, S=S, hv=hv, q=q)


# ─── GEX Signal  (auto-suggest Call vs Put from Gamma Exposure) ───────────────

def gex_signal(symbol: str, S: float, r: float, hv: float, q: float = 0.0) -> dict | None:
    """
    Calculate net Gamma Exposure (GEX) from the nearest *usable* expiry chain.

    Convention (SqueezeMetrics "retail-flow" / index assumption):
    Retail are net call BUYERS and (less consistently) put BUYERS, so dealers
    are net SHORT calls and net SHORT puts. Under that sign convention:
      Call GEX = OI × Γ × 100 × S   (treated POSITIVE for net-gamma sum)
      Put  GEX = OI × Γ × 100 × S   (treated NEGATIVE for net-gamma sum)
      Net GEX  = Σ Call GEX − Σ Put GEX

    Positive net GEX → dealers long gamma → suppress vol (mean-revert / pin)
    Negative net GEX → dealers short gamma → amplify moves (trend / breakout)

    LIMITATION (A8): this sign convention is empirically valid for SPX/NDX/SPY
    index options but breaks for single names and 0DTE in stress regimes —
    e.g. meme stocks (GME, MSTR) where retail buys both sides aggressively,
    and 0DTE-heavy days where the call/put taker mix flips intraday. Modern
    flow vendors (Glassnode "Taker-Flow-Based GEX", SpotGamma, vol.land) now
    INFER dealer side from CBOE COT-style or per-trade taker data rather than
    assuming it. Without that data here we publish two additional readings —
    `abs_call_wall` and `abs_put_wall` — which give the strikes of maximum
    |GEX| regardless of sign. Use those when dealer side is uncertain.

    Production upgrade path:
      • Plug an OCC/CBOE COT or vol.land "dealer-positioning" feed into the
        loop below and have it return signed (call_oi_buy − call_oi_sell) at
        each strike, then drop the assumed sign convention entirely.

    GEX flip level (zero-gamma level): the price where cumulative GEX crosses zero.
    Acts as a key support/resistance pivot. Price above = bullish structure; below = bearish.

    Call wall: strike with the largest call GEX concentration → ceiling/resistance.
    Put wall:  strike with the largest put  GEX concentration → floor/support.

    Tries up to 4 expiries, skipping those with < 3 DTE (0-DTE chains have near-zero
    T which breaks gamma / IV math) or too few OI-bearing contracts.
    """
    MIN_DTE       = 3    # skip expiries closer than this — gamma math degenerates
    MIN_CONTRACTS = 4    # need at least this many OI-bearing strikes to trust GEX
    MAX_TRIES     = 8    # scan up to this many expiries looking for a usable chain

    expirations = _ticker(symbol).options
    if not expirations:
        print("  (GEX: no option expiries found)")
        return None

    today = datetime.now().replace(tzinfo=None)

    # Scan forward through expiries until one has enough liquid strikes
    call_gex: dict[float, float] = {}
    put_gex:  dict[float, float] = {}
    chosen_exp = None
    chosen_T   = None

    for exp in expirations[:MAX_TRIES]:
        exp_dt = datetime.strptime(exp, '%Y-%m-%d')
        dte    = (exp_dt - today).days
        if dte < MIN_DTE:
            continue                          # skip same-week 0/1/2 DTE

        T_try = max(dte / 365.0, MIN_DTE / 365.0)

        try:
            chain = _ticker(symbol).option_chain(exp)
        except Exception:
            continue

        cg: dict[float, float] = {}
        pg: dict[float, float] = {}

        for _, row in chain.calls.iterrows():
            K  = row['strike']
            oi = row.get('openInterest', 0) or 0
            if oi == 0:
                continue
            price = mid_price(row)
            if price <= 0:
                continue
            iv    = implied_volatility(S, K, T_try, r, price, 'call', q)
            sig   = iv if (not np.isnan(iv) and iv > 0) else hv
            gamma = bs_greeks(S, K, T_try, r, sig, 'call', q)['gamma']
            if np.isnan(gamma):
                continue
            cg[K] = oi * gamma * 100 * S

        for _, row in chain.puts.iterrows():
            K  = row['strike']
            oi = row.get('openInterest', 0) or 0
            if oi == 0:
                continue
            price = mid_price(row)
            if price <= 0:
                continue
            iv    = implied_volatility(S, K, T_try, r, price, 'put', q)
            sig   = iv if (not np.isnan(iv) and iv > 0) else hv
            gamma = bs_greeks(S, K, T_try, r, sig, 'put', q)['gamma']
            if np.isnan(gamma):
                continue
            pg[K] = oi * gamma * 100 * S

        if len(cg) + len(pg) >= MIN_CONTRACTS:
            call_gex   = cg
            put_gex    = pg
            chosen_exp = exp
            chosen_T   = T_try
            break
        # too thin — try the next expiry silently

    if chosen_exp is None:
        print(f"  (GEX: no expiry in the first {MAX_TRIES} had ≥{MIN_CONTRACTS} liquid strikes)")
        return None

    all_strikes = sorted(set(call_gex) | set(put_gex))
    net_by_k    = {k: call_gex.get(k, 0) - put_gex.get(k, 0) for k in all_strikes}

    total_call_gex = sum(call_gex.values())
    total_put_gex  = sum(put_gex.values())
    net_gex        = total_call_gex - total_put_gex

    # GEX flip level: interpolated zero-crossing of cumulative net GEX
    cum      = 0.0
    prev_cum = 0.0
    prev_k   = all_strikes[0]
    flip_level = None
    for k in all_strikes:
        prev_cum = cum
        cum += net_by_k[k]
        if prev_cum * cum < 0:  # sign change — flip is between prev_k and k
            span = cum - prev_cum
            flip_level = prev_k + (k - prev_k) * (-prev_cum / span) if span != 0 else k
            break
        prev_k = k
    if flip_level is None:
        flip_level = all_strikes[len(all_strikes) // 2]  # mid-chain fallback

    call_wall = max(call_gex, key=call_gex.get) if call_gex else S * 1.05
    put_wall  = max(put_gex,  key=put_gex.get)  if put_gex  else S * 0.95

    # Secondary walls: 2nd-largest GEX concentration at a different strike
    _sc = sorted(call_gex, key=call_gex.get, reverse=True)
    _sp = sorted(put_gex,  key=put_gex.get,  reverse=True)
    secondary_call_wall = _sc[1] if len(_sc) > 1 else None
    secondary_put_wall  = _sp[1] if len(_sp)  > 1 else None

    # Direction: price vs flip level is the primary signal
    if S > flip_level:
        suggested = 'call'
        reason = f"price ${S:.2f} is ABOVE GEX flip ${flip_level:.2f} → bullish structure"
    else:
        suggested = 'put'
        reason = f"price ${S:.2f} is BELOW GEX flip ${flip_level:.2f} → bearish structure"

    # Override near walls (within 1.5%): call wall = ceiling, put wall = floor
    if abs(S - call_wall) / S < 0.015:
        suggested = 'put'
        reason = f"price within 1.5% of call wall ${call_wall:.2f} (resistance — expect cap)"
    elif abs(S - put_wall) / S < 0.015:
        suggested = 'call'
        reason = f"price within 1.5% of put wall ${put_wall:.2f} (support — expect bounce)"

    regime       = 'POSITIVE' if net_gex > 0 else 'NEGATIVE'
    regime_label = 'pin/mean-revert (selling edge)' if net_gex > 0 else 'trend/breakout (buying edge)'

    # Sign-agnostic walls (A8 fallback): strikes with maximum |GEX| on each
    # side regardless of assumed dealer direction. Useful for single-name
    # tickers where retail/dealer split is uncertain.
    _abs = {k: abs(call_gex.get(k, 0)) + abs(put_gex.get(k, 0)) for k in all_strikes}
    abs_top = sorted(_abs, key=_abs.get, reverse=True)
    abs_call_wall = next((k for k in abs_top if k > S), None)
    abs_put_wall  = next((k for k in abs_top if k < S), None)

    return {
        'net_gex':        net_gex,
        'regime':         regime,
        'regime_label':   regime_label,
        'flip_level':     round(flip_level, 2),
        'call_wall':      call_wall,
        'put_wall':       put_wall,
        'abs_call_wall':       abs_call_wall,
        'abs_put_wall':        abs_put_wall,
        'suggested':           suggested,
        'reason':              reason,
        'expiry_used':         chosen_exp,
        'secondary_call_wall': secondary_call_wall,
        'secondary_put_wall':  secondary_put_wall,
    }


# ─── Smart Exit Levels  (GEX + Greeks + Mathematics) ─────────────────────────

def smart_exit_levels(S: float, K: float, T: float, r: float, sigma: float,
                      option_type: str, premium_paid: float, gex: dict,
                      skew_slope: float = 0.30) -> dict:
    """
    GEX-based + Greeks-based optimal profit-taking calculator for a long option.

    Mathematical framework (published research):
    ─ Bollen & Whaley (2004, J.Finance): dealer delta-hedging at GEX walls creates
      measurable price resistance/support. Call wall = dealers sell hedge; put wall =
      dealers buy hedge. In negative GEX the effect is amplified, not dampened.
    ─ Black & Scholes (1973): residual option value at each GEX spot level computed
      with sticky-delta IV adjustment: σ'(S') = σ × (K/S')^skew_slope. The slope
      should be FIT from the actual chain (Gatheral 2006); a hard-coded 0.30 is
      wrong by 3–5 vol points for high-skew names. Caller passes `skew_slope`
      derived from a least-squares fit of log(IV) vs log(K/S) around ATM.
    ─ Median first-passage time of drift-free GBM (Karatzas-Shreve 1991):
      T_½ = 2.198 × (ln(L/S)/σ)² × 252  trading days.
    ─ Theta/price decay rate (Corrado & Miller 1996 framework):
      |Θ|/V > 1.5%/day → theta dominates; hold no longer justified by carry.
    ─ Kelly (1956) partial-exit fraction:
      f* = (p × b − q) / b  where p = P(next level), b = upside ratio.
      Exit (1 − f*) of position at current level; hold f* for next target.
    ─ Breeden & Litzenberger (1978): risk-neutral P(reaching level) via digital
      option approximation: P(S_T ≥ L) ≈ N(d2) evaluated at L as pseudo-strike.
    """
    if not gex:
        return {}

    flip      = gex.get('flip_level', S)
    call_wall = gex.get('call_wall', S * 1.05)
    put_wall  = gex.get('put_wall',  S * 0.95)
    sec_call  = gex.get('secondary_call_wall')
    sec_put   = gex.get('secondary_put_wall')
    regime    = gex.get('regime', 'UNKNOWN')
    neg_gex   = (regime == 'NEGATIVE')

    current_val  = bs_price(S, K, T, r, sigma, option_type)
    g0           = bs_greeks(S, K, T, r, sigma, option_type)
    theta0       = g0['theta']
    daily_1sigma = S * sigma * np.sqrt(1.0 / 252.0)
    intrinsic0   = max(0.0, S - K) if option_type == 'call' else max(0.0, K - S)
    time_val0    = max(0.0, (current_val or 0) - intrinsic0)

    def _days(target_S: float) -> float:
        """
        Median first-passage time of drift-free GBM to a log-barrier,
        in trading days. For BM with annualised vol σ the first hit time
        to a level a in log space follows a Lévy distribution; its
        *median* is 2.198 × (a/σ)² years (Karatzas-Shreve 1991, §2.6).
        The deterministic |L−S|/σ_day formula previously used here treats
        price as travelling in a straight line and understates the actual
        time-to-hit by roughly 2–5×.
        """
        if target_S <= 0 or sigma <= 0 or target_S == S:
            return np.nan
        log_d = abs(np.log(target_S / S))
        return float(2.198 * (log_d / sigma) ** 2 * 252)

    def _val(target_S: float, days_elapsed: float) -> float:
        """BS value at target spot with sticky-delta IV skew adjustment.
        Uses the per-symbol skew_slope passed in by the caller (default 0.30).
        """
        if target_S <= 0:
            return np.nan
        T_rem     = max(T - days_elapsed / 365.0, 0.5 / 365.0)
        skew_adj  = float(np.clip((K / target_S) ** skew_slope, 0.60, 2.00))
        return bs_price(target_S, K, T_rem, r, sigma * skew_adj, option_type)

    def _pnl(target_S: float, days_elapsed: float) -> float:
        v = _val(target_S, days_elapsed)
        return np.nan if np.isnan(v) else (v - premium_paid) * 100

    def _prob(target_S: float) -> float:
        """Breeden-Litzenberger digital-option P(S_T ≥ target_S)."""
        if T <= 0 or sigma <= 0 or target_S <= 0:
            return 0.0
        d2 = (np.log(S / target_S) + (r - 0.5 * sigma**2) * T) / (sigma * np.sqrt(T))
        return float(norm.cdf(d2))

    def _kelly_exit(p_cont: float, v_now: float, v_next: float) -> float:
        """
        Fraction to EXIT now (1 − f*).
        f* = Kelly fraction to HOLD for the next level.
        """
        if v_now <= 0 or v_next <= v_now:
            return 1.0
        b = (v_next - v_now) / v_now   # upside ratio to next target
        q = 1 - p_cont
        f = max(0.0, (p_cont * b - q) / b)
        return float(np.clip(1.0 - f, 0.25, 1.0))

    levels = []

    if option_type == 'put':
        # ── TP1: Put Wall ─────────────────────────────────────────────────────
        if put_wall and put_wall < S:
            d, v, p, pr = _days(put_wall), _val(put_wall, _days(put_wall)), \
                          _pnl(put_wall, _days(put_wall)), _prob(put_wall)
            v2 = _val(sec_put, _days(sec_put)) if sec_put and sec_put < put_wall else None
            p_cont = _prob(sec_put) / max(pr, 1e-6) if (sec_put and sec_put < put_wall and pr > 0) else 0.2
            kf = _kelly_exit(p_cont, v or 0, v2 or 0) if v2 else 0.65
            if neg_gex:
                note = "Neg GEX — wall is weak. Dealers not defending. Take 30–40% here, hold rest."
                action = f'PARTIAL EXIT (~{kf*100:.0f}%)'
            else:
                note = "Pos GEX — dealers buy-to-hedge here (strong floor). Likely bounces. Take 60–70%."
                action = f'MAJOR EXIT (~{kf*100:.0f}%)'
            levels.append({'name': 'TP1  Put Wall', 'spot': put_wall,
                           'days': d, 'opt_val': v, 'pnl': p, 'prob': pr,
                           'action': action, 'kelly_exit': kf, 'note': note})

        # ── TP2: Secondary Put Wall ───────────────────────────────────────────
        if sec_put and put_wall and sec_put < put_wall:
            d, v, p, pr = _days(sec_put), _val(sec_put, _days(sec_put)), \
                          _pnl(sec_put, _days(sec_put)), _prob(sec_put)
            levels.append({'name': 'TP2  Secondary Put Wall', 'spot': sec_put,
                           'days': d, 'opt_val': v, 'pnl': p, 'prob': pr,
                           'action': 'EXIT REMAINING 100%',
                           'kelly_exit': 1.0,
                           'note': 'Deeper floor. Exit fully — IV crush risk on oversold bounce.'})

        # ── Flip: regime-change warning or stop ───────────────────────────────
        if flip and flip < S:
            d, v, p = _days(flip), _val(flip, _days(flip)), _pnl(flip, _days(flip))
            levels.append({'name': '  GEX Flip — REGIME CHANGE', 'spot': flip,
                           'days': d, 'opt_val': v, 'pnl': p, 'prob': _prob(flip),
                           'action': 'EXIT ALL — negative GEX ends below flip',
                           'kelly_exit': 1.0,
                           'note': 'Below flip dealer amplification stops. Puts lose momentum. Exit all.'})
        elif flip and flip > S:
            d, v, p = _days(flip), _val(flip, _days(flip)), _pnl(flip, _days(flip))
            levels.append({'name': '  STOP  GEX Flip (above price)', 'spot': flip,
                           'days': d, 'opt_val': v, 'pnl': p, 'prob': np.nan,
                           'action': 'STOP LOSS — exit puts if price rallies to flip',
                           'kelly_exit': 1.0,
                           'note': 'Above flip → bullish regime. Put thesis invalidated.'})

    else:  # call
        # ── TP1: Call Wall ────────────────────────────────────────────────────
        if call_wall and call_wall > S:
            d, v, p, pr = _days(call_wall), _val(call_wall, _days(call_wall)), \
                          _pnl(call_wall, _days(call_wall)), _prob(call_wall)
            v2 = _val(sec_call, _days(sec_call)) if sec_call and sec_call > call_wall else None
            p_cont = _prob(sec_call) / max(pr, 1e-6) if (sec_call and sec_call > call_wall and pr > 0) else 0.2
            kf = _kelly_exit(p_cont, v or 0, v2 or 0) if v2 else 0.65
            if neg_gex:
                note = "Neg GEX — wall resistance is weak. Price may push through. Take 30–40% here."
                action = f'PARTIAL EXIT (~{kf*100:.0f}%)'
            else:
                note = "Pos GEX — dealers sell-to-hedge hard at call wall. Strong ceiling. Take 60–70%."
                action = f'MAJOR EXIT (~{kf*100:.0f}%)'
            levels.append({'name': 'TP1  Call Wall', 'spot': call_wall,
                           'days': d, 'opt_val': v, 'pnl': p, 'prob': pr,
                           'action': action, 'kelly_exit': kf, 'note': note})

        # ── TP2: Secondary Call Wall ──────────────────────────────────────────
        if sec_call and call_wall and sec_call > call_wall:
            d, v, p, pr = _days(sec_call), _val(sec_call, _days(sec_call)), \
                          _pnl(sec_call, _days(sec_call)), _prob(sec_call)
            levels.append({'name': 'TP2  Secondary Call Wall', 'spot': sec_call,
                           'days': d, 'opt_val': v, 'pnl': p, 'prob': pr,
                           'action': 'EXIT REMAINING 100%',
                           'kelly_exit': 1.0,
                           'note': 'Upper resistance. Exit fully — IV crush risk on overbought reversal.'})

        # ── Flip ──────────────────────────────────────────────────────────────
        if flip and flip > S:
            d, v, p = _days(flip), _val(flip, _days(flip)), _pnl(flip, _days(flip))
            levels.append({'name': '  GEX Flip — REGIME CHANGE', 'spot': flip,
                           'days': d, 'opt_val': v, 'pnl': p, 'prob': _prob(flip),
                           'action': 'EXIT ALL — regime turns pinning above flip',
                           'kelly_exit': 1.0,
                           'note': 'Above flip dealers pin price. Call momentum stalls. Exit all.'})
        elif flip and flip < S:
            d, v, p = _days(flip), _val(flip, _days(flip)), _pnl(flip, _days(flip))
            levels.append({'name': '  STOP  GEX Flip (below price)', 'spot': flip,
                           'days': d, 'opt_val': v, 'pnl': p, 'prob': np.nan,
                           'action': 'STOP LOSS — exit calls if price falls to flip',
                           'kelly_exit': 1.0,
                           'note': 'Below flip → bearish regime. Call thesis invalidated.'})

    # ── Theta danger ──────────────────────────────────────────────────────────
    theta_pct  = abs(theta0) / max(current_val or 0.001, 0.001) * 100
    tv_death   = time_val0 / abs(theta0) if (theta0 and theta0 != 0) else np.nan
    theta_exit = max(int(T * 365 - (tv_death or 0) * 0.5), 7)

    return {
        'levels':         levels,
        'current_val':    round(current_val, 4) if current_val and not np.isnan(current_val) else None,
        'premium_paid':   premium_paid,
        'theta_pct_day':  round(theta_pct, 2),
        'theta_danger':   theta_pct > 1.5,
        'tv_death_days':  round(tv_death, 1) if not np.isnan(tv_death) else None,
        'theta_exit_dte': theta_exit,
        'daily_1sigma':   round(daily_1sigma, 2),
        'regime':         regime,
        'neg_gex':        neg_gex,
    }


# ─── Institutional signal helpers ─────────────────────────────────────────────

def iv_rank_percentile(symbol: str, current_iv: float,
                        period: str = '1y', prices: list = None,
                        asof_index: int = None):
    """
    IV Rank and IV Percentile — the single most-used institutional filter.

    Goldman Sachs, Citadel, and every vol desk screen on this before sizing
    any long-vol trade. IVR < 30 = cheap; IVR > 60 = expensive, don't buy.

    IMPORTANT — HV-as-IV substitution bias:
    Uses rolling 21d HV as a proxy for IV because free Yahoo data has no
    historical IV feed. HV (backward-looking realised) and IV (forward-looking
    implied) diverge sharply around earnings, FOMC, and regime breaks, so the
    output is a *direction-of-vol-regime* indicator, not a precise vol-quote
    rank. For production use, pull MarketChameleon / Barchart / ORATS IV30
    history and feed that here instead of a price series.

    Walk-forward safety:
    Pass `asof_index = i` when running a backtest at historical date prices[i]
    to truncate the price series to [0:i+1]. Without this guard the rolling-
    HV min/max would include the future, leaking ~1-3% of spurious "cheap-vol"
    signal into the score.

    Returns (iv_rank 0-100, iv_percentile 0-100) or (nan, nan) on failure.
    """
    try:
        if prices and len(prices) >= 30:
            hist = pd.Series(prices)
        else:
            hist = _ticker(symbol).history(period=period)['Close'].dropna()
        if asof_index is not None:
            hist = hist.iloc[: asof_index + 1]
        if len(hist) < 30:
            return np.nan, np.nan
        lr      = np.log(hist / hist.shift(1)).dropna()
        hv_roll = lr.rolling(21).std().dropna() * np.sqrt(252)
        if len(hv_roll) < 2:
            return np.nan, np.nan
        lo, hi = float(hv_roll.min()), float(hv_roll.max())
        ivr = (current_iv - lo) / (hi - lo) * 100 if hi > lo else 50.0
        ivp = float((hv_roll < current_iv).mean() * 100)
        return round(ivr, 1), round(ivp, 1)
    except Exception:
        return np.nan, np.nan


def real_world_option_ev(S: float, K: float, T: float, r: float,
                          sigma: float, option_type: str,
                          mu_annual: float, market_price: float,
                          q: float = 0.0) -> float:
    """
    Real-world expected value of an option (Hakansson-Rubinstein P-measure).

    Institutions run this to find positive EV. The key insight:
      - Black-Scholes prices under Q (risk-neutral) measure: drift = r
      - Real-world (P) measure uses actual historical drift = mu
      - If the stock trends, real-world call value > BS price → edge for buyer

    EV > 0 means the option is mathematically cheap given the stock's history.
    EV < 0 means the market is pricing it fairly or expensive.
    """
    if T <= 0 or sigma <= 0 or S <= 0 or K <= 0:
        return np.nan
    mu_adj = mu_annual - q
    d1 = (np.log(S / K) + (mu_adj + 0.5 * sigma**2) * T) / (sigma * np.sqrt(T))
    d2 = d1 - sigma * np.sqrt(T)
    fwd_factor = np.exp((mu_adj - r) * T)
    if option_type == 'call':
        rw_val = S * fwd_factor * norm.cdf(d1) - K * np.exp(-r * T) * norm.cdf(d2)
    else:
        rw_val = K * np.exp(-r * T) * norm.cdf(-d2) - S * fwd_factor * norm.cdf(-d1)
    return round(float(rw_val - market_price), 4)


def earnings_within_dte(symbol: str, dte_max: int):
    """
    Returns (True, 'YYYY-MM-DD') if earnings fall within next dte_max days.

    Institutions always flag this — earnings are a vol event already priced in.
    Buying options before earnings is paying double: the stock IV is inflated
    specifically BECAUSE of the earnings uncertainty. After announcement,
    IV collapses ('vol crush'), killing option value even if you're right on direction.
    """
    try:
        cal = _ticker(symbol).calendar
        if cal is None:
            return False, None
        today = datetime.now().date()
        # yfinance returns calendar as dict with list values
        if isinstance(cal, dict):
            dates = []
            for v in cal.values():
                if isinstance(v, list):
                    dates.extend(v)
                else:
                    dates.append(v)
        elif hasattr(cal, 'columns'):
            dates = list(cal.columns)
        else:
            return False, None
        for d in dates:
            try:
                if hasattr(d, 'date'):
                    dd = d.date()
                else:
                    dd = datetime.strptime(str(d)[:10], '%Y-%m-%d').date()
                days_away = (dd - today).days
                if 0 <= days_away <= dte_max:
                    return True, str(dd)
            except Exception:
                continue
        return False, None
    except Exception:
        return False, None


def compute_rsi(prices: list, period: int = 14) -> float:
    """
    Wilder RSI — momentum position signal used by every institution.

    Institutions use RSI not to time entries but to avoid paying for
    momentum that's already exhausted:
      - Buying a call when RSI > 70 = chasing an overbought move
      - Buying a put when RSI < 30 = buying fear at maximum panic
    Best entries: RSI 40–60 for calls (fresh uptrend), RSI 40–60 for puts (fresh downtrend).
    """
    if len(prices) < period + 1:
        return 50.0
    deltas = [prices[i] - prices[i-1] for i in range(1, len(prices))]
    recent = deltas[-period:]
    gains  = [d if d > 0 else 0.0 for d in recent]
    losses = [-d if d < 0 else 0.0 for d in recent]
    avg_g  = sum(gains) / period
    avg_l  = sum(losses) / period
    if avg_l == 0:
        return 100.0
    return round(100.0 - 100.0 / (1.0 + avg_g / avg_l), 1)


def kelly_fraction_options(p_win: float, win_multiple: float,
                            loss_multiple: float = 1.0) -> float:
    """
    Kelly criterion adapted for options (Thorp 1969; Bouchaud-Potters fat-tail
    caveats; Haghani 2017 empirical study).

    p_win MUST be the probability the trade is *profitable* — i.e. for a long
    call, P(S_T > K + premium), NOT P(S_T > K). Substituting P(ITM) for P(win)
    over-sizes by the ratio P(ITM)/P(profit), typically 1.2–1.5×.
    Caller should compute p_win via prob_expire_itm(S, K + premium, ...).

    Institutions never risk more than half Kelly under known fat tails.
    For retail: quarter-Kelly is the standard safety margin (Bouchaud 2002,
    "Theory of financial risks" §13; cap 25% retained here).

    win_multiple:  expected multiple of premium on winners (e.g. 2.0 = 200% gain)
    loss_multiple: multiple of premium lost on losers (1.0 = full premium loss)
    """
    if win_multiple <= 0 or p_win <= 0 or p_win >= 1:
        return 0.0
    b = win_multiple / loss_multiple
    f_full = (b * p_win - (1.0 - p_win)) / b      # full Kelly
    f_half = f_full / 2.0                         # half Kelly (institutional standard)
    return round(max(0.0, min(f_half, 0.25)), 3)  # cap at 25%


# ─── Trade Finder  (main entry point for live trading decisions) ──────────────

def find_trade(symbol=None, S=None, hv=None, q=None):
    """
    Interactive trade finder — answers the question:
    "Which call/put should I buy or sell, at what strike and expiry?"

    Workflow:
      1. Enter symbol + fetch live spot, HV, div yield (skipped if passed in)
      2. Specify: buy or sell
      3. GEX analysis auto-suggests Call or Put — press Enter to accept or override
      4. Specify: expiry range (DTE) and premium budget
      5. Optionally: target delta
      6. Outputs ranked contracts + full trade scorecard
    """
    print("\n" + "="*60)
    print("  OPTION TRADE FINDER")
    print("="*60)

    if symbol is None:
        symbol = input("Symbol (e.g. AAPL, SPY, TSLA): ").strip().upper()
    action = input("Buy or Sell? [b/s]: ").strip().lower()

    # ── Fetch market data before Q3 so GEX can run ──
    if S is None or hv is None or q is None:
        print(f"\nFetching live data for {symbol}...")
        hist = _ticker(symbol).history(period='2d')
        if hist.empty:
            print("Error: could not fetch data.")
            return
        if S  is None: S  = hist['Close'].iloc[-1]
        if q  is None: q  = get_dividend_yield(symbol)
        if hv is None: hv = historical_volatility(symbol)
    else:
        print(f"\nUsing data from Full Analysis for {symbol}...")

    global _yield_curve_fn
    if _yield_curve_fn is None:
        _yield_curve_fn = _build_yield_curve()
    r_short = get_risk_free_rate(30 / 365.0)  # short-term rate for GEX chain

    print(f"Spot: ${S:.2f}  |  HV: {hv*100:.1f}%  |  Div yield: {q*100:.2f}%")

    # ── VIX term structure regime ──────────────────────────────────────────────
    vts = vix_term_structure()
    if vts:
        print(f"VIX: {vts['VIX']}  |  VIX3M: {vts['VIX3M']}  "
              f"|  Slope: {vts['slope']:+.1%}  →  {vts['regime']}")

    # ── Q3: GEX auto-suggest ──
    print("\nRunning GEX analysis on nearest expiry chain...")
    gex = gex_signal(symbol, S, r_short, hv, q)
    if gex:
        print(f"\n  ┌─ GEX SIGNAL ({gex['expiry_used']}) {'─'*30}")
        print(f"  │  Regime:      {gex['regime']} gamma  →  {gex['regime_label']}")
        print(f"  │  Net GEX:     ${gex['net_gex']/1e6:+.1f}M")
        print(f"  │  GEX Flip:    ${gex['flip_level']:.2f}  (zero-gamma level — key pivot)")
        print(f"  │  Call Wall:   ${gex['call_wall']:.2f}  (resistance — dealers sell above)")
        print(f"  │  Put Wall:    ${gex['put_wall']:.2f}  (support — dealers buy below)")
        print(f"  │  Signal:      → {gex['suggested'].upper()}")
        print(f"  │  Reason:      {gex['reason']}")
        print(f"  └{'─'*45}")

        # 25Δ Risk Reversal — second directional signal (Bali & Murray 2013)
        # Fetch nearest expiry chain we already have from GEX, reuse it
        try:
            _near_chain = _ticker(symbol).option_chain(gex['expiry_used'])
            _near_df    = analyze_chain(symbol, gex['expiry_used'], S, r_short, hv, q)
            rr = risk_reversal_25d(_near_df)
            if rr:
                rr_agree = ((rr['RR'] < -2 and gex['suggested'] == 'put') or
                            (rr['RR'] > 2  and gex['suggested'] == 'call'))
                agree_str = ' agrees with GEX' if rr_agree else ' conflicts with GEX'
                print(f"\n  25Δ Risk Reversal: {rr['RR']:+.1f}%  "
                      f"(put IV {rr['iv_put_25']}% vs call IV {rr['iv_call_25']}%)")
                print(f"  Skew signal: {rr['signal']}  [{agree_str}]")
        except Exception:
            pass

        ans = input(f"\n  GEX suggests {gex['suggested'].upper()} — press Enter to accept or type [c/p] to override: ").strip().lower()
        option_type = ans if ans in ('c', 'p') else gex['suggested'][0]
    else:
        print("  (GEX unavailable — answer manually)")
        option_type = input("Call or Put?  [c/p]: ").strip().lower()

    dte_min = int(input("Min days to expiry (e.g. 20):  ") or 20)
    dte_max = int(input("Max days to expiry (e.g. 60):  ") or 60)

    if action == 'b':
        budget  = float(input("Max premium willing to pay per contract ($, e.g. 5.00): ") or 9999)
        min_prem = 0.01
        max_prem = budget
    else:
        min_prem = float(input("Min credit you want to receive ($, e.g. 1.00): ") or 0.01)
        max_prem = 9999

    delta_target = input("Target delta (optional, e.g. 0.30 for call, -0.25 for put — press Enter to skip): ").strip()
    delta_target = float(delta_target) if delta_target else None

    # ── Market hours check ──────────────────────────────────────────────────────
    # Options only trade 9:30 AM – 4:00 PM ET Mon–Fri.
    # Outside those hours bid/ask are $0 for every contract — prices are stale.
    import pytz
    _et   = pytz.timezone('America/New_York')
    _now  = datetime.now(_et)
    _wday = _now.weekday()                        # 0=Mon … 6=Sun
    _t    = _now.hour * 60 + _now.minute          # minutes since midnight ET
    market_open = (_wday < 5) and (570 <= _t < 960)   # 9:30=570, 16:00=960

    if not market_open:
        if _wday >= 5:
            _status = f"weekend ({['Mon','Tue','Wed','Thu','Fri','Sat','Sun'][_wday]})"
        elif _t < 570:
            _opens_in = 570 - _t
            _status = f"pre-market — opens in {_opens_in//60}h {_opens_in%60}m (9:30 AM ET)"
        else:
            _status = "after-hours (closed at 4:00 PM ET)"
        print(f"\n    Market is CLOSED ({_status})")
        print(f"     Bid/Ask quotes are $0.00 — using yesterday's lastPrice.")
        print(f"     Scores and spreads are estimates only. Re-run after 9:30 AM ET for live quotes.\n")

    # ── Institutional pre-scan signals (computed once, reused in loop) ──────────
    opt_str_pre = 'call' if option_type.startswith('c') else 'put'

    # Fetch 2y history ONCE — reused for IVR, RSI, mu, GARCH (no extra API calls)
    try:
        _px_hist  = list(_ticker(symbol).history(period='2y')['Close'].dropna())
        _lr_hist  = [np.log(_px_hist[k]/_px_hist[k-1]) for k in range(1, len(_px_hist))]
        # Bayesian shrinkage for drift. Merton (1980) showed the standard
        # error of a 252-day mean is σ/√252 ≈ 1.9%/yr for typical σ=0.30,
        # so the 95% CI on a one-year sample mean is roughly ±3.8%/yr.
        # Using the raw sample mean as "drift" in the Real-World EV
        # formula produces noise-dominated answers on individual stocks.
        # Shrink toward the long-run equity risk premium (≈8%/yr) and
        # hard-cap at ±15%/yr — anything outside that range is almost
        # certainly estimation error, not a true drift.
        if len(_lr_hist) >= 252:
            _mu_sample = float(np.mean(_lr_hist[-252:]) * 252)
            mu_annual  = float(np.clip(0.7 * 0.08 + 0.3 * _mu_sample, -0.15, 0.15))
        else:
            mu_annual = 0.08
        rsi_14 = compute_rsi(_px_hist)
    except Exception:
        _px_hist  = []
        _lr_hist  = []
        mu_annual = 0.08
        rsi_14    = 50.0

    # 1. IVR / IVP — Goldman/Citadel screen #1: is vol cheap or expensive?
    print("  Computing IV Rank / IV Percentile...")
    ivr, ivp = iv_rank_percentile(symbol, hv, prices=_px_hist)
    if not np.isnan(ivr):
        ivr_label = ('CHEAP — good time to buy vol' if ivr < 30
                     else 'FAIR' if ivr < 60
                     else 'EXPENSIVE — vol buyers get edge only if IVR < 40')
        print(f"  IVR: {ivr:.0f}  |  IVP: {ivp:.0f}  →  {ivr_label}")

    print(f"  RSI(14): {rsi_14:.1f}  |  1y mu (drift): {mu_annual*100:+.1f}%/yr")

    # 2. Earnings calendar risk
    has_earnings, earnings_date = earnings_within_dte(symbol, dte_max)
    if has_earnings:
        print(f"    Earnings within DTE window: {earnings_date}  "
              f"— IV may be inflated; expect vol crush after announcement")
    else:
        print(f"    No earnings within {dte_min}–{dte_max} DTE window")

    # Collect matching contracts across all expiries in DTE range
    today       = datetime.now().replace(tzinfo=None)
    expirations = _ticker(symbol).options
    results     = []

    for exp in expirations:
        exp_dt = datetime.strptime(exp, '%Y-%m-%d')
        dte    = (exp_dt - today).days
        if not (dte_min <= dte <= dte_max):
            continue

        T = max(dte / 365.0, 1 / 365.0)
        r = get_risk_free_rate(T)

        try:
            chain   = _ticker(symbol).option_chain(exp)
            opt_df  = chain.calls if option_type.startswith('c') else chain.puts
            opt_str = 'call'     if option_type.startswith('c') else 'put'
        except Exception:
            continue

        for _, row in opt_df.iterrows():
            K     = row['strike']
            bid   = row.get('bid', 0) or 0
            ask   = row.get('ask', 0) or 0
            oi    = row.get('openInterest', 0) or 0
            vol   = row.get('volume', 0) or 0

            unquoted = (bid == 0 and ask == 0)

            price = mid_price(row)
            if not (min_prem <= price <= max_prem):
                continue
            if price <= 0:
                continue

            iv   = implied_volatility(S, K, T, r, price, opt_str, q)
            sig  = iv if not np.isnan(iv) else hv
            g    = bs_greeks(S, K, T, r, sig, opt_str, q)
            delta = g['delta']

            if delta_target is not None:
                if abs(delta - delta_target) > 0.15:
                    continue

            intrinsic  = max(0, S - K) if opt_str == 'call' else max(0, K - S)
            time_val   = max(0, price - intrinsic)
            p_itm      = prob_expire_itm(S, K, T, r, sig, opt_str, q)

            if action == 'b':
                p_itm_val = p_itm * 100 if not np.isnan(p_itm) else 0
                if p_itm_val < 15:
                    continue
                if oi == 0 and vol == 0:
                    continue

            amer_price = american_option_price(S, K, T, r, sig, opt_str, q)
            eep        = american_early_exercise_premium(S, K, T, r, sig, opt_str, q)
            exp_mv     = expected_move(S, sig, T)

            breakeven          = (K + price) if opt_str == 'call' else (K - price)
            breakeven_move_pct = abs(breakeven - S) / S * 100

            # Real-world EV (P-measure edge)
            rw_ev = real_world_option_ev(S, K, T, r, sig, opt_str, mu_annual, price, q)

            # ── Institutional scoring ──────────────────────────────────────────
            liq_penalty  = ((50 if oi == 0 else 0)
                            + (20 if vol < 10 else 0)
                            + (25 if unquoted else 0))
            iv_hv_spread = (iv - hv) * 100 if not np.isnan(iv) else 0

            if action == 'b':
                p_itm_pct = (p_itm * 100) if not np.isnan(p_itm) else 0

                # Signal 1: IV Rank — buying cheap vol is the core institutional edge
                ivr_score = 0.0
                if not np.isnan(ivr):
                    if ivr < 20:   ivr_score = 30    # vol very cheap
                    elif ivr < 35: ivr_score = 20
                    elif ivr < 50: ivr_score = 5
                    elif ivr < 65: ivr_score = -15
                    else:          ivr_score = -35   # vol very expensive

                # Signal 2: Real-world EV (P-measure drift vs market price)
                ev_score = 0.0
                if not np.isnan(rw_ev):
                    ev_score = max(-25.0, min(25.0, (rw_ev / price) * 30))

                # Signal 3: Delta/prob — reward sweet spot, not lottery tickets
                # ATM to slight OTM (delta 0.25–0.50) is the institutional sweet spot
                delta_abs = abs(delta)
                if 0.30 <= delta_abs <= 0.55:
                    delta_score = 25
                elif 0.20 <= delta_abs < 0.30:
                    delta_score = 12
                elif 0.15 <= delta_abs < 0.20:
                    delta_score = 0
                else:
                    delta_score = -20   # deep OTM lottery ticket

                # Signal 4: Momentum position via RSI (mean-reversion aware)
                # Institutions use RSI to AVOID exhausted moves, not chase them.
                # Best call entry: RSI 35–55 (stock rising from rest, not overbought).
                # Best put entry: RSI 45–65 (stock falling from elevated levels).
                rsi_score = 0.0
                if opt_str == 'call':
                    if 35 <= rsi_14 <= 55:     rsi_score = 20   # ideal: fresh / not extended
                    elif 55 < rsi_14 <= 65:    rsi_score = 8    # slightly extended but OK
                    elif 65 < rsi_14 <= 72:    rsi_score = -8   # getting overbought
                    elif rsi_14 > 72:          rsi_score = -25  # overbought — momentum exhausted
                    elif 25 <= rsi_14 < 35:    rsi_score = -8   # oversold bounce? risky call
                    elif rsi_14 < 25:          rsi_score = -20  # falling knife — don't buy calls
                else:  # put
                    if 45 <= rsi_14 <= 65:     rsi_score = 20   # ideal: turning down from elevated
                    elif 35 <= rsi_14 < 45:    rsi_score = 8    # slightly extended downside
                    elif 28 <= rsi_14 < 35:    rsi_score = -8   # oversold — puts now expensive
                    elif rsi_14 < 28:          rsi_score = -25  # panic oversold — puts at peak price
                    elif 65 < rsi_14 <= 75:    rsi_score = -5   # put on overbought? possible
                    elif rsi_14 > 75:          rsi_score = -15  # put on very strong uptrend, bad entry

                # Signal 5: Move needed — penalize far-OTM breakevens non-linearly
                # Under 5%: free money territory; 10–15%: fair; 20%+: lottery ticket
                move_score = max(-30.0, -(breakeven_move_pct ** 1.3) * 0.8)

                # Signal 6: IV vs HV — symmetric: cheap IV rewarded equally to expensive IV penalized
                # Fix: Previous code rewarded cheap vol only 0.3x but penalized expensive 0.5x.
                # Now both use 0.5x — symmetrical edge scoring.
                iv_score = -abs(max(iv_hv_spread, 0)) * 0.5 + abs(min(0, iv_hv_spread)) * 0.5

                # Signal 7: Earnings penalty — magnitude-scaled by IV inflation
                # If IV is very high (IVR > 70), earnings crush will be severe → bigger penalty
                earn_base    = -25 if has_earnings else 0
                earn_penalty = earn_base * (1.5 if not np.isnan(ivr) and ivr > 70 else 1.0)

                # Signal 8: Extra reward for buying genuinely cheap vol (IV-HV < -8%)
                # Only applies when IVR also confirms cheapness
                cheap_bonus = (15 if iv_hv_spread < -8 and not np.isnan(ivr) and ivr < 30
                               else 8 if iv_hv_spread < -5
                               else 0)

                score = (
                    ivr_score
                    + ev_score
                    + delta_score
                    + rsi_score
                    + move_score
                    + iv_score
                    + earn_penalty
                    + cheap_bonus
                    - liq_penalty
                )

                # Kelly fraction for display in scorecard.
                # Use TRUE P(profit) = P(S_T crosses breakeven), not P(ITM).
                # For a long call, profit threshold is K + premium; for a put,
                # K − premium. Otherwise Kelly systematically over-sizes
                # by the ratio P(ITM)/P(profit), typically 1.2–1.5×.
                K_be       = (K + price) if opt_str == 'call' else (K - price)
                p_profit   = prob_expire_itm(S, K_be, T, r, sig, opt_str, q)
                p_win_kelly = (p_profit if (not np.isnan(p_profit) and 0 < p_profit < 1)
                               else p_itm_pct / 100)
                avg_win_mult = 1.5   # assume 1.5x avg winner on profitable exits
                kelly = kelly_fraction_options(p_win_kelly, avg_win_mult)

            else:  # sellers
                # Signal 1: IV premium (core seller edge)
                iv_prem_score = max(iv_hv_spread, 0) * 0.5

                # Signal 2: IVR — sell high vol
                sell_ivr = 0.0
                if not np.isnan(ivr):
                    if ivr > 70:   sell_ivr = 30
                    elif ivr > 50: sell_ivr = 15
                    elif ivr > 35: sell_ivr = 5
                    else:          sell_ivr = -20  # vol too cheap to sell

                score = (
                    g['theta'] * -100     # theta income
                    + iv_prem_score
                    + sell_ivr
                    - liq_penalty
                )
                kelly = 0.0
                rw_ev = rw_ev if not np.isnan(rw_ev) else 0.0

            results.append({
                'Expiry':                 exp,
                'DTE':                    dte,
                'Strike':                 K,
                'Moneyness':              classify_moneyness(S, K, opt_str),
                'Mid Price ($)':          round(price, 2),
                'Bid':                    round(row.get('bid', 0) or 0, 2),
                'Ask':                    round(row.get('ask', 0) or 0, 2),
                'Bid-Ask Spread ($)':     round((row.get('ask', 0) or 0) - (row.get('bid', 0) or 0), 2),
                'IV (%)':                 round(iv * 100, 1) if not np.isnan(iv) else np.nan,
                'HV (%)':                 round(hv * 100, 1),
                'IV-HV Spread (%)':       round(iv_hv_spread, 1),
                'IV Rank':                ivr if not np.isnan(ivr) else np.nan,
                'IV Percentile':          ivp if not np.isnan(ivp) else np.nan,
                'RSI(14)':                rsi_14,
                'Real-World EV ($)':      round(rw_ev, 3) if not np.isnan(rw_ev) else np.nan,
                'Kelly Fraction':         round(kelly, 3) if action == 'b' else np.nan,
                'Earnings Risk':          earnings_date if has_earnings else 'None',
                'Delta':                  round(delta, 3),
                'Theta ($/day)':          round(g['theta'], 3),
                'Gamma':                  round(g['gamma'], 5),
                'Vega ($/1%vol)':         round(g['vega'], 3),
                'American Price ($)':     round(amer_price, 2) if not np.isnan(amer_price) else np.nan,
                'Early Exercise Prem ($)':round(eep, 3) if not np.isnan(eep) else np.nan,
                'Intrinsic ($)':          round(intrinsic, 2),
                'Time Value ($)':         round(time_val, 2),
                'Breakeven @ Expiry':     round(breakeven, 2),
                'Move Needed (%)':        round(breakeven_move_pct, 1),
                'Prob ITM (%)':           round(p_itm * 100, 1) if not np.isnan(p_itm) else np.nan,
                'Expected Move ($)':      round(exp_mv, 2),
                'Volume':                 int(pd.to_numeric(row.get('volume', 0), errors='coerce') or 0),
                'Open Interest':          int(pd.to_numeric(row.get('openInterest', 0), errors='coerce') or 0),
                'Score':                  round(score, 3),
            })

    if not results:
        # Help the user understand what budget they'd need
        print("\n    No tradeable contracts found within your criteria.")
        print(f"     Filters: {dte_min}–{dte_max} DTE, max ${max_prem:.2f} premium, "
              f"Prob ITM ≥ 15%, OI or volume > 0")
        # Find cheapest qualifying contract (ignoring budget cap) to guide user
        _suggestions = []
        for exp in expirations:
            try:
                _chain = _ticker(symbol).option_chain(exp)
                _df = _chain.puts if option_type.startswith('p') else _chain.calls
                _trade_dt = datetime.now().replace(tzinfo=None)
                _exp_dt   = datetime.strptime(exp, '%Y-%m-%d')
                _dte      = (_exp_dt - _trade_dt).days
                if not (dte_min <= _dte <= dte_max):
                    continue
                for _, _r in _df.iterrows():
                    _p = mid_price(_r)
                    if _p <= 0:
                        continue
                    _T  = max(_dte / 365.0, 1/365.0)
                    _K  = _r['strike']
                    _iv = implied_volatility(S, _K, _T, r_short, _p, opt_str, q)
                    _sg = _iv if not np.isnan(_iv) else hv
                    _pi = prob_expire_itm(S, _K, _T, r_short, _sg, opt_str, q)
                    if not np.isnan(_pi) and _pi * 100 >= 15:
                        _suggestions.append((_p, _K, exp, _dte, round(_pi*100,1)))
            except Exception:
                continue
        if _suggestions:
            _suggestions.sort()
            _cheapest = _suggestions[0]
            print(f"\n     Cheapest qualifying contract in your DTE window:")
            print(f"     {symbol} {_cheapest[2]} {opt_str.upper()} K={_cheapest[1]}  "
                  f"→  ${_cheapest[0]:.2f}/share = ${_cheapest[0]*100:.0f}/contract  "
                  f"(Prob ITM {_cheapest[4]}%)")
            if _cheapest[0] > max_prem:
                print(f"\n     → Increase your max premium to at least ${_cheapest[0]:.2f} and re-run.")
            else:
                print(f"\n     → Premium fits your budget. The chain may have no live bid/ask right now.")
                print(f"       Re-run after market open (9:30 AM ET) for live quotes.")
        else:
            print("     → No tradeable puts/calls found. Try wider DTE (e.g. 20–60 days).")
        return

    df = pd.DataFrame(results).sort_values('Score', ascending=False).reset_index(drop=True)

    # ── Print top 10 to console ──
    action_str = 'BUY' if action == 'b' else 'SELL'
    opt_str_display = 'CALL' if option_type.startswith('c') else 'PUT'
    print(f"\n{'─'*60}")
    print(f"  TOP CONTRACTS TO {action_str} — {symbol} {opt_str_display}")
    print(f"  Spot: ${S:.2f}  |  DTE: {dte_min}–{dte_max} days")
    print(f"{'─'*60}")

    display_cols = ['Expiry', 'DTE', 'Strike', 'Moneyness', 'Mid Price ($)',
                    'Bid-Ask Spread ($)', 'IV (%)', 'IV Rank', 'IV-HV Spread (%)',
                    'Delta', 'Theta ($/day)', 'Prob ITM (%)',
                    'Real-World EV ($)', 'Move Needed (%)', 'Volume']
    top10 = df[display_cols].head(10)

    print(top10.to_string(index=True))
    print(f"\n{'─'*60}")

    # ── Trade scorecard for #1 ranked contract ──
    best = df.iloc[0]
    print(f"\n  TRADE SCORECARD — #{1} Ranked Contract")
    print(f"  {'─'*40}")
    print(f"  {action_str} {symbol} {best['Strike']} {opt_str_display} exp {best['Expiry']}")
    print(f"  ─── Pricing ───────────────────────────")
    print(f"  Mid Price:           ${best['Mid Price ($)']:.2f}  (Bid ${best['Bid']:.2f} / Ask ${best['Ask']:.2f})")
    print(f"  Bid-Ask Spread:      ${best['Bid-Ask Spread ($)']:.2f}  {' wide' if best['Bid-Ask Spread ($)'] > best['Mid Price ($)']*0.10 else ' tight'}")
    print(f"  American Price:      ${best['American Price ($)']:.2f}  (European: ${best['Mid Price ($)'] - best['Early Exercise Prem ($)']:.2f})")
    print(f"  Early Exercise Prem: ${best['Early Exercise Prem ($)']:.3f}")

    print(f"  ─── Institutional Vol Signals ──────────")
    print(f"  Implied Vol:         {best['IV (%)']:.1f}%")
    print(f"  Historical Vol:      {best['HV (%)']:.1f}%")
    iv_signal = ('EXPENSIVE — vol sellers edge' if best['IV-HV Spread (%)'] > 3
                 else 'CHEAP — vol buyers edge' if best['IV-HV Spread (%)'] < -3
                 else 'FAIR')
    print(f"  IV vs HV:            {best['IV-HV Spread (%)']:+.1f}%  → {iv_signal}")
    # IV Rank / Percentile — primary institutional screen
    if not np.isnan(best.get('IV Rank', np.nan)):
        ivr_v = best['IV Rank']
        ivp_v = best['IV Percentile']
        ivr_color = (' CHEAP — ideal for buyers' if ivr_v < 30
                     else ' FAIR' if ivr_v < 60
                     else ' EXPENSIVE — risk of IV crush')
        print(f"  IV Rank (IVR):       {ivr_v:.0f}/100  →  {ivr_color}")
        print(f"  IV Percentile (IVP): {ivp_v:.0f}%  (vol cheaper than {ivp_v:.0f}% of past year)")

    print(f"  ─── Greeks ─────────────────────────────")
    print(f"  Delta:               {best['Delta']:+.3f}  (${abs(best['Delta'])*100:.0f} P&L per $1 move, 100 shares)")
    print(f"  Theta:               {best['Theta ($/day)']:.3f}  (${best['Theta ($/day)']*100:.2f}/day per contract)")
    print(f"  Gamma:               {best['Gamma']:.5f}")
    print(f"  Vega:                {best['Vega ($/1%vol)']:.3f}  (${best['Vega ($/1%vol)']*100:.2f} per 1% IV change)")

    print(f"  ─── Momentum (RSI) ─────────────────────")
    rsi_v = best.get('RSI(14)', rsi_14)
    if opt_str_pre == 'call':
        rsi_label = (' Ideal zone — fresh move, not extended'   if 35 <= rsi_v <= 55
                     else ' Slightly extended — OK to enter'    if 55 < rsi_v <= 65
                     else ' Getting overbought — momentum tiring' if 65 < rsi_v <= 72
                     else ' Overbought — do NOT chase calls'    if rsi_v > 72
                     else ' Oversold bounce — risky call entry' if 25 <= rsi_v < 35
                     else ' Falling knife — calls very risky')
    else:
        rsi_label = (' Ideal zone — turning down from elevated'  if 45 <= rsi_v <= 65
                     else ' Extended downside — still tradeable' if 35 <= rsi_v < 45
                     else ' Oversold — puts getting expensive'  if 28 <= rsi_v < 35
                     else ' Panic oversold — puts at peak price' if rsi_v < 28
                     else ' Put on strong uptrend — careful'    if rsi_v > 75
                     else ' Pullback potential — OK for puts')
    print(f"  RSI(14):             {rsi_v:.1f}  →  {rsi_label}")

    print(f"  ─── Risk / Reward ──────────────────────")
    print(f"  Intrinsic Value:     ${best['Intrinsic ($)']:.2f}")
    print(f"  Time Value:          ${best['Time Value ($)']:.2f}")
    print(f"  Breakeven @ expiry:  ${best['Breakeven @ Expiry']:.2f}  (need {best['Move Needed (%)']:.1f}% move)")
    print(f"  Prob of expiring ITM:{best['Prob ITM (%)']:.1f}%")
    print(f"  Expected 1σ move:    ${best['Expected Move ($)']:.2f}")

    # Real-world EV and Kelly — institutional sizing and edge check
    if action == 'b':
        ev_v = best.get('Real-World EV ($)', np.nan)
        kf_v = best.get('Kelly Fraction', 0.0)
        print(f"  ─── Edge & Sizing (Institutional) ─────")
        if not np.isnan(ev_v if ev_v is not None else np.nan):
            ev_prem_pct = ev_v / best['Mid Price ($)'] * 100
            ev_label = ('POSITIVE EDGE — stock drift favors buyer' if ev_v > 0
                        else 'NEGATIVE EDGE — market pricing is fair/expensive')
            print(f"  Real-World EV:       ${ev_v:+.3f}/share  ({ev_prem_pct:+.1f}% of premium)  →  {ev_label}")
            print(f"  [Hakansson-Rubinstein P-measure: uses {mu_annual*100:+.1f}%/yr historical drift]")
        if not np.isnan(kf_v if kf_v is not None else np.nan):
            kf_pct = kf_v * 100
            kf_contracts = max(1, int(kf_v * 10000 / (best['Mid Price ($)'] * 100)))
            print(f"  Kelly Fraction:      {kf_pct:.1f}% of account  →  ~{kf_contracts} contract(s) per $10k")
            print(f"  [Half-Kelly: Thorp 1969. Quarter-Kelly recommended for retail traders]")
        if has_earnings:
            print(f"    Earnings on {earnings_date} — IV will collapse after announcement.")
            print(f"     If trading through earnings: buy a straddle, not a directional contract.")
    print(f"  ─── Liquidity ──────────────────────────")
    _vol = int(best['Volume'])       if (pd.notna(best['Volume'])       and best['Volume'] == best['Volume']) else 0
    _oi  = int(best['Open Interest']) if (pd.notna(best['Open Interest']) and best['Open Interest'] == best['Open Interest']) else 0
    print(f"  Volume:              {_vol:,}")
    print(f"  Open Interest:       {_oi:,}")
    bid_ok  = best['Bid'] > 0 and best['Ask'] > 0
    oi_ok   = _oi > 50
    vol_ok  = _vol > 20
    spread_ok = best['Bid-Ask Spread ($)'] < best['Mid Price ($)'] * 0.10
    if not bid_ok and not market_open:
        liq = 'UNQUOTED (market closed — check bid/ask after 9:30 AM ET)'
    elif not bid_ok:
        liq = 'DEAD — no bid/ask quote, DO NOT TRADE'
    elif not oi_ok and not vol_ok:
        liq = 'POOR — OI and volume both very low, risk of bad fill'
    elif not spread_ok:
        liq = 'POOR — wide spread, pay limit price between bid and ask'
    elif vol_ok and oi_ok and spread_ok:
        liq = 'GOOD'
    else:
        liq = 'OK — check bid/ask before submitting'
    print(f"  Liquidity:           {liq}")
    print(f"  ─── Cost (1 contract = 100 shares) ─────")
    cost = best['Mid Price ($)'] * 100
    print(f"  Premium cost:        ${cost:,.2f}")
    if action == 'b':
        print(f"  Max loss:            ${cost:,.2f}  (premium paid)")
        print(f"  Max profit:          Unlimited (call) / Strike limited (put)")
    else:
        print(f"  Max profit:          ${cost:,.2f}  (premium received)")
        print(f"  Max loss:            Unlimited / large — use spreads to cap risk")
    # ── Quant Risk Engine (buyers only) ──────────────────────────────────────
    if action == 'b':
        opt_type_str = 'put' if option_type.startswith('p') else 'call'
        sig_b  = best['IV (%)'] / 100 if not np.isnan(best.get('IV (%)', np.nan)) else hv
        T_b    = best['DTE'] / 365.0
        r_b    = get_risk_free_rate(T_b)
        K_b    = float(best['Strike'])
        prem_b = float(best['Mid Price ($)'])

        # Second-order Greeks
        g2 = bs_second_order(S, K_b, T_b, r_b, sig_b, opt_type_str, q)
        # Gamma breakeven
        g1 = bs_greeks(S, K_b, T_b, r_b, sig_b, opt_type_str, q)
        gbe = gamma_breakeven_move(g1['gamma'], g1['theta'], S)
        # Delta-Gamma VaR (99%, 1-day)
        var1d = delta_gamma_var(g1['delta'], g1['gamma'], S, sig_b)
        # GARCH vol forecast — reuse _px_hist / _lr_hist already fetched above
        log_rets = _lr_hist if _lr_hist else []
        garch_fc = garch_vol_forecast(log_rets, horizon=int(T_b * 252) or 21) if log_rets else None
        # D-P exit threshold
        mu_est = mu_annual  # already computed from _lr_hist[-252:]
        dp_exit = dp_exit_threshold(K_b, mu_est, sig_b, r_b) if opt_type_str == 'call' else None
        # Pin risk
        nearest_chain_df = None
        for _df in (all_dfs if 'all_dfs' in dir() else []):
            if not _df.empty:
                nearest_chain_df = _df
                break
        try:
            _exp_chain = _ticker(symbol).option_chain(best['Expiry'])
            _combined  = pd.concat([_exp_chain.calls.assign(type='call'),
                                    _exp_chain.puts.assign(type='put')])
            _combined.rename(columns={'openInterest': 'Open Interest', 'strike': 'Strike'}, inplace=True)
            # Pass T/r/σ/q so pin kernel uses true Avellaneda-Lipkin
            # Γ × OI weighting instead of the Gaussian-proxy fallback.
            pin = pin_risk_score(S, _combined[['Strike','Open Interest']],
                                 T=T_b, r=r_b, sigma=sig_b, q=q)
        except Exception:
            pin = None

        print(f"\n  ─── Quant Risk Engine ───────────────────────────")
        print(f"  [Britten-Jones 1999 · Bollerslev 1986 · Haug 2007 · Dixit-Pindyck 1994]")

        # Gamma breakeven
        if not np.isnan(gbe.get('delta_s_be', np.nan)):
            be_cmp = ' PROFITABLE' if sig_b * 100 > gbe['sigma_be_pct'] else ' NOT COVERED'
            print(f"\n  Gamma Breakeven (Taleb 1997):")
            print(f"    Min daily move needed: ${gbe['delta_s_be']:.2f}  ({gbe['sigma_be_pct']:.1f}% ann.)")
            print(f"    Realised IV:           {sig_b*100:.1f}%  → long gamma {be_cmp}")

        # 1-day VaR
        if var1d is not None:
            print(f"\n  Delta-Gamma 1-day VaR (99%):  -${abs(var1d):.2f}/contract")
            print(f"    (expected max loss 99% of days; vs premium ${prem_b*100:.0f})")

        # Second-order Greeks
        if not np.isnan(g2.get('vanna', np.nan)):
            print(f"\n  Second-Order Greeks (Haug 2007 / Castagna & Mercurio 2007):")
            print(f"    Vanna:  {g2['vanna']:+.6f}  (delta sensitivity to 1% vol move)")
            print(f"    Volga:  {g2['volga']:+.6f}  (vega sensitivity to 1% vol move)")
            print(f"    Charm:  {g2['charm']:+.6f}  (delta decay per day — pin risk signal)")
            print(f"    Speed:  {g2['speed']:+.8f}  (gamma sensitivity to $1 move)")

        # GARCH forecast
        if garch_fc:
            garch_vs_iv = garch_fc['vol'] * 100 - sig_b * 100
            edge_str = ('vol likely rising → buying edge' if garch_vs_iv > 2
                        else 'vol likely falling → sellers edge' if garch_vs_iv < -2
                        else 'vol roughly fair')
            print(f"\n  GARCH(1,1) Vol Forecast (Bollerslev 1986):")
            print(f"    {int(T_b*252)}d horizon forecast: {garch_fc['vol']*100:.1f}%  "
                  f"(persistence α+β = {garch_fc['persistence']:.3f})")
            print(f"    vs Current IV {sig_b*100:.1f}%  → {edge_str}")

        # D-P exit threshold (calls only)
        if dp_exit and not np.isnan(dp_exit):
            print(f"\n  Dixit-Pindyck Optimal Exit Threshold (1994):")
            print(f"    Continuation → exercise boundary: S* = ${dp_exit:.2f}")
            print(f"    Holding call above ${dp_exit:.2f} has negative expected utility")
            print(f"    → SELL CALL if {symbol} approaches ${dp_exit:.2f}")

        # Pin risk
        if pin:
            dist_str = f"${pin['pin_strike']:.2f} ({pin['distance_pct']:.1f}% away)"
            strength_str = 'STRONG' if pin['pin_strength'] > 0.4 else 'MODERATE' if pin['pin_strength'] > 0.2 else 'WEAK'
            print(f"\n  Pin Risk at Expiry (Avellaneda & Lipkin 2003):")
            print(f"    Highest-gravity strike: {dist_str}  [{strength_str} pin, score {pin['pin_strength']:.2f}]")
            if pin['distance_pct'] < 3.0:
                print(f"     Price near pin strike — expect slow drift toward ${pin['pin_strike']:.2f} into expiry")

    # ── Smart Exit Levels (buyers only) ──────────────────────────────────────
    if action == 'b' and gex:
        opt_type_str = 'put' if option_type.startswith('p') else 'call'
        sig_best  = best['IV (%)'] / 100 if not np.isnan(best.get('IV (%)', np.nan)) else hv
        T_best    = best['DTE'] / 365.0
        r_best    = get_risk_free_rate(T_best)

        # ── Fit per-symbol skew slope from the live chain ───────────────────
        # log(IV) ≈ const − slope × log(K/S) near ATM.
        # Single-stock skew varies widely (0.05 for low-vol names to 0.50 for
        # TSLA/NVDA-class); a hard-coded 0.30 mis-prices the GEX exit ladder
        # by 3-5 vol points on high-skew names. Use the already-scored
        # contracts (same Expiry as `best`) to extract the local slope.
        try:
            _skew_df  = df[(df['Expiry'] == best['Expiry']) &
                           df['IV (%)'].notna() &
                           df['Strike'].between(S * 0.85, S * 1.15)].copy()
            if len(_skew_df) >= 4:
                _logm = np.log(_skew_df['Strike'].values / S)
                _logv = np.log(_skew_df['IV (%)'].values / 100.0)
                # least-squares slope of log(IV) on −log(K/S)
                _slope_fit = float(np.polyfit(-_logm, _logv, 1)[0])
                skew_slope = float(np.clip(_slope_fit, 0.05, 0.50))
            else:
                skew_slope = 0.30
        except Exception:
            skew_slope = 0.30

        exits = smart_exit_levels(
            S=S, K=float(best['Strike']), T=T_best, r=r_best, sigma=sig_best,
            option_type=opt_type_str, premium_paid=float(best['Mid Price ($)']),
            gex=gex, skew_slope=skew_slope,
        )
        if exits and exits.get('levels'):
            print(f"\n  ─── Smart Exit Plan (GEX + Mathematics) ────────")
            print(f"  Research: Bollen & Whaley 2004 · Black-Scholes 1973 · Kelly 1956")
            print(f"  Daily 1σ move: ${exits['daily_1sigma']:.2f}  |  GEX Regime: {exits['regime']}")
            print(f"  ─────────────────────────────────────────────────")
            for lv in exits['levels']:
                pnl_v   = lv.get('pnl')
                opt_v   = lv.get('opt_val')
                prob_v  = lv.get('prob')
                days_v  = lv.get('days')
                kf_v    = lv.get('kelly_exit', 1.0)
                pnl_s   = f"${pnl_v:+.0f}/contract" if pnl_v is not None and not np.isnan(pnl_v) else "n/a"
                optv_s  = f"${opt_v:.2f}" if opt_v is not None and not np.isnan(opt_v) else "n/a"
                prob_s  = f"{prob_v*100:.0f}%" if prob_v is not None and not np.isnan(prob_v) else "?"
                days_s  = f"~{days_v:.0f}d" if days_v is not None and not np.isnan(days_v) else "?"
                print(f"\n  {lv['name']}")
                print(f"    Target spot:  ${lv['spot']:.2f}  |  Est. days: {days_s}  |  P(reach): {prob_s}")
                print(f"    Option value: {optv_s}  |  P&L: {pnl_s}")
                print(f"    → {lv['action']}  (Kelly: sell {kf_v*100:.0f}% of position)")
                print(f"    [{lv['note']}]")
            print(f"\n  ─── Theta Clock ─────────────────────────────────")
            td  = exits['theta_pct_day']
            tvd = exits.get('tv_death_days')
            print(f"  Decay rate:  {td:.2f}%/day  {' HIGH — decay dominates' if exits['theta_danger'] else ' manageable'}")
            if tvd:
                print(f"  Time value fully consumed in: {tvd:.0f} days")
                print(f"  Hard exit trigger: ≤ {exits['theta_exit_dte']} DTE regardless of price")
    print(f"\n{'='*60}\n")

    # ── Save to Excel ──
    fname = f"{symbol}_{opt_str_display}_{action_str}_trade_finder.xlsx"
    df.to_excel(fname, index=False)
    print(f"Full results saved: {fname}")

    try:
        from google.colab import files
        files.download(fname)
    except ImportError:
        pass


# ─── Backtest ─────────────────────────────────────────────────────────────────

def backtest_model():
    """
    Replay the trade-finder scoring model against historical prices.

    Since free data sources don't provide historical option chains, we build
    a synthetic chain at each past entry date using Black-Scholes with 30-day
    realised volatility computed from the stock's own price history.

    Limitations (be honest with yourself):
      • No real historical bid/ask — prices are theoretical BS values.
      • GEX cannot be replayed (no historical open-interest data for free).
        Direction is either user-supplied or set by a 20-day momentum proxy.
      • Slippage / commissions are not modelled.
      • Results will be optimistic vs. live trading.
    """
    print("\n" + "="*60)
    print("  STRATEGY BACKTEST  (synthetic Black-Scholes options)")
    print("="*60)
    print("  Note: uses theoretical BS prices — no real bid/ask history.")
    print("  GEX cannot be replayed; use 'auto' for momentum proxy.\n")

    symbol     = input("Symbol (e.g. AAPL, SPY, ONDS): ").strip().upper()
    action     = input("Buy or Sell? [b/s]: ").strip().lower()
    dir_input  = input("Direction — Call / Put / Auto-momentum [c/p/a]: ").strip().lower()
    dte_target = int(input("Target DTE at entry (e.g. 30): ") or 30)
    lookback   = int(input("Lookback in trading days (e.g. 252 = 1 year): ") or 252)
    tp_input    = input("Take profit % of premium (e.g. 100 = exit at 100% gain, 0 = none): ").strip()
    take_profit = float(tp_input) / 100 if tp_input and tp_input != '0' else None
    be_input    = input("Break-even trigger % (e.g. 25 = once up 25% move stop to entry, 0 = none): ").strip()
    be_trigger  = float(be_input) / 100 if be_input and be_input != '0' else None
    stop_loss   = input("Stop loss % of premium (e.g. 50 = exit when option loses 50%, 0 = none): ").strip()
    stop_loss   = float(stop_loss) / 100 if stop_loss and stop_loss != '0' else None
    exit_dte    = int(input("Exit at X DTE remaining (0 = hold to expiry): ") or 0)

    print(f"\nFetching {lookback + dte_target + 60} days of price history for {symbol}...")
    raw = _ticker(symbol).history(period=f"{lookback + dte_target + 90}d")
    if raw.empty or len(raw) < 60:
        print("Not enough history to backtest.")
        return

    closes = raw['Close'].dropna()
    dates  = list(closes.index)
    # Normalise timezone so date comparisons work
    dates  = [d.replace(tzinfo=None) if hasattr(d, 'tzinfo') else d for d in dates]
    prices = list(closes.values)

    global _yield_curve_fn
    if _yield_curve_fn is None:
        _yield_curve_fn = _build_yield_curve()
    r = get_risk_free_rate(dte_target / 365.0)

    # ── Build entry index list (every ~14 trading days, after HAR-RV warmup) ──
    hv_window     = 50    # need ≥50 days for HAR-RV to fit meaningfully
    entry_indices = list(range(hv_window, len(dates) - dte_target - 5, 14))
    cutoff        = len(dates) - dte_target - 5
    entry_indices = [i for i in entry_indices if i <= cutoff]
    entry_indices = entry_indices[-(lookback // 14):]

    print(f"Simulating {len(entry_indices)} trades over {lookback} trading days...\n")
    print(f"  Vol model: HAR-RV (Corsi 2009) — daily/weekly/monthly RV components")
    print(f"  Scoring:   VRP signal boosts/penalises entry (Bakshi & Kapadia 2003)\n")

    STRIKE_PCTS = [0.80, 0.85, 0.90, 0.95, 1.00, 1.05, 1.10, 1.15, 1.20]

    trades = []
    equity = 0.0

    for ei in entry_indices:
        S = prices[ei]

        # HAR-RV forecast at entry  (Corsi 2009)
        hv = har_rv_forecast(prices[:ei])
        if np.isnan(hv) or hv <= 0.01:
            continue

        # ATM IV proxy: Use the HAR-RV *forward* forecast as our IV stand-in.
        # This eliminates look-ahead bias: the trader on entry date would form
        # a vol forecast using data BEFORE the entry date, not realised vol AFTER.
        # rv_30 (the previous approach) used realized vol from the same 30 days
        # we're about to trade — that's looking into the future.
        # Correct approach: GARCH(1,1) conditioned on pre-entry returns.
        window_prices  = prices[max(0, ei-90): ei]
        log_rets_pre   = [np.log(window_prices[k] / window_prices[k-1])
                          for k in range(1, len(window_prices))]
        garch_pre = garch_vol_forecast(log_rets_pre, horizon=21)
        iv_proxy  = garch_pre['vol'] if garch_pre else hv   # forward-looking GARCH as IV
        # VRP = forecasted vol (IV proxy) − HAR-RV baseline forecast
        vrp = iv_proxy - hv   # positive → IV rich → sellers edge
                               # negative → IV cheap → buyers edge

        # Direction
        if dir_input == 'a':
            ma20     = np.mean(prices[ei - 20: ei])
            opt_type = 'call' if S > ma20 else 'put'
        elif dir_input == 'p':
            opt_type = 'put'
        else:
            opt_type = 'call'

        T_entry = dte_target / 365.0

        # Score synthetic chain and pick best strike
        best_score         = -1e9
        best_K             = None
        best_entry_px      = None
        best_delta         = None
        best_p_itm         = None
        best_be_move       = None

        # B7 fix: price at the *IV proxy* we score against, not at HV.
        # Otherwise a "cheap vol" boost in scoring isn't realised in P&L —
        # the entry premium would be the same whether VRP is +5 or -5.
        sig_entry = iv_proxy if (iv_proxy and not np.isnan(iv_proxy) and iv_proxy > 0) else hv

        for pct in STRIKE_PCTS:
            # Snap to OCC listed-strike grid so we don't backtest fills
            # at strikes that don't exist (e.g. $180.50 on LLY).
            K     = nearest_listed_strike(S, S * pct)
            price = bs_price(S, K, T_entry, r, sig_entry, opt_type)
            if np.isnan(price) or price <= 0.01:
                continue
            g     = bs_greeks(S, K, T_entry, r, sig_entry, opt_type)
            delta = g['delta']
            p_itm = prob_expire_itm(S, K, T_entry, r, sig_entry, opt_type)
            if np.isnan(p_itm):
                continue
            breakeven  = (K + price) if opt_type == 'call' else (K - price)
            be_pct     = abs(breakeven - S) / S * 100

            if action == 'b':
                # VRP edge (Bakshi & Kapadia 2003):
                #   negative VRP (IV < HAR-RV) → options cheap → boost score
                #   positive VRP (IV > HAR-RV) → options expensive → penalise
                vrp_adj = -vrp * 30   # −30 per 1.0 unit of VRP
                score = (abs(delta) * 50
                         + p_itm * 100 * 0.5
                         - be_pct * 1.5
                         + vrp_adj)
            else:
                vrp_adj = vrp * 30    # sellers rewarded when VRP positive
                score = g['theta'] * -100 + vrp_adj

            if score > best_score:
                best_score    = score
                best_K        = K
                best_entry_px = price
                best_delta    = delta
                best_p_itm    = p_itm
                best_be_move  = be_pct

        if best_K is None:
            continue

        # ── Simulate day-by-day until exit ──────────────────────────────────
        expiry_i    = min(ei + dte_target, len(dates) - 1)
        exit_px     = None
        exit_date   = None
        exit_reason = 'expiry'
        be_active   = False          # True once BE trigger is hit

        # B2 partial fix: simulate IV with a mean-reverting log-OU process
        # so vega P&L isn't zero. Half-life ≈ 21d, vol-of-vol ≈ 30%/yr (matches
        # SPX VIX-of-VIX empirical regime, see Carr-Wu 2016 §4).
        rng_iv = np.random.default_rng(seed=ei)   # reproducible per trade
        kappa, theta_iv, eta = 1.0/21.0, np.log(sig_entry), 0.30
        log_iv_path = [np.log(sig_entry)]
        for _ in range(ei + 1, expiry_i + 1):
            prev = log_iv_path[-1]
            log_iv_path.append(
                prev + kappa * (theta_iv - prev) + eta * np.sqrt(1/252) * rng_iv.standard_normal()
            )

        for j in range(ei + 1, expiry_i + 1):
            S_j      = prices[j]
            dte_left = expiry_i - j
            T_j      = max(dte_left / 365.0, 0.0)
            sig_j    = float(np.clip(np.exp(log_iv_path[j - ei]), 0.05, 3.0))

            if T_j <= 0:
                intrinsic   = max(0, S_j - best_K) if opt_type == 'call' else max(0, best_K - S_j)
                exit_px     = intrinsic
                exit_date   = dates[j]
                exit_reason = 'expiry'
                break

            curr_px = bs_price(S_j, best_K, T_j, r, sig_j, opt_type)
            if np.isnan(curr_px):
                continue

            # 1. Take profit
            if action == 'b' and take_profit and curr_px >= best_entry_px * (1 + take_profit):
                exit_px     = curr_px
                exit_date   = dates[j]
                exit_reason = f'{int(take_profit*100)}% profit'
                break

            # 2. Break-even: once up X%, lock stop at entry price
            if action == 'b' and be_trigger and not be_active:
                if curr_px >= best_entry_px * (1 + be_trigger):
                    be_active = True   # stop now trails up to entry price

            # 3. BE stop — exit at or near entry (no loss)
            if action == 'b' and be_active and curr_px <= best_entry_px:
                exit_px     = curr_px
                exit_date   = dates[j]
                exit_reason = 'BE stop'
                break

            # 4. Regular stop loss (only before BE is triggered)
            if action == 'b' and stop_loss and not be_active:
                if curr_px <= best_entry_px * (1 - stop_loss):
                    exit_px     = curr_px
                    exit_date   = dates[j]
                    exit_reason = f'stop {int(stop_loss*100)}%'
                    break

            # 5. DTE exit (sellers)
            if action == 's' and exit_dte > 0 and dte_left <= exit_dte:
                exit_px     = curr_px
                exit_date   = dates[j]
                exit_reason = f'{exit_dte} DTE'
                break

        if exit_px is None:
            S_exp     = prices[expiry_i]
            intrinsic = max(0, S_exp - best_K) if opt_type == 'call' else max(0, best_K - S_exp)
            exit_px   = intrinsic
            exit_date = dates[expiry_i]
            exit_reason = 'expiry'

        # ── Transaction-cost realism (Muravyev & Pearson 2020, RFS) ─────────
        # Equity-option effective spread ≈ 2.2% of option value for traders
        # who time executions; we conservatively model HALF the quoted spread
        # as paid on entry AND on exit (full round-trip slippage), proxied as
        # 2.2% × premium per side because we don't have historical bid/ask.
        # Add TastyWorks-style commissions: $1 to open, free to close, plus
        # $0.03 reg fee/leg.
        _slip_per_side = 0.022   # 2.2% of premium per side (M&P 2020)
        _entry_eff     = best_entry_px * (1 + _slip_per_side) if action == 'b' \
                         else best_entry_px * (1 - _slip_per_side)
        _exit_eff      = exit_px * (1 - _slip_per_side) if action == 'b' \
                         else exit_px * (1 + _slip_per_side)
        gross_pnl      = ((_exit_eff - _entry_eff) if action == 'b'
                          else (_entry_eff - _exit_eff)) * 100
        _comm = commission_cost(n_legs=1, n_contracts=1, open_or_close='open') \
              + commission_cost(n_legs=1, n_contracts=1, open_or_close='close')
        pnl = gross_pnl - _comm
        equity += pnl

        trades.append({
            'Entry Date':    dates[ei].strftime('%Y-%m-%d') if hasattr(dates[ei], 'strftime') else str(dates[ei])[:10],
            'Exit Date':     exit_date.strftime('%Y-%m-%d') if hasattr(exit_date, 'strftime') else str(exit_date)[:10],
            'Direction':     opt_type.upper(),
            'Strike':        round(best_K, 2),
            'Spot @ Entry':  round(S, 2),
            'Entry Price':   round(best_entry_px, 3),
            'Exit Price':    round(exit_px, 3),
            'P&L ($)':       round(pnl, 2),
            'Cumul P&L ($)': round(equity, 2),
            'Exit Reason':   exit_reason,
            'Delta':         round(best_delta, 3),
            'Prob ITM (%)':  round(best_p_itm * 100, 1),
            'HAR-RV (%)':    round(hv * 100, 1),
            'VRP (%)':       round(vrp * 100, 1),
        })

    if not trades:
        print("No trades generated — try a larger lookback or different DTE.")
        return

    df = pd.DataFrame(trades)

    # ── Performance metrics ────────────────────────────────────────────────────
    wins     = df[df['P&L ($)'] > 0]
    losses   = df[df['P&L ($)'] <= 0]
    win_rate = len(wins) / len(df) * 100
    avg_win  = wins['P&L ($)'].mean()  if len(wins)   > 0 else 0.0
    avg_loss = losses['P&L ($)'].mean() if len(losses) > 0 else 0.0
    pf_denom = abs(losses['P&L ($)'].sum())
    profit_factor = wins['P&L ($)'].sum() / pf_denom if pf_denom > 0 else float('inf')
    total_pnl = df['P&L ($)'].sum()

    cum = df['P&L ($)'].cumsum()
    max_dd = (cum - cum.cummax()).min()

    consec_loss = max_consec = cur = 0
    for p in df['P&L ($)']:
        cur = cur + 1 if p <= 0 else 0
        max_consec = max(max_consec, cur)

    action_str = 'BUY' if action == 'b' else 'SELL'
    tp_label   = f"TP {int(take_profit*100)}%" if take_profit else "no TP"
    be_label   = f"BE @{int(be_trigger*100)}%" if be_trigger  else "no BE"
    sl_label   = f"SL {int(stop_loss*100)}%"  if stop_loss  else "no SL"
    dte_label  = f"{exit_dte} DTE" if exit_dte else "hold to expiry"
    print(f"\n{'='*60}")
    print(f"  BACKTEST — {symbol}  |  {action_str}  |  {lookback}d  |  DTE {dte_target}")
    print(f"  Exit rules: {tp_label}  |  {be_label}  |  {sl_label}  |  {dte_label}")
    print(f"{'='*60}")
    print(f"  Trades:            {len(df)}")
    print(f"  Win Rate:          {win_rate:.1f}%  ({len(wins)}W / {len(losses)}L)")
    print(f"  Avg Win:           ${avg_win:,.2f}")
    print(f"  Avg Loss:          ${avg_loss:,.2f}")
    print(f"  Profit Factor:     {profit_factor:.2f}  (>1.5 is good)")
    print(f"  Total P&L:         ${total_pnl:,.2f}")
    print(f"  Max Drawdown:      ${max_dd:,.2f}")
    print(f"  Max Consec Losses: {max_consec}")
    print(f"{'─'*60}")
    print(f"  Exit breakdown:")
    for reason, cnt in df['Exit Reason'].value_counts().items():
        print(f"    {reason:20s} {cnt:3d}  ({cnt/len(df)*100:.0f}%)")
    print(f"{'='*60}")

    # ── Fractional Kelly position sizing (arXiv 2025) ─────────────────────────
    account_size = float(input("\n  Your account size ($) for Kelly sizing (press Enter to skip): ").strip() or 0)
    if account_size > 0:
        avg_entry_px = df['Entry Price'].mean()
        vts_now      = vix_term_structure()
        vix_now      = vts_now['VIX'] if vts_now else 20.0
        kelly        = kelly_contracts(win_rate / 100, avg_win, abs(avg_loss),
                                       account_size, avg_entry_px,
                                       fraction=0.25, vix=vix_now)
        print(f"\n  ── Fractional Kelly Sizing (25% Kelly) ──────────────────")
        print(f"  Account:           ${account_size:,.2f}")
        print(f"  Full Kelly f*:     {kelly['kelly_f']:.3f}  ({kelly['kelly_f']*100:.1f}% of account)")
        print(f"  25% Fractional:    {kelly['kelly_frac']:.3f}  ({kelly['risk_pct']:.1f}% of account)")
        if vix_now > 30:
            print(f"  VIX {vix_now:.0f} > 30 → fraction halved to 12.5% (high-vol regime)")
        print(f"  Recommended:       {kelly['contracts']} contract(s)  "
              f"(${avg_entry_px*100*kelly['contracts']:,.0f} at risk)")
        print(f"  Source: arXiv 2025 — Kelly+VIX hybrid for index options")
        print(f"  {'─'*50}")

    print(f"\n  Trade Log (last 10):")
    print(df[['Entry Date','Direction','Strike','Entry Price','Exit Price',
              'P&L ($)','Cumul P&L ($)','Exit Reason']].tail(10).to_string(index=False))

    # ── Excel export ───────────────────────────────────────────────────────────
    fname = f"{symbol}_backtest_{lookback}d.xlsx"
    perf_rows = [{
        'Metric': k, 'Value': v
    } for k, v in {
        'Symbol':             symbol,
        'Action':             action_str,
        'Lookback (days)':    lookback,
        'DTE Target':         dte_target,
        'Take Profit':        tp_label,
        'Break-Even Trigger': be_label,
        'Stop Loss':          sl_label,
        'DTE Exit':           f'{exit_dte} DTE' if exit_dte else 'hold to expiry',
        'Total Trades':       len(df),
        'Win Rate (%)':       round(win_rate, 1),
        'Avg Win ($)':        round(avg_win, 2),
        'Avg Loss ($)':       round(avg_loss, 2),
        'Profit Factor':      round(profit_factor, 2),
        'Total P&L ($)':      round(total_pnl, 2),
        'Max Drawdown ($)':   round(max_dd, 2),
        'Max Consec Losses':  max_consec,
        'NOTE':               'Synthetic BS prices — no real bid/ask history',
    }.items()]

    with pd.ExcelWriter(fname, engine='openpyxl') as writer:
        pd.DataFrame(perf_rows).to_excel(writer, index=False, sheet_name='Performance')
        df.to_excel(writer, index=False, sheet_name='Trade Log')
        df[['Entry Date','P&L ($)','Cumul P&L ($)']].to_excel(
            writer, index=False, sheet_name='Equity Curve')
        for ws in writer.sheets.values():
            for col in ws.columns:
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(
                    max((len(str(c.value)) for c in col if c.value), default=8) + 3, 28)
            for cell in ws[1]:
                cell.font = Font(bold=True)

    print(f"\nSaved: {fname}")

    try:
        from google.colab import files
        files.download(fname)
    except ImportError:
        pass


# ─── Market Scanner ───────────────────────────────────────────────────────────

# Fallback curated list (used when all live sources fail)
_SCAN_FALLBACK = [
    'AAPL','MSFT','NVDA','GOOGL','AMZN','META','TSLA','AVGO','ORCL','AMD',
    'JPM','GS','BAC','MS','V','MA','C','WFC','AXP','BRK-B',
    'UNH','LLY','JNJ','MRK','ABBV','PFE','AMGN','MRNA','GILD','BIIB',
    'XOM','CVX','BA','CAT','DE','GE','RTX','HON','LMT','HAL',
    'NFLX','DIS','SBUX','MCD','NKE','COST','WMT','TGT','HD','LOW',
    'PLTR','SOFI','GME','MARA','RIOT','COIN','HOOD','RDDT','IREN','ONDS',
    'SPY','QQQ','IWM','GLD','TLT','XLF','XLE','ARKK','SOXL','UVXY',
]


def _fetch_scan_universe(limit: int = 500) -> list:
    """
    Build a dynamic universe of optionable stocks from 3 live sources:

    Source 1 — CBOE most-active equity options (real-time, ~top 50 by volume)
      CBOE publishes the day's most-traded option contracts via their public
      market-statistics API. We extract the underlying symbols and keep only
      pure equity tickers (no index, no ETN, length ≤ 5 chars).

    Source 2 — S&P 500 constituent list from Wikipedia
      Parsed with pandas read_html. Gives ~500 large-caps that virtually all
      have liquid listed options. Symbols normalised (. → -) for yfinance.

    Source 3 — Nasdaq-100 from Wikipedia
      Additional 100 growth/tech names with very high options volume.

    All three are merged, deduplicated, and capped at `limit` symbols.
    Falls back to the 70-stock curated list if every live source fails.

    This means the scanner covers:
      - ALL of today's most-actively traded options (CBOE source)
      - ALL S&P 500 members (Wikipedia source)
      - ALL Nasdaq-100 members (Wikipedia source)
      - Your curated high-vol retail favourites (fallback always included)
    """
    syms: set[str] = set(_SCAN_FALLBACK)   # always include curated as base

    # ── Source 1: CBOE most-active equity options ──────────────────────────
    _CBOE_MA_URLS = [
        'https://cdn.cboe.com/api/global/delayed_quotes/most_active_equities.json',
        'https://www.cboe.com/us/options/market_statistics/most_active/?mkt=cone',
    ]
    for url in _CBOE_MA_URLS:
        try:
            r = requests.get(url, timeout=8,
                             headers={'User-Agent': 'Mozilla/5.0',
                                      'Accept': 'application/json'})
            if not r.ok:
                continue
            try:
                data = r.json()
            except Exception:
                continue
            # The JSON shape varies by endpoint; try common key paths
            for key in ('data', 'most_active', 'results', 'quotes'):
                items = data.get(key, [])
                if items:
                    for item in items:
                        # underlying symbol may be under 'symbol', 'name', 'underlying'
                        for field in ('symbol', 'name', 'underlying', 'ticker'):
                            raw = str(item.get(field, '')).strip().upper()
                            raw = raw.split()[0]          # drop suffix like "C" or "P"
                            raw = raw.split('/')[0]       # drop exchange suffix
                            if raw.isalpha() and 1 <= len(raw) <= 5:
                                syms.add(raw)
                    break
            if len(syms) > len(_SCAN_FALLBACK) + 5:
                break   # got useful data from this URL
        except Exception:
            continue

    # ── Source 2: S&P 500 from Wikipedia ──────────────────────────────────
    try:
        sp_tables = pd.read_html(
            'https://en.wikipedia.org/wiki/List_of_S%26P_500_companies',
            attrs={'id': 'constituents'}
        )
        for sym in sp_tables[0]['Symbol'].tolist():
            syms.add(str(sym).replace('.', '-').strip().upper())
        print(f"  [Universe] S&P 500 loaded ({len(sp_tables[0])} symbols)")
    except Exception:
        pass

    # ── Source 3: Nasdaq-100 from Wikipedia ───────────────────────────────
    try:
        ndx_tables = pd.read_html(
            'https://en.wikipedia.org/wiki/Nasdaq-100',
            attrs={'id': 'constituents'}
        )
        for sym in ndx_tables[0]['Ticker'].tolist():
            syms.add(str(sym).replace('.', '-').strip().upper())
        print(f"  [Universe] Nasdaq-100 loaded")
    except Exception:
        pass

    # Remove obviously invalid tokens (indices, etc.)
    clean = sorted(s for s in syms if s.isalpha() or (len(s) <= 5 and '-' in s))
    result = clean[:limit]
    print(f"  [Universe] Total unique symbols: {len(result)}")
    return result


def _quick_options_filter(symbols: list, min_atm_oi: int = 50) -> list:
    """
    Fast pre-filter: keep only symbols where the nearest expiry ATM option has OI > min_atm_oi.
    One API call per symbol — much cheaper than full scan.
    Runs in parallel-ish by batching and skipping on first failure.
    """
    today = datetime.now().replace(tzinfo=None)
    liquid = []
    for sym in symbols:
        try:
            exps = _ticker(sym).options
            if not exps:
                continue
            # Find first expiry 7–45 DTE
            for exp in exps[:6]:
                dte = (datetime.strptime(exp, '%Y-%m-%d') - today).days
                if 7 <= dte <= 45:
                    ch = _ticker(sym).option_chain(exp)
                    # Check ATM OI (just check first 3 near-ATM calls/puts)
                    total_oi = int((ch.calls['openInterest'].fillna(0)
                                   + ch.puts['openInterest'].fillna(0)).sum())
                    total_vol = int((ch.calls.get('volume', pd.Series([0])).fillna(0)
                                    + ch.puts.get('volume', pd.Series([0])).fillna(0)).sum())
                    if total_oi >= min_atm_oi or total_vol > 0:
                        liquid.append(sym)
                    break
        except Exception:
            continue
    return liquid


def market_scanner(budget: float = 5.00, max_stocks: int = 150,
                   min_score: float = 30.0, watchlist: list | None = None,
                   prefilter: bool = True):
    """
    Scan the market for the best call/put buying opportunities.

    Universe (when watchlist=None):
      - CBOE most-active equity options today (real-time top by volume)
      - All S&P 500 constituents (Wikipedia)
      - All Nasdaq-100 constituents (Wikipedia)
      - Curated high-vol retail list (always included)
      → ~600 symbols, pre-filtered to those with real options OI/volume
      → Top max_stocks by options activity are deep-scanned

    Scoring model (max 100 pts) using 5 quant signals:

    Signal 1 — VRP / IV-HV spread (Carr & Wu 2009):
        IV < HV → options cheap vs realised vol → buyers edge
        +30 if IV-HV < −10%, +15 if < 0%, −20 if > +10%

    Signal 2 — GEX regime (Bollen & Whaley 2004 / SpotGamma):
        Negative GEX → dealers short gamma → trending regime → buyers edge
        +20 if NEG GEX, −10 if POS GEX

    Signal 3 — GARCH(1,1) vs IV (Bollerslev 1986):
        GARCH > IV → vol expected to rise → buy options
        +15 if GARCH > IV×1.05, −10 if GARCH < IV×0.95

    Signal 4 — HAR-RV forecast vs IV (Corsi 2009):
        +10 if HAR-RV > IV×1.02

    Signal 5 — Liquidity:
        +10 if ATM OI > 500, +5 if 100–500, −15 if < 100

    Bonus:   +10 if ATM premium ≤ budget
    Penalty: −25 if no GEX data
    """
    if watchlist:
        universe = list(dict.fromkeys(watchlist))
        print(f"\n  [Universe] Custom watchlist: {len(universe)} symbols")
    else:
        print("\n  [Universe] Building dynamic universe from live sources...")
        universe = _fetch_scan_universe(limit=600)

    r_base = get_risk_free_rate(0.25)

    # ── Pre-filter: only keep symbols with real options OI/volume ─────────────
    if prefilter and len(universe) > max_stocks:
        print(f"  [Pre-filter] Checking options activity on {len(universe)} symbols"
              f" — keeping top {max_stocks} with live OI/volume...")
        liquid = _quick_options_filter(universe, min_atm_oi=50)
        print(f"  [Pre-filter] {len(liquid)} symbols passed (have options with OI ≥ 50)")
        universe = liquid[:max_stocks]
    else:
        universe = universe[:max_stocks]

    print(f"\n{'='*62}")
    print(f"  MARKET SCANNER — Best Options Opportunities")
    print(f"{'='*62}")
    print(f"  Scanning: {len(universe)} symbols  |  Budget: ${budget:.2f}/contract")
    print(f"  Signals: IV-HV · GEX regime · GARCH · HAR-RV · Liquidity")
    print(f"{'─'*62}")

    results = []
    for i, sym in enumerate(universe, 1):
        bar = ('█' * int(i / len(universe) * 30)).ljust(30)
        print(f"  [{bar}] {i:>3}/{len(universe)}  {sym:<8}", end='\r', flush=True)

        try:
            # ── Spot + HV ──────────────────────────────────────────────────
            hist = _ticker(sym).history(period='1y')
            if len(hist) < 30:
                continue
            S  = float(hist['Close'].iloc[-1])
            q  = get_dividend_yield(sym)
            log_rets = list(np.log(hist['Close'] / hist['Close'].shift(1)).dropna())
            hv = float(np.std(log_rets) * np.sqrt(252))

            # ── Nearest ATM IV ────────────────────────────────────────────
            exps = _ticker(sym).options
            if not exps:
                continue
            # Pick first expiry with DTE 7–45 days
            chosen_exp, chosen_T, chain_calls, chain_puts = None, None, None, None
            today = datetime.now().replace(tzinfo=None)
            for exp in exps[:8]:
                exp_dt = datetime.strptime(exp, '%Y-%m-%d')
                dte = (exp_dt - today).days
                if 7 <= dte <= 45:
                    try:
                        ch = _ticker(sym).option_chain(exp)
                        chosen_exp = exp
                        chosen_T   = max(dte / 365.0, 7/365.0)
                        chain_calls = ch.calls
                        chain_puts  = ch.puts
                    except Exception:
                        pass
                    break
            if chosen_exp is None:
                continue

            # ATM call and put (nearest strike to spot)
            atm_call_row = chain_calls.iloc[(chain_calls['strike'] - S).abs().argsort().iloc[0]] if not chain_calls.empty else None
            atm_put_row  = chain_puts.iloc[(chain_puts['strike']  - S).abs().argsort().iloc[0]] if not chain_puts.empty else None

            if atm_call_row is None and atm_put_row is None:
                continue

            # ATM IV (average of call and put IV for robustness)
            atm_ivs = []
            for row, otype in [(atm_call_row, 'call'), (atm_put_row, 'put')]:
                if row is None:
                    continue
                p = mid_price(row)
                if p > 0:
                    iv = implied_volatility(S, row['strike'], chosen_T, r_base, p, otype, q)
                    if not np.isnan(iv):
                        atm_ivs.append(iv)
            if not atm_ivs:
                continue
            atm_iv = float(np.mean(atm_ivs))

            # ATM premium (call for call signal, put for put signal — use lower)
            atm_call_prem = mid_price(atm_call_row) if atm_call_row is not None else np.nan
            atm_put_prem  = mid_price(atm_put_row)  if atm_put_row  is not None else np.nan
            atm_prem_min  = float(np.nanmin([atm_call_prem, atm_put_prem]))

            # ATM OI
            atm_oi_c = int(atm_call_row.get('openInterest', 0) or 0) if atm_call_row is not None else 0
            atm_oi_p = int(atm_put_row.get('openInterest', 0)  or 0) if atm_put_row  is not None else 0
            atm_oi   = max(atm_oi_c, atm_oi_p)

            # ── GARCH forecast ────────────────────────────────────────────
            garch_fc = garch_vol_forecast(log_rets, horizon=21)
            garch_vol = garch_fc['vol'] if garch_fc else None

            # ── HAR-RV forecast ───────────────────────────────────────────
            prices_list = list(hist['Close'].dropna())
            har_vol = har_rv_forecast(prices_list)

            # ── GEX signal ────────────────────────────────────────────────
            gex = gex_signal(sym, S, r_base, hv, q)

            # ── SCORING ───────────────────────────────────────────────────
            score   = 0.0
            reasons = []

            # Signal 1: VRP / IV-HV
            iv_hv_diff = (atm_iv - hv) * 100
            if iv_hv_diff < -10:
                score += 30
                reasons.append(f"IV very cheap ({iv_hv_diff:+.1f}%)")
            elif iv_hv_diff < 0:
                score += 15
                reasons.append(f"IV cheap ({iv_hv_diff:+.1f}%)")
            elif iv_hv_diff > 10:
                score -= 20
                reasons.append(f"IV expensive ({iv_hv_diff:+.1f}%)")

            # Signal 2: GEX regime
            direction = 'AUTO'
            if gex:
                if gex['regime'] == 'NEGATIVE':
                    score += 20
                    reasons.append("neg GEX (trend)")
                else:
                    score -= 10
                    reasons.append("pos GEX (pin)")
                direction = gex['suggested'].upper()
            else:
                score -= 25
                reasons.append("no GEX data")
                # fallback direction: trend-follow (HAR > HV → expecting vol → no clear dir)
                direction = 'CALL' if S > (S * 0.98) else 'PUT'   # neutral placeholder

            # Signal 3: GARCH vs IV
            if garch_vol:
                if garch_vol > atm_iv * 1.05:
                    score += 15
                    reasons.append(f"GARCH↑ ({garch_vol*100:.0f}%>IV)")
                elif garch_vol < atm_iv * 0.95:
                    score -= 10
                    reasons.append(f"GARCH↓ ({garch_vol*100:.0f}%<IV)")

            # Signal 4: HAR-RV vs IV
            if har_vol and not np.isnan(har_vol):
                if har_vol > atm_iv * 1.02:
                    score += 10
                    reasons.append(f"HAR↑ ({har_vol*100:.0f}%)")

            # Signal 5: Liquidity
            if atm_oi > 500:
                score += 10
            elif atm_oi > 100:
                score += 5
            else:
                score -= 15
                reasons.append("low OI")

            # Budget bonus
            if atm_prem_min <= budget:
                score += 10
                reasons.append(f"fits ${budget:.0f} budget")
            else:
                reasons.append(f"need ${atm_prem_min*100:.0f}/contract")

            results.append({
                'Symbol':          sym,
                'Spot ($)':        round(S, 2),
                'Direction':       direction,
                'ATM IV (%)':      round(atm_iv * 100, 1),
                'HV (%)':          round(hv * 100, 1),
                'IV-HV (%)':       round(iv_hv_diff, 1),
                'GARCH fcst (%)':  round(garch_vol * 100, 1) if garch_vol else np.nan,
                'HAR fcst (%)':    round(har_vol * 100, 1) if har_vol and not np.isnan(har_vol) else np.nan,
                'GEX Regime':      gex['regime'] if gex else 'N/A',
                'GEX Flip ($)':    gex['flip_level'] if gex else np.nan,
                'ATM OI':          atm_oi,
                'ATM Prem ($)':    round(atm_prem_min, 2),
                'Score':           round(score, 1),
                'Expiry':          chosen_exp,
                'DTE':             (datetime.strptime(chosen_exp, '%Y-%m-%d') - today).days,
                'Signals':         ' | '.join(reasons),
            })

        except Exception:
            continue

    print(f"\n{'─'*62}")

    if not results:
        print("  No results found. Check internet connection or widen universe.")
        return

    df = pd.DataFrame(results).sort_values('Score', ascending=False).reset_index(drop=True)
    df_good = df[df['Score'] >= min_score]

    # ── Console output ────────────────────────────────────────────────────────
    print(f"\n  TOP OPPORTUNITIES (score ≥ {min_score:.0f})")
    print(f"{'─'*62}")
    display_cols = ['Symbol','Spot ($)','Direction','ATM IV (%)','HV (%)','IV-HV (%)','GEX Regime','Score','ATM Prem ($)','Expiry']
    top_show = df_good[display_cols].head(15) if not df_good.empty else df[display_cols].head(5)
    print(top_show.to_string(index=True))

    print(f"\n{'─'*62}")
    print(f"  SCORECARD — #1 Ranked Setup")
    print(f"{'─'*62}")
    best = df.iloc[0]
    print(f"  {best['Direction']}  {best['Symbol']}  @ ${best['Spot ($)']:.2f}")
    print(f"  Expiry: {best['Expiry']} ({best['DTE']} DTE)  |  ATM prem: ${best['ATM Prem ($)']:.2f}/share = ${best['ATM Prem ($)']*100:.0f}/contract")
    print(f"  ATM IV {best['ATM IV (%)']:.1f}%  vs  HV {best['HV (%)']:.1f}%  →  {best['IV-HV (%)']:+.1f}% spread")
    print(f"  GARCH: {best['GARCH fcst (%)']:.1f}%  |  HAR-RV: {best['HAR fcst (%)']:.1f}%  |  GEX: {best['GEX Regime']}")
    print(f"  Score: {best['Score']:.0f}/100  |  Signals: {best['Signals']}")
    print(f"\n  → Run Trade Finder (option 2) on {best['Symbol']} to get exact strike + expiry.")
    print(f"{'='*62}\n")

    # ── Excel ─────────────────────────────────────────────────────────────────
    fname = f"market_scan_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    with pd.ExcelWriter(fname, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='All Results')
        if not df_good.empty:
            df_good.to_excel(writer, index=False, sheet_name='Top Picks')
        # Score breakdown
        df[['Symbol','Direction','Score','Signals']].to_excel(writer, index=False, sheet_name='Score Breakdown')
    print(f"  Full scan saved: {fname}")

    try:
        from google.colab import files
        files.download(fname)
    except ImportError:
        pass

    return df


# ─── Scanner Backtest ─────────────────────────────────────────────────────────

def backtest_scanner(watchlist=None, lookback_days=252, holding_days=14,
                     scan_freq=14, top_n=3, account=1000.0,
                     take_profit=0.50, stop_loss=0.60, min_score=35.0,
                     budget=500.0):
    """
    Improved scanner backtest v2.

    Key fixes vs v1:
      1. Hard premium cap — skip if prem*100 > min(budget, acct*20%). Never
         forces a trade bigger than the account can absorb.
      2. Always 1 contract. No max(1,...) that bought 13 LLY contracts on $190.
      3. Stop widened to 60% + 7-DTE time-stop + breakeven stop after +25%.
      4. Vol-cheap gate — only enter if GARCH & HAR both forecast higher vol
         than current HV_30 (proxy for IV cheap relative to expected RV).
      5. Vol-percentile filter — skip if HV_30 > 70th-pct of its 252d range
         (buying vol at the highs is bad).
      6. SPY regime filter — skip calls when SPY 10d HV > 30d HV × 1.3
         (market in stress; calls get crushed by vol collapse on bounces).
      7. Direction requires 3-period alignment (3d, 5d, 20d all same sign).
         Ambiguous setups are skipped entirely.
      8. Min score = 35; never force trades when nothing qualifies.
      9. Anti-correlation — cap at 2 picks in the same direction per scan date.
    """
    wl = watchlist or _SCAN_FALLBACK
    r  = 0.05
    dte_target = max(holding_days + 14, 30)

    end_dt = datetime.now().replace(tzinfo=None)

    print(f"\n{'='*62}")
    print(f"  SCANNER BACKTEST  (v2 — improved model)")
    print(f"{'='*62}")
    print(f"  Universe:    {len(wl)} symbols")
    print(f"  Period:      last {lookback_days} calendar days")
    print(f"  Scan every:  {scan_freq} trading days")
    print(f"  Hold:        max {holding_days} days  (7-DTE time stop)")
    print(f"  Top N/scan:  {top_n}  |  TP: {take_profit*100:.0f}%  SL: {stop_loss*100:.0f}%")
    print(f"  Account:     ${account:,.2f}  |  Max cost/trade: ${budget:.2f}")
    print(f"  Min score:   {min_score}  |  Vol-cheap gate: ON  |  SPY regime: ON")
    print(f"{'─'*62}")

    # ── Fetch all historical prices upfront ───────────────────────────────────
    print("  Fetching historical prices...")
    price_data: dict[str, pd.Series] = {}
    symbols_needed = list(set(wl) | {'SPY'})
    for i, sym in enumerate(symbols_needed, 1):
        print(f"    {i:>3}/{len(symbols_needed)}  {sym:<8}", end='\r', flush=True)
        try:
            hist = _ticker(sym).history(period='2y')
            if len(hist) >= 90:
                price_data[sym] = hist['Close'].dropna()
        except Exception:
            pass
    print(f"\n  Loaded price data: {len(price_data)} symbols")

    if len(price_data) < 3:
        print("  Not enough data to backtest.")
        return

    # ── Build scan dates ──────────────────────────────────────────────────────
    ref_sym   = 'SPY' if 'SPY' in price_data else list(price_data.keys())[0]
    all_dates = price_data[ref_sym].index
    all_dates_naive = [d.replace(tzinfo=None) if hasattr(d, 'tzinfo') else d
                       for d in all_dates]
    cutoff = end_dt - pd.Timedelta(days=lookback_days)
    scan_indices = [i for i, d in enumerate(all_dates_naive)
                    if d >= cutoff and i + holding_days + 10 < len(all_dates)]
    scan_indices = scan_indices[::scan_freq]

    if not scan_indices:
        print("  No valid scan dates in lookback window.")
        return

    print(f"  Scan dates: {len(scan_indices)}  "
          f"({all_dates[scan_indices[0]].strftime('%Y-%m-%d')} → "
          f"{all_dates[scan_indices[-1]].strftime('%Y-%m-%d')})")
    print(f"{'─'*62}")

    # ── Main backtest loop ────────────────────────────────────────────────────
    trades         = []
    equity         = [account]
    acct           = account
    max_acct       = account
    max_dd         = 0.0
    skipped_afford = 0
    skipped_volpct = 0
    skipped_regime = 0
    skipped_gate   = 0

    for scan_idx in scan_indices:
        scan_date = all_dates[scan_idx]

        # ── SPY market regime check ───────────────────────────────────────────
        spy_bearish = False
        if 'SPY' in price_data:
            try:
                spy_sub = list(price_data['SPY'][price_data['SPY'].index <= scan_date].values)
                if len(spy_sub) >= 35:
                    spy_lr  = [np.log(spy_sub[k] / spy_sub[k-1]) for k in range(1, len(spy_sub))]
                    spy_h10 = float(np.std(spy_lr[-10:]) * np.sqrt(252))
                    spy_h30 = float(np.std(spy_lr[-30:]) * np.sqrt(252))
                    spy_bearish = spy_h10 > spy_h30 * 1.30
            except Exception:
                pass

        # ── Score every symbol ────────────────────────────────────────────────
        candidates = []
        for sym, prices in price_data.items():
            if sym == 'SPY':
                continue
            try:
                px = prices[prices.index <= scan_date]
                if len(px) < 90:
                    continue
                px_list = list(px.values)
                S_entry = float(px_list[-1])

                log_rets = [np.log(px_list[k] / px_list[k-1])
                            for k in range(1, len(px_list))]
                hv_10  = float(np.std(log_rets[-10:])  * np.sqrt(252))
                hv_30  = float(np.std(log_rets[-30:])  * np.sqrt(252))
                hv_252 = float(np.std(log_rets[-252:]) * np.sqrt(252)) if len(log_rets) >= 252 else hv_30

                # Vol percentile filter — skip if HV_30 already at highs
                if len(log_rets) >= 282:
                    hv30_hist = [float(np.std(log_rets[j-30:j]) * np.sqrt(252))
                                 for j in range(60, len(log_rets), 10)]
                    if hv30_hist and hv_30 > float(np.percentile(hv30_hist, 70)):
                        skipped_volpct += 1
                        continue

                har = har_rv_forecast(px_list[-120:])
                if np.isnan(har):
                    har = hv_30

                garch_res     = garch_vol_forecast(log_rets[-252:], horizon=21)
                garch_vol_est = garch_res['vol'] if garch_res else hv_30

                # Vol-cheap gate — both forecasters must say vol will rise
                forecast_avg = (garch_vol_est + har) / 2.0
                if forecast_avg <= hv_30 * 1.02:
                    skipped_gate += 1
                    continue

                # Momentum with 3-period alignment for direction
                ret_3d  = (px_list[-1] / px_list[-4]  - 1) if len(px_list) >= 4  else 0
                ret_5d  = (px_list[-1] / px_list[-6]  - 1) if len(px_list) >= 6  else 0
                ret_20d = (px_list[-1] / px_list[-21] - 1) if len(px_list) >= 21 else 0

                if   ret_3d > 0 and ret_5d > 0.01 and ret_20d > 0:
                    direction = 'call'
                elif ret_3d < 0 and ret_5d < -0.01 and ret_20d < 0:
                    direction = 'put'
                elif abs(ret_5d) > abs(ret_20d) * 2 and abs(ret_5d) > 0.03:
                    direction = 'call' if ret_5d > 0 else 'put'
                else:
                    continue   # ambiguous direction → skip

                # SPY regime: penalise calls in stress (not hard-skip — valid setups still pass min_score)
                spy_regime_penalty = -20 if (spy_bearish and direction == 'call') else 0
                if spy_regime_penalty < 0:
                    skipped_regime += 1   # count but don't skip

                # Scoring
                vrp_proxy    = (forecast_avg - hv_30) * 100
                mom_strength = abs(ret_5d) + abs(ret_20d) * 0.5
                score = float(spy_regime_penalty)   # start from regime adjustment

                if vrp_proxy > 10:
                    score += 35
                elif vrp_proxy > 5:
                    score += 25
                elif vrp_proxy > 2:
                    score += 15

                if garch_vol_est > har * 1.10:
                    score += 20
                elif garch_vol_est > har * 1.05:
                    score += 12

                if mom_strength > 0.07:
                    score += 20
                elif mom_strength > 0.04:
                    score += 12
                elif mom_strength > 0.02:
                    score += 6

                if hv_10 > hv_30 * 1.20:
                    score += 10   # vol just spiked — early-move signal

                if 0.25 <= hv_30 <= 0.80:
                    score += 5
                elif hv_30 < 0.15:
                    score -= 20   # too quiet — theta will kill position

                candidates.append({
                    'sym': sym, 'S': S_entry, 'hv': hv_30,
                    'har': har, 'garch': garch_vol_est,
                    'direction': direction, 'score': score,
                    'ret_5d': ret_5d, 'ret_20d': ret_20d,
                    'vrp': vrp_proxy,
                })
            except Exception:
                continue

        if not candidates:
            continue

        candidates.sort(key=lambda x: x['score'], reverse=True)
        picks = [c for c in candidates if c['score'] >= min_score][:top_n]
        if not picks:
            continue   # nothing qualifies — skip this scan date

        # Anti-correlation: cap 2 in same direction per scan
        for direction_cap in ('call', 'put'):
            same = [p for p in picks if p['direction'] == direction_cap]
            if len(same) > 2:
                other = [p for p in picks if p['direction'] != direction_cap]
                picks = (same[:2] + other)[:top_n]

        # ── Simulate each pick ────────────────────────────────────────────────
        for pick in picks:
            sym       = pick['sym']
            S0        = pick['S']
            # B7 fix: enter the trade at the forecast (forward-looking) vol
            # rather than 30d historical. Otherwise vol-cheap-gate signals
            # don't actually translate to cheaper premium at entry — and the
            # "edge" you score on never appears in P&L. Use the avg of the
            # GARCH and HAR forecasts (already computed in the scoring pass).
            sig0      = float((pick['garch'] + pick['har']) / 2.0)
            T0        = dte_target / 365.0
            direction = pick['direction']
            # Snap ATM strike to a real OCC grid increment (B3 fix).
            K0        = nearest_listed_strike(S0, S0)
            prem      = bs_price(S0, K0, T0, r, sig0, direction)
            if prem <= 0 or np.isnan(prem):
                continue

            contract_cost = prem * 100

            # Hard affordability gate — never let one trade exceed 20% of account
            max_per_trade = min(budget, acct * 0.20)
            if contract_cost > max_per_trade:
                skipped_afford += 1
                continue

            n_contracts = 1   # always 1 — no leverage on a small account
            cost        = contract_cost

            px_fwd_idx = scan_idx + 1
            try:
                px_fwd = price_data[sym].iloc[px_fwd_idx : px_fwd_idx + holding_days + 10]
            except Exception:
                continue
            if len(px_fwd) < 2:
                continue

            # Daily simulation
            entry_val      = prem
            exit_val       = None
            exit_day       = 0
            exit_reason    = 'expiry'
            be_stop_active = False
            final_pnl_pct  = 0.0

            # B2 partial fix: mean-reverting log-OU IV path so vega isn't zero
            rng_iv = np.random.default_rng(seed=scan_idx ^ hash(sym) & 0xFFFFFFFF)
            kappa, theta_iv, eta = 1.0/21.0, np.log(sig0), 0.30
            log_iv_path = [np.log(sig0)]
            for _ in range(len(px_fwd)):
                prev = log_iv_path[-1]
                log_iv_path.append(
                    prev + kappa * (theta_iv - prev) + eta * np.sqrt(1/252) * rng_iv.standard_normal()
                )

            for day_i, (date_i, S_i) in enumerate(px_fwd.items()):
                dte_rem = dte_target - (day_i + 1)
                T_rem   = max(dte_rem / 365.0, 0.5 / 365.0)

                # Sticky-strike vol surface tilt + stochastic IV path
                dist     = abs(float(S_i) - S0) / S0
                sig_base = float(np.clip(np.exp(log_iv_path[day_i + 1]), 0.05, 3.0))
                sig_i    = sig_base * (1 + 0.20 * dist)
                val_i    = bs_price(float(S_i), K0, T_rem, r, sig_i, direction)
                if np.isnan(val_i) or val_i <= 0:
                    val_i = 0.01

                pnl_pct = (val_i - entry_val) / entry_val
                final_pnl_pct = pnl_pct

                # Activate breakeven stop once up 25%
                if pnl_pct >= 0.25 and not be_stop_active:
                    be_stop_active = True

                # 7-DTE time stop — exit before gamma/theta crush in final week
                if dte_rem <= 7:
                    exit_val    = val_i
                    exit_day    = day_i + 1
                    exit_reason = '7-DTE stop'
                    break

                if pnl_pct >= take_profit:
                    exit_val    = val_i
                    exit_day    = day_i + 1
                    exit_reason = f'TP +{take_profit*100:.0f}%'
                    break

                stop_level = 0.0 if be_stop_active else -stop_loss
                if pnl_pct <= stop_level:
                    exit_val    = val_i
                    exit_day    = day_i + 1
                    exit_reason = 'BE stop' if be_stop_active else f'SL -{stop_loss*100:.0f}%'
                    break

                if day_i + 1 >= holding_days:
                    exit_val    = val_i
                    exit_day    = day_i + 1
                    exit_reason = 'hold end'
                    break

            if exit_val is None:
                exit_val    = 0.01
                exit_reason = 'expired'

            # ── Transaction-cost realism (Muravyev & Pearson 2020) ──────────
            # 2.2% effective spread per side + TW commissions ($1 open, free
            # close, $0.03 reg fee/leg). Without this, retail backtests are
            # systematically optimistic by 5-15% of P&L.
            _slip_side = 0.022
            _e_eff     = entry_val * (1 + _slip_side)     # buyer pays ask-ish
            _x_eff     = exit_val  * (1 - _slip_side)     # closer hits bid-ish
            gross_pnl  = (_x_eff - _e_eff) * 100 * n_contracts
            _comm = commission_cost(n_legs=1, n_contracts=n_contracts, open_or_close='open') \
                  + commission_cost(n_legs=1, n_contracts=n_contracts, open_or_close='close')
            pnl_total = gross_pnl - _comm
            acct     += pnl_total
            max_acct  = max(max_acct, acct)
            dd        = (max_acct - acct) / max_acct * 100 if max_acct > 0 else 0
            max_dd    = max(max_dd, dd)
            equity.append(acct)

            trades.append({
                'Scan Date':        scan_date.strftime('%Y-%m-%d') if hasattr(scan_date, 'strftime') else str(scan_date)[:10],
                'Symbol':           sym,
                'Direction':        direction.upper(),
                'Entry Spot ($)':   round(S0, 2),
                'Entry Prem ($)':   round(entry_val, 4),
                'Exit Prem ($)':    round(exit_val, 4),
                'Cost ($)':         round(cost, 2),
                'P&L ($)':          round(pnl_total, 2),
                'P&L (%)':          round(final_pnl_pct * 100, 1),
                'Exit Day':         exit_day,
                'Exit Reason':      exit_reason,
                'Score':            round(pick['score'], 1),
                'VRP proxy':        round(pick['vrp'], 1),
                'HV (%)':           round(pick['hv'] * 100, 1),
                'GARCH (%)':        round(pick['garch'] * 100, 1),
                'HAR-RV (%)':       round(pick['har'] * 100, 1),
                '5d Ret (%)':       round(pick['ret_5d'] * 100, 2),
                'Acct Balance ($)': round(acct, 2),
            })

    # ── Filter report ──────────────────────────────────────────────────────────
    print(f"\n  Filter stats:")
    print(f"    Skipped (unaffordable):  {skipped_afford}")
    print(f"    Skipped (vol at highs):  {skipped_volpct}")
    print(f"    Skipped (vol not cheap): {skipped_gate}")
    print(f"    Skipped (SPY regime):    {skipped_regime}")

    # ── Results ────────────────────────────────────────────────────────────────
    if not trades:
        print("  No trades generated after filters.")
        print("  Try: lower min_score, higher budget, or more symbols in watchlist.")
        return

    df        = pd.DataFrame(trades)
    wins      = df[df['P&L ($)'] > 0]
    losses    = df[df['P&L ($)'] <= 0]
    win_rate  = len(wins) / len(df) * 100
    avg_win   = wins['P&L ($)'].mean()   if not wins.empty   else 0
    avg_loss  = losses['P&L ($)'].mean() if not losses.empty else 0
    gross_win = wins['P&L ($)'].sum()    if not wins.empty   else 0
    gross_los = losses['P&L ($)'].sum()  if not losses.empty else 0
    pf        = abs(gross_win / gross_los) if gross_los != 0 else np.inf
    total_pnl = df['P&L ($)'].sum()
    total_ret = (acct - account) / account * 100

    print(f"\n{'='*62}")
    print(f"  SCANNER BACKTEST RESULTS  (v2)")
    print(f"{'='*62}")
    print(f"  Trades:         {len(df)}")
    print(f"  Win Rate:       {win_rate:.1f}%  ({len(wins)}W / {len(losses)}L)")
    print(f"  Avg Win:        ${avg_win:+.2f}    Avg Loss: ${avg_loss:+.2f}")
    print(f"  Profit Factor:  {pf:.2f}")
    print(f"  Total P&L:      ${total_pnl:+.2f}")
    print(f"  Total Return:   {total_ret:+.1f}%  (${account:,.0f} → ${acct:,.0f})")
    print(f"  Max Drawdown:   -{max_dd:.1f}%")
    print(f"{'─'*62}")

    best_syms = (df.groupby('Symbol')['P&L ($)'].sum()
                   .sort_values(ascending=False).head(5))
    print(f"  Best symbols by total P&L:")
    for sym, pnl in best_syms.items():
        sym_trades = df[df['Symbol'] == sym]
        sym_wr = len(sym_trades[sym_trades['P&L ($)'] > 0]) / len(sym_trades) * 100
        print(f"    {sym:<8}  ${pnl:+.2f}  ({sym_wr:.0f}% WR, {len(sym_trades)} trades)")

    print(f"\n  Exit reasons:")
    for reason, cnt in df['Exit Reason'].value_counts().items():
        pnl_r = df[df['Exit Reason'] == reason]['P&L ($)'].sum()
        print(f"    {reason:<24}  {cnt:>3} trades  ${pnl_r:+.2f}")

    print(f"{'─'*62}")
    print(f"   Synthetic backtest — IV proxy = HV, vol surface approximated.")
    print(f"    No historical GEX or real fills. Use as signal, not exact P&L.")
    print(f"{'='*62}\n")

    print("  RECENT TRADES (last 10):")
    show_cols = ['Scan Date','Symbol','Direction','Entry Spot ($)','Entry Prem ($)',
                 'Cost ($)','P&L ($)','P&L (%)','Exit Reason','Score']
    print(df[show_cols].tail(10).to_string(index=False))

    # ── Excel export ──────────────────────────────────────────────────────────
    fname = f"scanner_backtest_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    with pd.ExcelWriter(fname, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='All Trades')

        summary = pd.DataFrame([{
            'Metric': k, 'Value': v
        } for k, v in {
            'Total Trades':        len(df),
            'Win Rate (%)':        round(win_rate, 1),
            'Profit Factor':       round(pf, 2),
            'Total P&L ($)':       round(total_pnl, 2),
            'Total Return (%)':    round(total_ret, 1),
            'Start Balance':       account,
            'End Balance':         round(acct, 2),
            'Max Drawdown (%)':    round(max_dd, 1),
            'Avg Win ($)':         round(avg_win, 2),
            'Avg Loss ($)':        round(avg_loss, 2),
            'Skipped (afford)':    skipped_afford,
            'Skipped (vol pct)':   skipped_volpct,
            'Skipped (vol flat)':  skipped_gate,
            'Skipped (regime)':    skipped_regime,
        }.items()])
        summary.to_excel(writer, index=False, sheet_name='Summary')

        sym_stats = (df.groupby('Symbol').agg(
            Trades=('P&L ($)', 'count'),
            Total_PnL=('P&L ($)', 'sum'),
            Win_Rate=('P&L ($)', lambda x: (x > 0).mean() * 100),
            Avg_Score=('Score', 'mean'),
            Avg_VRP=('VRP proxy', 'mean'),
        ).sort_values('Total_PnL', ascending=False))
        sym_stats.to_excel(writer, sheet_name='By Symbol')

    print(f"\n  Saved: {fname}")
    return df


if __name__ == '__main__':
    print("1. Full Analysis      (all expiries, Greeks, screeners, margin)")
    print("2. Trade Finder       (find the right call/put to buy or sell now)")
    print("3. Backtest           (replay the model on one stock)")
    print("4. Market Scanner     (scan market for best opportunities)")
    print("5. Scanner Backtest   (backtest scanner signals historically)")
    choice = input("Choose [1/2/3/4/5]: ").strip()

    if choice == '2':
        find_trade()
    elif choice == '3':
        backtest_model()
    elif choice == '4':
        _budget = input("Max premium per contract $ (e.g. 1.10): ").strip()
        _budget = float(_budget) if _budget else 5.00
        print("  Universe options:")
        print("    [Enter]        = S&P500 + Nasdaq-100 + CBOE most-active + retail list (~600 stocks)")
        print("    Custom tickers = e.g.  AAPL,TSLA,SOFI,ONDS")
        _watch = input("  Your choice: ").strip()
        _wlist = [t.strip().upper() for t in _watch.split(',') if t.strip()] if _watch else None
        market_scanner(budget=_budget, watchlist=_wlist)
    elif choice == '5':
        print("\n  Scanner Backtest Settings  (v2)")
        print("  ─────────────────────────────────────────────")
        _watch = input("  Watchlist (e.g. AAPL,TSLA,SOFI or Enter for default 70 stocks): ").strip()
        _wlist = [t.strip().upper() for t in _watch.split(',') if t.strip()] if _watch else None
        _lb    = input("  Lookback days [252]: ").strip()
        _lb    = int(_lb) if _lb else 252
        _hold  = input("  Holding days per trade [14]: ").strip()
        _hold  = int(_hold) if _hold else 14
        _freq  = input("  Scan frequency (every N trading days) [14]: ").strip()
        _freq  = int(_freq) if _freq else 14
        _topn  = input("  Top N stocks to trade per scan [3]: ").strip()
        _topn  = int(_topn) if _topn else 3
        _acct  = input("  Starting account size [$190]: ").strip()
        _acct  = float(_acct) if _acct else 190.0
        _bud   = input("  Max cost per contract $ [5.00]: ").strip()
        _bud   = float(_bud) if _bud else 5.00
        _tp    = input("  Take profit % [50]: ").strip()
        _tp    = float(_tp) / 100 if _tp else 0.50
        _sl    = input("  Stop loss % [60]: ").strip()
        _sl    = float(_sl) / 100 if _sl else 0.60
        backtest_scanner(
            watchlist=_wlist, lookback_days=_lb, holding_days=_hold,
            scan_freq=_freq, top_n=_topn, account=_acct, budget=_bud,
            take_profit=_tp, stop_loss=_sl,
        )
    else:
        main()
